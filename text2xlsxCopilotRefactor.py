import os
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from openpyxl.worksheet.hyperlink import Hyperlink

# 设置要读取的文件夹路径
folder_path = r'C:\mobile pic'  # 修改为你的txt文件夹路径
output_xlsx = r'C:\Users\SNOW\Desktop\test.xlsx' # 修改为你想保存的xlsx路径

wb = Workbook()
ws = wb.active
ws.append([
    '文件夹绝对路径',
    '图片绝对路径',
    '图片文件超链接',
    'TXT绝对路径',
    'TXT内容',
    '清洗后的数据',
    '提示词类型'
])

def detect_types(line, cleaned):
    types = []
    lower_line = line.lower()
    # R18相关词汇 
    if any(word in lower_line for word in [
        'censor', 'nipple', 'pussy', 'penis', 'hetero', 'sex', 'anus'
    ]):
        types.append('R18')
    # boy类型
    if any(boy_word in lower_line for boy_word in ['1boy', '2boys', 'multiple_boys']):
        types.append('boy')
    # no_human类型
    if 'no_human' in lower_line:
        types.append('no_human')
    return ','.join(types)

def clean_tags(line):
    tags = [tag.strip() for tag in line.strip().split(',')]
    # 去掉含有censor的tag
    cleaned_tags = [tag for tag in tags if 'censor' not in tag.lower()]
    # 检查是否含有敏感词
    has_sensitive = any(
        any(word in tag.lower() for word in [
            'censor', 'nipple', 'pussy', 'penis', 'hetero', 'sex', 'anus'
        ])
        for tag in tags
    )
    if has_sensitive:
        cleaned_tags.append('uncensored')
    cleaned_line = ', '.join([tag for tag in cleaned_tags if tag])
    return cleaned_line, has_sensitive


for root, dirs, files in os.walk(folder_path):
    images = [f for f in files if f.lower().endswith(('.png', '.jpg', '.jpeg', '.bmp', '.gif', '.webp'))]
    txts = [f for f in files if f.lower().endswith('.txt')]
    for img in images:
        img_name, _ = os.path.splitext(img)
        txt_file = img_name + '.txt'
        if txt_file in txts:
            img_path = os.path.join(root, img)
            txt_path = os.path.join(root, txt_file)
            with open(txt_path, 'r', encoding='utf-8') as f:
                for line in f:
                    cleaned, _ = clean_tags(line)
                    prompt_type = detect_types(line, cleaned)
                    # 添加图片超链接
                    img_abs_path = os.path.abspath(img_path)
                    hyperlink = f'=HYPERLINK("{img_abs_path}", "打开图片")'
                    ws.append([
                        os.path.abspath(root),
                        img_abs_path,
                        hyperlink,
                        os.path.abspath(txt_path),
                        line.strip(),
                        cleaned,
                        prompt_type
                    ])

# 设置超链接列为蓝色字体
for row in ws.iter_rows(min_row=2, min_col=3, max_col=3):
    for cell in row:
        cell.font = Font(color="0000FF", underline="single")

wb.save(output_xlsx)
print('合并完成，保存为:', output_xlsx)