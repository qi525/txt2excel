import os
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from openpyxl.worksheet.hyperlink import Hyperlink
import datetime

# 通过input询问用户要处理的文件夹路径
folder_path = input("请输入您要处理的文件夹路径: ")

# 验证路径是否存在
if not os.path.isdir(folder_path):
    print(f"错误: 您输入的路径 '{folder_path}' 不是一个有效的文件夹。程序将退出。")
    exit() # 直接退出程序

# 定义反推记录子文件夹路径
output_folder = os.path.join(folder_path, "反推记录")

# 检查并创建反推记录文件夹
if not os.path.exists(output_folder):
    os.makedirs(output_folder)
    print(f"已创建输出文件夹: {output_folder}")

# 根据您的要求，log和xlsx文件需要自动打开并方便检查结果
# 为了方便调试和检查，log文件和xlsx文件将自动命名并打开
current_time = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
output_xlsx = os.path.join(output_folder, f'scan_results_{current_time}.xlsx')
log_file = os.path.join(output_folder, f'scan_log_{current_time}.txt') # log文件路径

# 创建一个日志文件
def write_log(message):
    with open(log_file, 'a', encoding='utf-8') as f:
        f.write(f"{datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')} - {message}\n")

wb = Workbook()
ws = wb.active
ws.append([
    '文件夹绝对路径',
    '文件绝对路径',
    '文件超链接',
    'TXT绝对路径',
    'TXT内容',
    '清洗后的数据',
    '提示词类型',
    '是否找到匹配TXT'
])

def detect_types(line, cleaned):
    types = []
    lower_line = line.lower()
    # R18相关词汇 
    if any(word in lower_line for word in [
        'censor', 'nipple', 'pussy', 'penis', 'hetero', 'sex', 'anus'
    ]):
        types.append('R18')
    # boy类型
    if any(boy_word in lower_line for boy_word in ['1boy', '2boys', 'multiple_boys']):
        types.append('boy')
    # no_human类型
    if 'no_human' in lower_line:
        types.append('no_human')
    return ','.join(types)

def clean_tags(line):
    tags = [tag.strip() for tag in line.strip().split(',')]
    # 去掉含有censor的tag
    cleaned_tags = [tag for tag in tags if 'censor' not in tag.lower()]
    # 检查是否含有敏感词
    has_sensitive = any(
        any(word in tag.lower() for word in [
            'censor', 'nipple', 'pussy', 'penis', 'hetero', 'sex', 'anus'
        ])
        for tag in tags
    )
    if has_sensitive:
        cleaned_tags.append('uncensored')
    cleaned_line = ', '.join([tag for tag in cleaned_tags if tag])
    return cleaned_line, has_sensitive

# 遍历所有文件，而不仅仅是图片
for root, dirs, files in os.walk(folder_path):
    # 将当前文件夹下的所有txt文件收集起来，方便查找
    current_txt_files = {os.path.splitext(f)[0].lower(): os.path.join(root, f) for f in files if f.lower().endswith('.txt')}

    for f_name in files:
        # 跳过txt文件本身，因为我们主要关注其他文件类型以及它们是否关联了txt
        if f_name.lower().endswith('.txt'):
            continue

        file_path = os.path.join(root, f_name)
        file_name_without_ext, file_ext = os.path.splitext(f_name)
        
        txt_content = ''
        cleaned_data = ''
        prompt_type = ''
        txt_absolute_path = ''
        found_txt = '否' # 默认没有找到匹配的TXT文件

        # 尝试查找同名的txt文件
        if file_name_without_ext.lower() in current_txt_files:
            txt_file_path = current_txt_files[file_name_without_ext.lower()]
            try:
                with open(txt_file_path, 'r', encoding='utf-8') as f:
                    # 假设一个txt文件只有一行内容，如果有多行，您可以根据需求修改
                    # 这里只读取第一行作为txt内容进行处理，如果txt文件有多个描述，需要调整逻辑
                    for line in f: # 遍历txt文件中的每一行
                        txt_content = line.strip() # 获取原始txt内容
                        cleaned_data, _ = clean_tags(txt_content) # 清洗数据
                        prompt_type = detect_types(txt_content, cleaned_data) # 检测提示词类型
                        txt_absolute_path = os.path.abspath(txt_file_path)
                        found_txt = '是'
                        break # 只处理txt文件中的第一行，如果需要处理所有行，请移除此行
            except Exception as e:
                write_log(f"Error reading TXT file {txt_file_path}: {e}")
                txt_content = f"Error reading TXT: {e}" # 记录读取错误信息
                found_txt = '否 (读取错误)'
        else:
            write_log(f"No matching TXT file found for: {file_path}")

        # 添加文件超链接
        file_abs_path = os.path.abspath(file_path)
        hyperlink = f'=HYPERLINK("{file_abs_path}", "打开文件")'
        
        ws.append([
            os.path.abspath(root),
            file_abs_path,
            hyperlink,
            txt_absolute_path,
            txt_content,
            cleaned_data,
            prompt_type,
            found_txt
        ])

# 设置超链接列为蓝色字体
for row in ws.iter_rows(min_row=2, min_col=3, max_col=3):
    for cell in row:
        cell.font = Font(color="0000FF", underline="single")

wb.save(output_xlsx)
print('合并完成，保存为:', output_xlsx)
write_log(f"Processing complete. Results saved to: {output_xlsx}")
write_log(f"Log saved to: {log_file}")

# 自动打开xlsx和log文件方便检查结果
try:
    os.startfile(output_xlsx)
    print(f"Automatically opened: {output_xlsx}")
    os.startfile(log_file)
    print(f"Automatically opened: {log_file}")
except Exception as e:
    print(f"Could not automatically open files. Please check manually. Error: {e}")
    write_log(f"Failed to auto-open files: {output_xlsx}, {log_file}. Error: {e}")