import os
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.styles import PatternFill

def write_txt_paths_and_content_to_excel(folder_path, excel_file_path):
    """
    将指定文件夹及其子文件夹中所有 .txt 文件的完整路径写入 Excel 表的 A 列，
    将其内容写入 B 列，并将 B 列内容复制到 C 列，根据关键词删除 C 列中的指定词组。
    同时，根据内容为 B 列和 C 列的单元格设置字体颜色。
    尝试匹配同名的图片文件，将路径以超链接形式写入 D 列。
    新增一列 (E 列)，显示 TXT 文件所在的文件夹路径。
    如果内容含有换行符，则删除。

    Args:
        folder_path (str): 包含 .txt 文件的根文件夹路径。
        excel_file_path (str): 输出的 Excel 文件路径（例如: "output.xlsx"）。
    """
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "TXT文件信息"

    # 设置表头，让Excel内容更清晰
    sheet.cell(row=1, column=1, value="TXT文件路径")
    sheet.cell(row=1, column=2, value="原始内容")
    sheet.cell(row=1, column=3, value="清洗后的内容")
    sheet.cell(row=1, column=4, value="对应图片文件")
    sheet.cell(row=1, column=5, value="所在文件夹") # 新增列：E列
    
    row_num = 2  # 从第二行开始写入数据，因为第一行是表头

    # 定义字体颜色样式
    red_font = Font(color="FF0000")
    blue_font = Font(color="0000FF")
    yellow_font = Font(color="FFFF00")
    
    # 定义超链接字体样式
    hyperlink_font = Font(color="0000FF", underline="single")
    
    # 定义未找到图片的背景色
    no_image_fill = PatternFill(start_color="FFFCCB", end_color="FFFCCB", fill_type="solid")

    # 定义常见的图片文件后缀
    image_extensions = [
        ".jpg", ".jpeg", ".png", ".gif", ".bmp", ".webp", ".tiff", ".tif", ".ico", ".svg"
    ]

    txt_files_processed = 0

    try:
        for root, dirs, files in os.walk(folder_path):
            for file in files:
                if file.lower().endswith(".txt"):
                    full_path_txt = os.path.join(root, file)
                    txt_files_processed += 1

                    # 提取TXT文件的主文件名 (不带后缀)
                    txt_basename = os.path.splitext(file)[0]

                    # 将文件路径写入 A 列
                    sheet.cell(row=row_num, column=1, value=full_path_txt)

                    # 读取 .txt 文件内容
                    txt_content = ""
                    try:
                        with open(full_path_txt, 'r', encoding='utf-8') as f:
                            txt_content = f.read()
                    except Exception as e:
                        print(f"警告: 无法读取文件 {full_path_txt}。错误: {e}")
                        txt_content = "读取失败" 

                    # 删除换行符
                    cleaned_content = txt_content.replace('\r\n', '').replace('\n', '').replace('\r', '')

                    # 将处理后的内容写入 B 列
                    b_cell = sheet.cell(row=row_num, column=2, value=cleaned_content)
                    
                    # 为 B 列设置字体颜色
                    lower_cleaned_content = cleaned_content.lower()
                    if "censor" in lower_cleaned_content:
                        b_cell.font = red_font
                    elif "no_humans" in lower_cleaned_content:
                        b_cell.font = blue_font
                    elif "boy" in lower_cleaned_content:
                        b_cell.font = yellow_font

                    # --- C 列处理逻辑 ---
                    content_for_c = cleaned_content
                    keywords_to_remove = ["censor", "mosaic"] 
                    exception_words = ["uncensored"]
                    tags = [tag.strip() for tag in content_for_c.split(',')]
                    
                    filtered_tags = []
                    for tag in tags:
                        is_unwanted = False
                        if tag in exception_words:
                            filtered_tags.append(tag)
                            continue
                        
                        for keyword in keywords_to_remove:
                            if keyword in tag: 
                                is_unwanted = True
                                break
                        
                        if not is_unwanted and tag: 
                            filtered_tags.append(tag)
                    
                    content_for_c = ','.join(filtered_tags)

                    # 将处理后的内容写入 C 列
                    c_cell = sheet.cell(row=row_num, column=3, value=content_for_c)

                    # 为 C 列设置字体颜色
                    if "censor" in content_for_c.lower():
                        c_cell.font = red_font

                    # --- 匹配对应的图片文件 (D列 - 超链接形式) ---
                    matched_image_path = "未找到对应图片"
                    image_found = False
                    
                    try:
                        for entry in os.scandir(root):
                            if entry.is_file():
                                other_file_basename, other_file_ext = os.path.splitext(entry.name)
                                
                                if txt_basename == other_file_basename and other_file_ext.lower() in image_extensions:
                                    matched_image_path = entry.path
                                    image_found = True
                                    break
                    except Exception as e:
                        print(f"警告: 无法扫描目录 {root} 或匹配图片。错误: {e}")
                        matched_image_path = "图片匹配失败"
                        image_found = False

                    d_cell = sheet.cell(row=row_num, column=4)
                    
                    if image_found:
                        d_cell.value = os.path.basename(matched_image_path)
                        d_cell.hyperlink = os.path.abspath(matched_image_path)
                        d_cell.font = hyperlink_font
                    else:
                        d_cell.value = matched_image_path
                        d_cell.fill = no_image_fill

                    # --- 图片文件匹配结束 ---

                    # --- 新增 E 列：所在文件夹路径 ---
                    # root 变量在 os.walk 循环中天然就是当前文件的所在文件夹路径
                    sheet.cell(row=row_num, column=5, value=root)
                    # --- E 列添加结束 ---

                    row_num += 1
        
        if txt_files_processed == 0:
            print(f"未在 '{folder_path}' 及其子文件夹中找到任何 .txt 文件。未生成 Excel 文件。")
            return

        output_dir = os.path.dirname(excel_file_path)
        if output_dir and not os.path.exists(output_dir):
            try:
                os.makedirs(output_dir)
                print(f"已创建输出目录: '{output_dir}'")
            except OSError as e:
                print(f"错误: 无法创建输出目录 '{output_dir}'。请检查权限。错误: {e}")
                return

        workbook.save(excel_file_path)
        print(f"所有 .txt 文件的路径、内容、对应图片路径（超链接形式）和所在文件夹路径已成功写入到 '{excel_file_path}'，并已设置字体颜色。")

    except Exception as e:
        print(f"发生了一个意外错误: {e}")
        print("请检查文件夹路径、文件权限或文件是否被其他程序占用。")



if __name__ == "__main__":
    # 请将 'C:\mobile pic\Pictures' 替换为你实际要扫描的文件夹路径
    folder_to_scan = r"C:\mobile pic\Pictures" 

    script_dir = os.path.dirname(os.path.abspath(__file__))
    cleaned_folder_name = folder_to_scan.replace(":", "").replace("\\", "_").replace("/", "_").strip()
    # 更新文件名，反映新增列
    output_excel_file = os.path.join(script_dir, f"{cleaned_folder_name}_processed_tags_with_hyperlinks_and_folders.xlsx") 

    if os.path.exists(output_excel_file):
        print(f"提示：文件 '{output_excel_file}' 已存在。它将被覆盖。")
    
    if not os.path.isdir(folder_to_scan):
        print(f"错误: 扫描文件夹 '{folder_to_scan}' 不存在或不是一个有效的目录。请检查路径。")
    else:
        write_txt_paths_and_content_to_excel(folder_to_scan, output_excel_file)