import os
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from openpyxl.worksheet.hyperlink import Hyperlink
import datetime
import shutil # 导入shutil模块用于文件复制
import subprocess # 用于跨平台打开文件

# 通过input询问用户要处理的文件夹路径
folder_path = input("请输入您要处理的文件夹路径: ")

# 验证路径是否存在
if not os.path.isdir(folder_path):
    print(f"错误: 您输入的路径 '{folder_path}' 不是一个有效的文件夹。程序将退出。")
    exit() # 直接退出程序

# 获取Python运行文件的目录
python_script_dir = os.path.dirname(os.path.abspath(__file__))
# 定义Python运行目录中的文件路径 (作为主保存位置)
current_time = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
main_output_xlsx = os.path.join(python_script_dir, f'scan_results_{current_time}.xlsx')
main_log_file = os.path.join(python_script_dir, f'scan_log_{current_time}.txt') # log文件路径

# 定义反推记录子文件夹路径 (作为次要保存位置)
output_folder = os.path.join(folder_path, "反推记录")
# 定义目标文件夹中的文件路径
target_output_xlsx = os.path.join(output_folder, f'scan_results_{current_time}.xlsx')
target_log_file = os.path.join(output_folder, f'scan_log_{current_time}.txt')

# 创建一个日志文件函数，现在它将主要写入到运行目录的log文件
def write_log(message):
    with open(main_log_file, 'a', encoding='utf-8') as f:
        f.write(f"{datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')} - {message}\n")

write_log("程序开始运行...") # 记录程序启动

wb = Workbook()
ws = wb.active
ws.title = "已匹配TXT文件" # 设置第一个工作表的标题

# 主工作表的列头
ws.append([
    '文件夹绝对路径',
    '文件绝对路径',
    '文件超链接',
    '文件后缀', # 新增列：文件后缀
    'TXT绝对路径',
    'TXT内容',
    '清洗后的数据',
    '提示词类型',
    '是否找到匹配TXT'
])

# 创建第二个工作表用于存放未找到TXT的文件
ws_no_txt = wb.create_sheet("未匹配TXT文件", 1) # 在第一个工作表后创建
ws_no_txt.append([
    '文件夹绝对路径',
    '文件绝对路径',
    '文件超链接',
    '文件后缀', # 新增列：文件后缀
    '是否找到匹配TXT'
])

def detect_types(line, cleaned):
    types = []
    lower_line = line.lower()
    # R18相关词汇 
    if any(word in lower_line for word in [
        'censor', 'nipple', 'pussy', 'penis', 'hetero', 'sex', 'anus'
    ]):
        types.append('R18')
    # boy类型
    if any(boy_word in lower_line for boy_word in ['1boy', '2boys', 'multiple_boys']):
        types.append('boy')
    # no_human类型
    if 'no_human' in lower_line:
        types.append('no_human')
    return ','.join(types)

def clean_tags(line):
    tags = [tag.strip() for tag in line.strip().split(',')]
    # 去掉含有censor的tag
    cleaned_tags = [tag for tag in tags if 'censor' not in tag.lower()]
    # 检查是否含有敏感词
    has_sensitive = any(
        any(word in tag.lower() for word in [
            'censor', 'nipple', 'pussy', 'penis', 'hetero', 'sex', 'anus'
        ])
        for tag in tags
    )
    if has_sensitive:
        cleaned_tags.append('uncensored')
    cleaned_line = ', '.join([tag for tag in cleaned_tags if tag])
    return cleaned_line, has_sensitive

# 遍历所有文件，而不仅仅是图片
for root, dirs, files in os.walk(folder_path):
    # 将当前文件夹下的所有txt文件收集起来，方便查找
    current_txt_files = {os.path.splitext(f)[0].lower(): os.path.join(root, f) for f in files if f.lower().endswith('.txt')}

    for f_name in files:
        # 跳过txt文件本身，因为我们主要关注其他文件类型以及它们是否关联了txt
        if f_name.lower().endswith('.txt'):
            continue

        file_path = os.path.join(root, f_name)
        file_name_without_ext, file_ext = os.path.splitext(f_name)
        file_ext = file_ext.lower() # 统一小写文件后缀

        txt_content = ''
        cleaned_data = ''
        prompt_type = ''
        txt_absolute_path = ''
        found_txt = '否' # 默认没有找到匹配的TXT文件

        # 添加文件超链接
        file_abs_path = os.path.abspath(file_path)
        hyperlink_formula = f'=HYPERLINK("{file_abs_path}", "打开文件")'
        
        # 尝试查找同名的txt文件
        if file_name_without_ext.lower() in current_txt_files:
            txt_file_path = current_txt_files[file_name_without_ext.lower()]
            try:
                with open(txt_file_path, 'r', encoding='utf-8') as f:
                    # 假设一个txt文件只有一行内容，如果有多行，您可以根据需求修改
                    # 这里只读取第一行作为txt内容进行处理，如果txt文件有多个描述，需要调整逻辑
                    for line in f: # 遍历txt文件中的每一行
                        txt_content = line.strip() # 获取原始txt内容
                        cleaned_data, _ = clean_tags(txt_content) # 清洗数据
                        prompt_type = detect_types(txt_content, cleaned_data) # 检测提示词类型
                        txt_absolute_path = os.path.abspath(txt_file_path)
                        found_txt = '是'
                        break # 只处理txt文件中的第一行，如果需要处理所有行，请移除此行
            except Exception as e:
                write_log(f"Error reading TXT file {txt_file_path}: {e}")
                txt_content = f"Error reading TXT: {e}" # 记录读取错误信息
                found_txt = '否 (读取错误)'
        else:
            write_log(f"No matching TXT file found for: {file_path}")

        # 根据是否找到TXT文件，将数据写入不同的工作表
        if found_txt == '是':
            ws.append([
                os.path.abspath(root),
                file_abs_path,
                hyperlink_formula,
                file_ext, # 新增列数据
                txt_absolute_path,
                txt_content,
                cleaned_data,
                prompt_type,
                found_txt
            ])
        else:
            ws_no_txt.append([
                os.path.abspath(root),
                file_abs_path,
                hyperlink_formula,
                file_ext, # 新增列数据
                found_txt
            ])

# 设置主工作表超链接列为蓝色字体
# 注意：超链接列现在是第4列（索引3）
for row in ws.iter_rows(min_row=2, min_col=3, max_col=3):
    for cell in row:
        cell.font = Font(color="0000FF", underline="single")

# 设置未匹配TXT工作表超链接列为蓝色字体
# 注意：超链接列现在是第3列（索引2）
for row in ws_no_txt.iter_rows(min_row=2, min_col=3, max_col=3):
    for cell in row:
        cell.font = Font(color="0000FF", underline="single")

# ====================================================================
# 优先保存到Python运行目录
# ====================================================================
try:
    wb.save(main_output_xlsx)
    print('合并完成，已保存至Python运行目录:', main_output_xlsx)
    write_log(f"Processing complete. Results saved to Python script directory: {main_output_xlsx}")
    write_log(f"Log saved to Python script directory: {main_log_file}")
except Exception as e:
    print(f"错误: 无法保存结果到Python运行目录 {main_output_xlsx}。错误: {e}")
    write_log(f"Error: Could not save results to Python script directory {main_output_xlsx}. Error: {e}")
    print("程序将退出，请检查Python运行目录的写入权限。")
    exit() # 如果连运行目录都无法写入，则直接退出

# ====================================================================
# 尝试保存一份到目标文件夹
# ====================================================================
try:
    # 检查并创建反推记录文件夹（如果之前没有创建的话，确保在尝试写入前存在）
    if not os.path.exists(output_folder):
        os.makedirs(output_folder)
        print(f"已创建输出文件夹: {output_folder}")

    # 复制已经生成在Python运行目录的XLSX文件到目标文件夹
    shutil.copy2(main_output_xlsx, target_output_xlsx)
    print(f"一份副本已保存至目标文件夹: {target_output_xlsx}")
    write_log(f"A copy of XLSX saved to target folder: {target_output_xlsx}")

    # 复制已经生成在Python运行目录的log文件到目标文件夹
    shutil.copy2(main_log_file, target_log_file)
    print(f"一份log副本已保存至目标文件夹: {target_log_file}")
    write_log(f"A copy of log saved to target folder: {target_log_file}")

except Exception as e:
    print(f"警告: 无法保存文件到目标文件夹 '{output_folder}'。错误: {e}")
    print("请检查目标文件夹的写入权限。Python运行目录已保存一份结果。")
    write_log(f"Warning: Could not save files to target folder '{output_folder}'. Error: {e}")
    write_log("Python script directory already contains a copy of results.")

# 自动打开Python运行目录的xlsx和log文件方便检查结果
try:
    # 使用subprocess.Popen以兼容更多系统
    if os.name == 'nt': # Windows
        os.startfile(main_output_xlsx)
        os.startfile(main_log_file)
    elif os.uname().sysname == 'Darwin': # macOS
        subprocess.Popen(['open', main_output_xlsx])
        subprocess.Popen(['open', main_log_file])
    else: # Linux/Unix
        subprocess.Popen(['xdg-open', main_output_xlsx])
        subprocess.Popen(['xdg-open', main_log_file])
    
    print(f"Automatically opened: {main_output_xlsx}")
    print(f"Automatically opened: {main_log_file}")
except Exception as e:
    print(f"Could not automatically open files from Python script directory. Please check manually. Error: {e}")
    write_log(f"Failed to auto-open files from Python script directory: {main_output_xlsx}, {main_log_file}. Error: {e}")