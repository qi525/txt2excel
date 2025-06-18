# utils/excel_utils.py
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.worksheet.worksheet import Worksheet # 导入Worksheet类型用于类型提示
from openpyxl.utils import get_column_letter # 尽管不直接使用，但可能在其他地方有用

def create_main_workbook() -> Workbook:
    """
    创建一个新的Excel工作簿。
    """
    return Workbook()

def setup_excel_sheets() -> tuple[Workbook, Worksheet, Worksheet, Worksheet]:
    """
    设置Excel工作簿，创建所需的三个工作表并设置列头。
    返回工作簿对象和三个工作表对象。
    """
    wb = Workbook()
    ws_matched = wb.active
    ws_matched.title = "已匹配TXT文件"

    ws_matched.append([
        '文件夹绝对路径',
        '文件绝对路径',
        '文件超链接',
        '文件后缀',
        'TXT绝对路径',
        'TXT内容',
        '清洗后的数据',
        '清洗后的数据字数',
        '提示词类型',
        '是否找到匹配TXT'
    ])

    ws_no_txt = wb.create_sheet("未匹配TXT文件", 1)
    ws_no_txt.append([
        '文件夹绝对路径',
        '文件绝对路径',
        '文件超链接',
        '文件后缀',
        '是否找到匹配TXT'
    ])

    ws_tag_frequency = wb.create_sheet("Tag 词频统计", 2)
    ws_tag_frequency.append([
        'Tag',
        '出现次数'
    ])
    return wb, ws_matched, ws_no_txt, ws_tag_frequency

def apply_hyperlink_style(ws: Worksheet, col_index: int):
    """
    为指定工作表的超链接列（从第二行开始）应用蓝色字体和下划线样式。
    col_index 是基于1的列索引。
    """
    for row in ws.iter_rows(min_row=2, min_col=col_index, max_col=col_index):
        for cell in row:
            cell.font = Font(color="0000FF", underline="single")