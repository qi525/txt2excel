# services/history_manager.py
import datetime
import sys # 依然保留sys
from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

from services.log_manager import LogManager

class HistoryManager:
    """
    负责程序运行历史记录的更新和管理。
    """
    def __init__(self, history_file_path: Path, log_manager: LogManager):
        self.history_file_path = history_file_path
        self.log_manager = log_manager
        self._ensure_history_file_exists()

    def _ensure_history_file_exists(self):
        """
        确保历史记录文件存在并具有正确的列头。
        """
        if not self.history_file_path.exists():
            try:
                history_wb = Workbook()
                history_ws = history_wb.active
                history_ws.title = "扫描历史记录"
                history_ws.append([
                    '运行时间',
                    '分析目录',
                    '总文件量',
                    '成功匹配TXT数量',
                    '失败匹配TXT数量',
                    'Log文件绝对路径',
                    'Log文件超链接',
                    '结果XLSX文件绝对路径',
                    '结果XLSX文件超链接'
                ])
                history_wb.save(str(self.history_file_path))
                self.log_manager.write_log(f"Created new history file: {self.history_file_path}")
            except Exception as e:
                self.log_manager.write_log(f"Error: Could not create new history file {self.history_file_path}. Error: {e}")
                print(f"错误: 无法创建新的历史记录文件 {self.history_file_path}。错误: {e}")

    def update_history(self, folder_path: Path, total_scanned: int,
                       found_count: int, not_found_count: int,
                       xlsx_file_path: Path, log_file_path: Path):
        """
        更新历史记录文件，添加新的扫描结果。
        """
        try:
            history_wb = load_workbook(str(self.history_file_path))
            history_ws = history_wb.active

            expected_headers = ['运行时间', '分析目录', '总文件量', '成功匹配TXT数量', '失败匹配TXT数量',
                                'Log文件绝对路径', 'Log文件超链接', '结果XLSX文件绝对路径', '结果XLSX文件超链接']
            current_headers = [cell.value for cell in history_ws[1]]
            if current_headers != expected_headers:
                self.log_manager.write_log(f"Warning: History file {self.history_file_path} might have outdated headers. Appending new row with full headers schema.")
                print("警告：检测到旧版历史记录文件格式，将追加新行。建议手动检查或删除旧文件以生成完整新格式。")

            current_run_time = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

            # --- 核心改动：直接使用Path对象的字符串表示作为超链接目标 ---
            # 与您旧版单文件脚本逻辑一致，Windows Excel可以直接处理这种路径
            log_link_path_for_excel = str(log_file_path)
            xlsx_link_path_for_excel = str(xlsx_file_path)
            # --- 结束核心改动 ---

            log_hyperlink_formula = f'=HYPERLINK("{log_link_path_for_excel}", "打开Log")'
            xlsx_hyperlink_formula = f'=HYPERLINK("{xlsx_link_path_for_excel}", "打开结果XLSX")'

            history_ws.append([
                current_run_time,
                str(folder_path),
                total_scanned,
                found_count,
                not_found_count,
                str(log_file_path),
                log_hyperlink_formula,
                str(xlsx_file_path),
                xlsx_hyperlink_formula
            ])

            new_row_idx = history_ws.max_row
            history_ws.cell(row=new_row_idx, column=7).font = Font(color="0000FF", underline="single")
            history_ws.cell(row=new_row_idx, column=9).font = Font(color="0000FF", underline="single")

            history_wb.save(str(self.history_file_path))
            print(f"历史记录已更新到: {self.history_file_path}")
            self.log_manager.write_log(f"History record updated: {self.history_file_path}")
        except Exception as e:
            self.log_manager.write_log(f"Error: Could not update history file {self.history_file_path}. Error: {e}")
            print(f"错误: 无法更新历史记录文件 {self.history_file_path}。错误: {e}")