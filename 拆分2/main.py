# InterrogateText2Xlsx7.0_main.py

import os
import sys
import datetime
import shutil
import subprocess
import re
import time
from pathlib import Path
from collections import defaultdict
from typing import Tuple, Dict, Optional, Set, List, Any
import hashlib

# 注入超链接样式和列宽大小
#from excel_utils import set_hyperlink_and_style,set_fixed_column_widths

# openpyxl 相关的导入
from openpyxl import Workbook, load_workbook
#from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill
from openpyxl.styles.colors import Color
from openpyxl.worksheet.worksheet import Worksheet

# 从 my_logger 导入 setup_logger 和 _error_log_file_path，以及 normalize_drive_letter
from my_logger import setup_logger, _error_log_file_path, normalize_drive_letter

# 从 loguru 导入 logger
from loguru import logger


# --- 主要改动点 START ---
# 从 excel_utils 导入 set_hyperlink_and_style, set_fixed_column_widths 和 FIXED_COLUMN_WIDTH
#from excel_utils import set_hyperlink_and_style, set_fixed_column_widths, FIXED_COLUMN_WIDTH
# 从 execution_history 导入 HistoryManager
#from execution_history import HistoryManager
# --- 主要改动点 END ---

# 从 utils 导入除了 normalize_drive_letter 的其他常量和函数
from utils import (
    generate_folder_prefix,
    HistoryManager,
    create_directory_if_not_exists,
    copy_file,
    create_main_workbook,
    setup_excel_sheets,
    scan_files_and_extract_data,
    read_batch_paths,
    HISTORY_FOLDER_NAME,
    HISTORY_EXCEL_NAME,
    OUTPUT_FOLDER_NAME,
    CACHE_FOLDER_NAME,
    MAX_SAVE_RETRIES,
    RETRY_DELAY_SECONDS,
    FIXED_COLUMN_WIDTH
)

# --- BEGIN MODIFICATION 1/X (解决 NameError) ---
# 新增 Loguru 和 my_logger 的导入
# [MODIFIED] 添加 get_error_log_file_path 到导入列表
from my_logger import setup_logger, get_error_log_file_path 
from loguru import logger
from my_logger import normalize_drive_letter # 导入 normalize_drive_letter

# 将 script_dir 和 log_output_dir 移到全局作用域
script_dir = Path(os.path.dirname(os.path.abspath(__file__)))
log_output_dir = script_dir / "logs"

# 初始化 Loguru 日志系统
setup_logger(log_output_dir)
# --- END MODIFICATION 1/X ---


# --- Configuration ---
# ... (这一部分在你的原始main.py中可能存在，请保留原有内容) ...
# HYPERLINK_FONT = Font(color="0000FF", underline="single") # 如果Font不是从openpyxl导入并全局定义，则可能需要在这里定义


# ... (其他导入) ...
import time # 确保导入了time模块

# ... (其他函数定义) ...

def open_output_files_automatically(file_paths: List[Path], logger_obj):
    """
    根据用户设置自动打开生成的输出文件（Excel和Log文件）。
    Args:
        file_paths (List[Path]): 包含要打开的文件路径的列表。
        logger_obj (logger): Loguru logger 实例。
    """
    if os.getenv("DISABLE_AUTO_OPEN", "0") == "1":
        logger_obj.info("已禁用自动打开文件功能。")
        return

    # 定义延迟时间 (秒)
    OPEN_FILE_DELAY_SECONDS = 2 # 你建议的2秒延迟

    for file_path in file_paths:
        # --- 主要修改点 START ---
        # 在尝试打开文件前增加延迟
        logger_obj.debug(f"尝试打开文件 '{normalize_drive_letter(str(file_path))}' 前，等待 {OPEN_FILE_DELAY_SECONDS} 秒。")
        time.sleep(OPEN_FILE_DELAY_SECONDS)
        # --- 主要修改点 END ---

        actual_path_to_open = file_path
        # 特殊处理 Loguru 压缩后的日志文件（根据之前的讨论，这部分逻辑应该已经存在）
        if file_path.suffix == '.txt' and not file_path.exists():
            zip_path = file_path.with_suffix('.zip')
            if zip_path.exists():
                actual_path_to_open = zip_path
                logger_obj.info(f"日志文件 '{normalize_drive_letter(str(file_path))}' 不存在，尝试打开压缩文件: {normalize_drive_letter(str(zip_path))}")

        if not actual_path_to_open.exists():
            logger_obj.warning(f"警告: 无法自动打开文件 '{normalize_drive_letter(str(actual_path_to_open))}'，因为文件不存在。")
            print(f"警告: 无法自动打开文件 '{actual_path_to_open}'，因为文件不存在。")
            continue

        try:
            normalized_path = normalize_drive_letter(str(actual_path_to_open))
            logger_obj.info(f"自动打开: {normalized_path}")
            print(f"自动打开: {actual_path_to_open}")
            if sys.platform == "win32":
                os.startfile(normalized_path)
            elif sys.platform == "darwin": # macOS
                subprocess.run(["open", normalized_path], check=True)
            else: # Linux
                subprocess.run(["xdg-open", normalized_path], check=True)
        except Exception as e:
            logger_obj.error(f"自动打开文件 '{normalize_drive_letter(str(actual_path_to_open))}' 失败: {e}")
            print(f"自动打开文件 '{actual_path_to_open}' 失败: {e}")

# ... (main 函数的其余部分) ...

# ... (保留原有导入，包括 from loguru import logger 和 from my_logger import setup_logger, get_error_log_file_path) ...

# main.py

# ... (imports and global definitions) ...

def main():
    script_dir = Path(os.path.dirname(os.path.abspath(__file__)))
    log_output_dir = script_dir / "logs"
    history_folder_path = script_dir / HISTORY_FOLDER_NAME
    output_base_dir = script_dir / OUTPUT_FOLDER_NAME
    final_history_excel_path = history_folder_path / HISTORY_EXCEL_NAME
    cache_folder_path = script_dir / CACHE_FOLDER_NAME

    # 新增：用于存储错误和警告日志文件路径
    error_warning_log_file_path = get_error_log_file_path()

    logger.info(f"程序启动. 脚本目录: {normalize_drive_letter(str(script_dir))}")
    logger.info(f"日志文件将保存到: {normalize_drive_letter(str(log_output_dir))}")
    logger.info(f"错误和警告日志将写入到独立的日志文件 (仅当出现警告或错误时创建)。")

    if not create_directory_if_not_exists(history_folder_path, logger):
        logger.critical("致命错误: 无法创建历史记录文件夹，程序退出。")
        sys.exit(1)

    if not create_directory_if_not_exists(cache_folder_path, logger):
        logger.critical("致命错误: 无法创建缓存文件夹，程序退出。")
        sys.exit(1)

    if not create_directory_if_not_exists(output_base_dir, logger):
        logger.critical("致命错误: 无法创建输出文件夹 (反推记录)，程序退出。")
        sys.exit(1)

    history_manager = HistoryManager(final_history_excel_path, logger)

    batch_file_path = script_dir / "batchPath.txt"
    folders_to_scan = read_batch_paths(batch_file_path, logger)

    logger.info(f"在 '{normalize_drive_letter(str(batch_file_path))}' 中检测到 {len(folders_to_scan)} 条有效地址。")

    if not folders_to_scan:
        logger.info("没有找到要扫描的文件夹路径，程序终止。")
        print("没有找到要扫描的文件夹路径，程序终止。")
        logger.close()
        sys.exit(0)

    current_folder_log_sink_id: Optional[int] = None

    # --- START MODIFICATION ---
    # Initialize final_files_to_open_at_end here, at the top level of main()
    final_files_to_open_at_end = []
    # --- END MODIFICATION ---

    for folder_path in folders_to_scan:
        # ... (rest of the loop content, no changes needed here) ...
        logger.info(f"\n--- 开始处理文件夹: {normalize_drive_letter(str(folder_path))} ---")
        print(f"\n--- 开始处理文件夹: {folder_path} ---")

        scan_timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        folder_prefix = generate_folder_prefix(folder_path)

        current_scan_log_file = output_base_dir / f"{folder_prefix}_scan_log_{scan_timestamp}.txt"
        current_excel_file = output_base_dir / f"{folder_prefix}_scan_results_{scan_timestamp}.xlsx"

        fallback_excel_file = log_output_dir / f"FALLBACK_{folder_prefix}_scan_results_{scan_timestamp}.xlsx"

        try:
            current_folder_log_sink_id = logger.add(
                str(current_scan_log_file),
                level="INFO",
                rotation="5 MB",
                compression="zip",
                enqueue=True,
                encoding="utf-8",
                format="{time:YYYY-MM-DD HH:mm:ss.SSS} | {level: <8} | {message}"
            )
            logger.info(f"针对当前文件夹的扫描日志将写入: {normalize_drive_letter(str(current_scan_log_file))}")
        except Exception as e:
            logger.error(f"无法为文件夹 '{normalize_drive_letter(str(folder_path))}' 添加扫描日志文件 sink: {e}")
            current_folder_log_sink_id = None

        logger.info(f"开始扫描 {normalize_drive_letter(str(folder_path))}")

        try:
            wb = create_main_workbook()
            ws_matched, ws_no_txt, ws_tag_frequency = setup_excel_sheets(wb)

            total_files, found_txt_count, not_found_txt_count, tag_counts = scan_files_and_extract_data(
                folder_path, ws_matched, ws_no_txt, logger
            )

            sorted_tags = sorted(tag_counts.items(), key=lambda item: item[1], reverse=True)
            for tag, count in sorted_tags:
                ws_tag_frequency.append([tag, count])

            for worksheet in [ws_matched, ws_no_txt, ws_tag_frequency]:
                from utils import set_fixed_column_widths
                set_fixed_column_widths(worksheet, FIXED_COLUMN_WIDTH, logger)

            save_successful = False
            actual_result_file_path = Path("N/A_SAVE_FAILED")

            for attempt in range(MAX_SAVE_RETRIES):
                try:
                    wb.save(str(current_excel_file))
                    logger.info(f"扫描结果已保存到: {normalize_drive_letter(str(current_excel_file))} (尝试 {attempt + 1}/{MAX_SAVE_RETRIES})")
                    print(f"扫描结果已保存到: {current_excel_file}")
                    actual_result_file_path = current_excel_file
                    save_successful = True
                    break
                except PermissionError as e:
                    logger.warning(
                        f"警告: 无法将扫描结果保存到 '{normalize_drive_letter(str(current_excel_file))}'，原因: 权限拒绝！请确保该文件未被其他程序（如Excel）打开。尝试 {attempt + 1}/{MAX_SAVE_RETRIES}。错误: {e}"
                    )
                    print(f"警告: 无法保存结果到 {current_excel_file}！原因: 权限拒绝。尝试 {attempt + 1}/{MAX_SAVE_RETRIES}。等待 {RETRY_DELAY_SECONDS} 秒后重试...")
                    time.sleep(RETRY_DELAY_SECONDS)
                except Exception as e:
                    logger.error(
                        f"错误: 将扫描结果保存到 '{normalize_drive_letter(str(current_excel_file))}' 失败: {e} (尝试 {attempt + 1}/{MAX_SAVE_RETRIES})",
                        level="ERROR"
                    )
                    print(f"错误: 无法保存结果到 {current_excel_file}. 错误: {e}")
                    break

            if not save_successful:
                logger.critical(
                    f"严重警告: 经过 {MAX_SAVE_RETRIES} 次尝试后，仍无法将扫描结果保存到 '{normalize_drive_letter(str(current_excel_file))}'。尝试保存到备用位置。"
                )
                print(f"严重警告: 经过 {MAX_SAVE_RETRIES} 次尝试后，仍无法将扫描结果保存到 {current_excel_file}。尝试保存到备用位置...")
                try:
                    wb.save(str(fallback_excel_file))
                    logger.warning(f"成功将扫描结果保存到备用位置: {normalize_drive_letter(str(fallback_excel_file))}")
                    print(f"成功将扫描结果保存到备用位置: {fallback_excel_file}")
                    actual_result_file_path = fallback_excel_file
                except Exception as fallback_e:
                    logger.critical(
                        f"致命错误: 尝试将扫描结果保存到备用位置 '{normalize_drive_letter(str(fallback_excel_file))}' 也失败了！错误: {fallback_e}"
                    )
                    print(f"致命错误: 无法保存结果到任何位置！错误: {fallback_e}")
                    actual_result_file_path = Path("N/A_SAVE_FAILED")

            history_manager.add_history_entry(
                folder_path, total_files, found_txt_count, not_found_txt_count, actual_result_file_path, current_scan_log_file
            )
            logger.info(f"本次扫描历史记录已成功添加至内存。")

            files_to_open_this_scan = [current_scan_log_file]
            if actual_result_file_path.exists():
                files_to_open_this_scan.append(actual_result_file_path)

            open_output_files_automatically(files_to_open_this_scan, logger)

        except Exception as e:
            logger.error(f"处理文件夹 {normalize_drive_letter(str(folder_path))} 时发生错误: {e}")
            print(f"处理文件夹 {folder_path} 时发生错误。错误: {e}")
        finally:
            if current_folder_log_sink_id is not None:
                logger.remove(current_folder_log_sink_id)
                current_folder_log_sink_id = None
            logger.info(f"--- 完成处理文件夹: {normalize_drive_letter(str(folder_path))} ---\n")
            print(f"--- 完成处理文件夹: {folder_path} ---\n")

    logger.info(f"所有扫描任务完成，开始将历史记录保存到最终的Excel文件: {normalize_drive_letter(str(final_history_excel_path))}")
    logger.info(f"准备保存 {len(history_manager.history_data)} 条历史记录到Excel。")
    logger.info(f"最初在 '{normalize_drive_letter(str(batch_file_path))}' 中检测到 {len(folders_to_scan)} 条有效地址。")

    save_history_success = history_manager.save_history_to_excel()

    # The list is now initialized at the beginning of main()
    # final_files_to_open_at_end = [] # This line is moved/removed

    if save_history_success:
        logger.info("历史记录已成功保存到Excel。")
        if create_directory_if_not_exists(cache_folder_path, logger):
            current_timestamp_for_cache = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            cached_history_file_name = f"scan_history_cached_{current_timestamp_for_cache}.xlsx"
            cached_history_file_path = cache_folder_path / cached_history_file_name

            logger.info(f"开始复制历史记录到缓存文件夹: {normalize_drive_letter(str(cached_history_file_path))}")
            copy_cache_success = copy_file(final_history_excel_path, cached_history_file_path, logger)

            if copy_cache_success:
                logger.info("历史记录已成功复制到缓存文件夹。")
                final_files_to_open_at_end.append(cached_history_file_path)
            else:
                logger.error("历史记录复制到缓存文件夹失败。")
        else:
            logger.error(f"无法创建缓存文件夹: {normalize_drive_letter(str(cache_folder_path))}，将无法复制历史记录。")
    else:
        logger.error("历史记录保存到Excel失败，将不会自动打开历史Excel文件。")

    # Add error/warning log file to the list
    if error_warning_log_file_path and error_warning_log_file_path.exists():
        final_files_to_open_at_end.append(error_warning_log_file_path)
    else:
        logger.warning("警告: 错误和警告日志文件不存在或路径无效，无法自动打开。")

    open_output_files_automatically(final_files_to_open_at_end, logger)

    logger.info("所有文件夹处理完毕，程序即将退出。")
    print("所有文件夹处理完毕，程序即将退出。")
    #logger.close()

if __name__ == "__main__":
    main()
