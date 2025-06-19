# InterrogateText2Xlsx6.33.py
import os
import sys
import datetime
import shutil
import subprocess
import re # 导入re模块用于正则表达式
import time # 新增：导入time模块用于延迟重试
from pathlib import Path
from collections import defaultdict
from typing import Tuple, Dict, Optional, Set, List, Any
import hashlib # 新增：导入hashlib用于生成文件夹名的哈希值

# openpyxl 相关的导入
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill
from openpyxl.styles.colors import Color # 导入Color
from openpyxl.worksheet.worksheet import Worksheet # 用于类型提示

# --- Configuration ---
HISTORY_FOLDER_NAME = "反推历史记录"
HISTORY_EXCEL_NAME = "scan_history.xlsx" # 现在直接使用Excel作为历史记录
OUTPUT_FOLDER_NAME = "反推记录" # 在目标文件夹内创建的输出文件夹名称
CACHE_FOLDER_NAME = "cache" # 用于缓存历史记录的文件夹名称 # 已恢复

# 全局或常量定义超链接字体样式
HYPERLINK_FONT = Font(color="0000FF", underline="single")
# 全局常量：所有Excel列的固定宽度
FIXED_COLUMN_WIDTH = 20

# 新增：文件保存重试参数
MAX_SAVE_RETRIES = 5  # 最大重试次数
RETRY_DELAY_SECONDS = 0.5  # 每次重试之间的延迟 (秒)

# --- Utility Function to Normalize Drive Letter ---
def normalize_drive_letter(path_str: str) -> str:
    """
    如果路径以驱动器号开头，将其转换为大写。
    例如: c:\\test -> C:\\test
    """
    if sys.platform.startswith('win') and len(path_str) >= 2 and path_str[1] == ':':
        return path_str[0].upper() + path_str[1:]
    return path_str

# --- NEW FUNCTION: Generate a safe and identifiable folder prefix for filenames ---
def generate_folder_prefix(folder_path: Path) -> str:
    """
    根据文件夹路径生成一个安全且可识别的前缀，用于文件名。
    原理：
        1. 获取文件夹的basename（即文件夹本身的名称）。
        2. 如果 basename 包含中文或特殊字符，为了确保文件名在各种文件系统中的兼容性，
           我们使用该 basename 的MD5哈希值的前8位作为唯一标识。
        3. 如果 basename 只包含ASCII字符（数字、字母、下划线、短横线），
           则直接使用 basename。
        4. 最终前缀会限制长度，避免文件名过长。
    Args:
        folder_path (Path): 文件夹的Path对象。
    Returns:
        str: 一个安全且短小的字符串，用于作为文件名前缀。
    """
    folder_name = folder_path.name
    # 检查是否包含非ASCII字符（例如中文），或者其他不适合作为文件名的字符
    if not re.fullmatch(r'[\w.-]+', folder_name): # 允许字母、数字、下划线、点、短横线
        # 如果包含特殊字符或中文，则使用哈希值
        return hashlib.md5(folder_name.encode('utf-8')).hexdigest()[:8]
    else:
        # 否则，使用文件夹名，并限制长度，防止文件名过长
        return folder_name[:30] # 限制为30个字符，避免过长

# --- LogManager Class ---
class LogManager:
    """
    负责程序的日志记录。
    """
    def __init__(self, log_directory: Path, log_file_name: str = None, 
                 error_log_manager: Optional['LogManager'] = None): # 新增参数
        self.log_directory = log_directory
        self.log_file_path = None
        self.file_handle = None
        self.error_log_manager = error_log_manager # 保存错误日志管理器实例

        # 尝试创建日志目录
        try:
            if not self.log_directory.exists():
                os.makedirs(self.log_directory)
                print(f"已创建日志文件夹: {normalize_drive_letter(str(self.log_directory))}")
                self.write_log(f"已创建日志文件夹: {normalize_drive_letter(str(self.log_directory))}", level="INFO", to_error_log=False) # 避免循环调用
        except Exception as e:
            print(f"关键错误: 无法创建日志文件夹 {normalize_drive_letter(str(self.log_directory))}. 日志将仅打印到控制台. 错误: {e}")
            self.write_log(f"关键错误: 无法创建日志文件夹 {normalize_drive_letter(str(self.log_directory))}. 日志将仅打印到控制台. 错误: {e}", level="CRITICAL", to_error_log=False) # 避免循环调用
            self.log_directory = None # 设为None，后续操作会跳过文件写入

        if self.log_directory: # 如果目录创建成功
            if log_file_name is None:
                self.log_file_path = self.log_directory / f"main_program_log_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.txt"
            else:
                self.log_file_path = self.log_directory / log_file_name
            
            self._open_log_file()
        else: # 目录创建失败，所有日志都将仅打印到控制台
            self.log_file_path = None

    def _open_log_file(self):
        """
        尝试打开日志文件，如果失败则打印到控制台。
        """
        if not self.log_file_path: # 如果没有有效的日志路径，直接返回
            return
        
        try:
            self.file_handle = open(self.log_file_path, 'a', encoding='utf-8')
            self.write_log(f"日志文件已打开: {normalize_drive_letter(str(self.log_file_path))}", level="INFO", to_file_only=True, to_error_log=False) # 避免循环调用
        except Exception as e:
            print(f"关键错误: 无法打开日志文件 {normalize_drive_letter(str(self.log_file_path))}. 所有后续日志将仅打印到控制台. 错误: {e}")
            self.write_log(f"关键错误: 无法打开日志文件 {normalize_drive_letter(str(self.log_file_path))}. 所有后续日志将仅打印到控制台. 错误: {e}", level="CRITICAL", to_file_only=True, to_error_log=False) # 避免循环调用
            self.file_handle = None # 无法打开，设为None

    def write_log(self, message: str, level: str = "INFO", to_file_only: bool = False, to_error_log: bool = True):
        """
        写入日志信息到文件，如果文件句柄无效则打印到控制台。
        Args:
            message (str): 日志消息。
            level (str): 日志级别 (INFO, WARNING, ERROR, CRITICAL).
            to_file_only (bool): 如果为True，则只写入文件，不打印到控制台。
            to_error_log (bool): 如果为True且存在error_log_manager，则将WARNING/ERROR日志写入错误日志。
        """
        timestamp = datetime.datetime.now().strftime("[%Y-%m-%d %H:%M:%S]")
        log_message = f"{timestamp} [{level}] {message}"
        
        # 将WARNING和ERROR级别的日志写入单独的错误日志文件
        if to_error_log and self.error_log_manager and level in ["WARNING", "ERROR", "CRITICAL"]:
            # 避免主日志管理器和错误日志管理器之间的循环写入
            if self.error_log_manager is not self: # 避免自己给自己写
                self.error_log_manager.write_log(message, level=level, to_file_only=True, to_error_log=False)

        if self.file_handle:
            try:
                self.file_handle.write(log_message + "\n")
                self.file_handle.flush() # 立即将缓冲区内容写入文件
                if not to_file_only:
                    print(log_message)
            except Exception as e:
                print(f"关键错误: 写入日志文件 {normalize_drive_letter(str(self.log_file_path))} 失败. 消息: {message}. 错误: {e}")
                self.write_log(f"关键错误: 写入日志文件 {normalize_drive_letter(str(self.log_file_path))} 失败. 消息: {message}. 错误: {e}", level="CRITICAL", to_file_only=True, to_error_log=False) # 避免循环调用
                if self.file_handle:
                    self.file_handle.close()
                self.file_handle = None
                print(f"日志消息重定向到控制台: {log_message}") # 失败后打印到控制台
        else:
            print(log_message)

    def close(self):
        """
        关闭日志文件句柄。
        """
        if self.file_handle:
            try:
                self.file_handle.close()
                self.file_handle = None
                print(f"日志文件已关闭: {normalize_drive_letter(str(self.log_file_path))}")
            except Exception as e:
                print(f"关闭日志文件 {normalize_drive_letter(str(self.log_file_path))} 失败. 错误: {e}")
                self.write_log(f"关闭日志文件 {normalize_drive_letter(str(self.log_file_path))} 失败. 错误: {e}", level="ERROR", to_file_only=True, to_error_log=False) # 避免循环调用

    def __del__(self):
        """
        析构函数，确保在对象被销毁时关闭文件句柄。
        """
        self.close()

# --- NEW FUNCTION: Set Hyperlink and Style ---
def set_hyperlink_and_style(
    cell, 
    location: Optional[str], # location 现在可以是 Optional[str]
    display_text: str, 
    log_manager: LogManager, 
    source_description: str = "未知源"
):
    """
    封装设置单元格超链接和样式的逻辑。
    Args:
        cell: openpyxl 单元格对象。
        location (Optional[str]): 超链接指向的实际位置（文件路径或URL）。如果为None或空字符串，则不设置超链接。
        display_text (str): 在单元格中显示的文本。
        log_manager (LogManager): 日志管理器实例。
        source_description (str): 描述超链接来源，用于日志记录。
    """
    try:
        cell.value = display_text # 首先设置单元格显示文本
        
        # 只有当 location 不为 None 且不为空时才设置超链接
        if location: # 检查 location 是否有效
            cell.hyperlink = location # 然后设置超链接目标
            cell.font = HYPERLINK_FONT # 最后应用预定义的超链接字体样式
            log_manager.write_log(
                f"成功设置超链接和样式 for '{display_text}' (Location: '{location}', Source: {source_description})", 
                level="DEBUG", to_file_only=True
            )
        else:
            # 如果没有 location，确保不设置超链接，并移除可能的超链接样式
            cell.hyperlink = None 
            cell.font = Font(color="000000") # 恢复默认字体颜色，去除下划线
            # 这条日志保留，因为仍然是提示没有设置超链接，但级别可以低一些
            log_manager.write_log(
                f"未为 '{display_text}' (Source: {source_description}) 设置超链接，因为location无效或为空。", 
                level="INFO", to_file_only=True
            )

    except Exception as e:
        log_manager.write_log(
            f"错误: 无法为单元格设置超链接或样式 for '{display_text}' (Location: '{location}', Source: {source_description}). 错误: {e}", 
            level="ERROR"
        )
        # 即使出错，也要确保单元格值被设置，即使没有超链接样式
        cell.value = display_text

# --- NEW FUNCTION: Set Fixed Column Widths for a Worksheet ---
def set_fixed_column_widths(worksheet: Worksheet, width: int, log_manager: LogManager):
    """
    为给定工作表的所有列设置固定宽度。
    Args:
        worksheet (Worksheet): openpyxl 工作表对象。
        width (int): 要设置的列宽。
        log_manager (LogManager): 日志管理器实例。
    """
    try:
        for col_idx in range(1, worksheet.max_column + 1): # 从1开始遍历所有列
            column_letter = get_column_letter(col_idx)
            worksheet.column_dimensions[column_letter].width = width
        log_manager.write_log(f"已为工作表 '{worksheet.title}' 设置所有列宽度为 {width}.", level="INFO", to_file_only=True)
    except Exception as e:
        log_manager.write_log(f"错误: 无法为工作表 '{worksheet.title}' 设置列宽: {e}", level="ERROR")
        print(f"错误: 无法为工作表 '{worksheet.title}' 设置列宽. 错误: {e}")

# --- HistoryManager (Excel Version) ---
class HistoryManager:
    """
    负责程序扫描历史记录的Excel文件存储。
    """
    def __init__(self, history_file_path: Path, log_manager: LogManager):
        self.history_file_path = history_file_path
        self.log_manager = log_manager
        self.history_data: List[Dict[str, Any]] = [] # 存储内存中的历史记录
        self._load_history_from_excel()

    def _load_history_from_excel(self):
        """
        从Excel文件加载历史记录到内存。
        """
        self.history_data = []
        if not self.history_file_path.exists():
            self.log_manager.write_log(f"历史记录Excel文件不存在: {normalize_drive_letter(str(self.history_file_path))}. 将创建新文件。", level="INFO")
            return

        try:
            wb = load_workbook(str(self.history_file_path))
            if "扫描历史" in wb.sheetnames:
                ws = wb["扫描历史"]
                headers = [cell.value for cell in ws[1]] # 获取表头
                if not headers:
                    self.log_manager.write_log(f"历史记录Excel文件 '{normalize_drive_letter(str(self.history_file_path))}' 的 '扫描历史' 工作表为空，无历史记录可加载。", level="WARNING")
                    return

                # 确保表头符合预期，避免因旧文件格式导致的问题
                expected_headers = [
                    "扫描时间",
                    "文件夹路径",
                    "总文件数",
                    "找到TXT文件数",
                    "未找到TXT文件数",
                    "Log文件绝对路径",
                    "Log文件超链接", # 此列不再是实际数据，而是超链接显示文本
                    "结果XLSX文件绝对路径",
                    "结果XLSX文件超链接" # 此列不再是实际数据，而是超链接显示文本
                ]
                # 简化检查，只需要检查前5列和两个绝对路径列是否存在，因为超链接列是动态生成的
                if not all(h in headers for h in expected_headers[:6] + [expected_headers[7]]): # 检查前6个和第8个（结果XLSX绝对路径）
                    self.log_manager.write_log(f"历史记录Excel文件 '{normalize_drive_letter(str(self.history_file_path))}' 表头不匹配预期，可能无法完全加载所有历史记录。", level="WARNING")
                    # 继续尝试加载，但可能不完整

                for row_idx in range(2, ws.max_row + 1): # 从第二行开始读取数据
                    row_values = [cell.value for cell in ws[row_idx]]
                    # 将行数据映射到字典
                    entry = {}
                    for i, header in enumerate(headers):
                        if i < len(row_values):
                            entry[header] = row_values[i]
                        else:
                            entry[header] = None # 如果某些列没有值，则设置为None

                    # 仅保留我们需要的字段，并确保路径是Path对象
                    self.history_data.append({
                        'scan_time': entry.get("扫描时间"),
                        'folder_path': Path(entry.get("文件夹路径")) if entry.get("文件夹路径") else None,
                        'total_files': entry.get("总文件数"),
                        'found_txt_count': entry.get("找到TXT文件数"),
                        'not_found_txt_count': entry.get("未找到TXT文件数"),
                        'log_file_abs_path': Path(entry.get("Log文件绝对路径")) if entry.get("Log文件绝对路径") else None,
                        'result_xlsx_abs_path': Path(entry.get("结果XLSX文件绝对路径")) if entry.get("结果XLSX文件绝对路径") else None
                    })
            self.log_manager.write_log(f"成功从历史记录Excel文件加载 {len(self.history_data)} 条历史记录。", level="INFO")
        except Exception as e:
            self.log_manager.write_log(f"错误: 从历史记录Excel文件 {normalize_drive_letter(str(self.history_file_path))} 加载历史记录失败: {e}", level="ERROR")
            self.history_data = [] # 加载失败则清空内存数据，避免脏数据

    def add_history_entry(self, folder_path: Path, total_scanned: int, found_txt_count: int,
                          not_found_txt_count: int, result_file_path: Path, log_file_path: Path):
        """
        向内存中的历史记录列表添加一条新的扫描历史记录。
        """
        scan_time = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        entry = {
            'scan_time': scan_time,
            'folder_path': folder_path,
            'total_files': total_scanned,
            'found_txt_count': found_txt_count,
            'not_found_txt_count': not_found_txt_count,
            'log_file_abs_path': log_file_path,
            'result_xlsx_abs_path': result_file_path
        }
        self.history_data.append(entry)
        self.log_manager.write_log(f"历史记录成功添加至内存: 文件夹'{folder_path.name}'", level="INFO")

    def save_history_to_excel(self) -> bool:
        """
        将内存中的所有历史记录保存到Excel文件。
        Returns:
            bool: 如果保存成功返回True，否则返回False。
        """
        self.log_manager.write_log(f"开始将内存中的历史记录保存到Excel: {normalize_drive_letter(str(self.history_file_path))}", level="INFO")

        # 尝试删除旧的历史Excel文件，以便重新写入
        if self.history_file_path.exists():
            try:
                os.remove(str(self.history_file_path))
                self.log_manager.write_log(f"已删除旧的历史记录Excel文件: {normalize_drive_letter(str(self.history_file_path))}", level="INFO")
            except PermissionError as e: # 明确捕获权限错误
                self.log_manager.write_log(f"警告: 无法删除旧的历史记录Excel文件 {normalize_drive_letter(str(self.history_file_path))}. 可能文件被占用. 错误: {e}", level="WARNING")
                print(f"警告: 无法删除旧的历史记录Excel文件 {self.history_file_path}. 可能文件被占用. 错误: {e}")
                self.log_manager.write_log("无法覆盖旧的历史文件，历史记录将无法保存。请关闭Excel中打开的历史文件。", level="CRITICAL")
                return False # 删除失败，返回False
            except Exception as e:
                self.log_manager.write_log(f"警告: 删除旧的历史记录Excel文件时发生未知错误 {normalize_drive_letter(str(self.history_file_path))}. 错误: {e}", level="WARNING")
                print(f"警告: 删除旧的历史记录Excel文件时发生未知错误 {self.history_file_path}. 错误: {e}")
                self.log_manager.write_log("删除旧文件失败，历史记录将无法保存。", level="CRITICAL")
                return False # 删除失败，返回False


        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "扫描历史"

            # 定义Excel表头
            excel_headers = [
                "扫描时间",
                "文件夹路径",
                "总文件数",
                "找到TXT文件数",
                "未找到TXT文件数",
                "Log文件绝对路径",
                "Log文件超链接",
                "结果XLSX文件绝对路径",
                "结果XLSX文件超链接"
            ]
            ws.append(excel_headers)

            for entry in self.history_data:
                log_file_abs_path = entry['log_file_abs_path']
                result_xlsx_abs_path = entry['result_xlsx_abs_path']

                # Log文件超链接的显示文本和location
                log_link_display_text = "打开Log" 
                log_link_location = None
                if log_file_abs_path and log_file_abs_path.exists():
                    log_link_location = normalize_drive_letter(str(log_file_abs_path)).replace("\\", "/")
                    if not sys.platform.startswith('win'): 
                        log_link_location = f'file://{log_link_location}'
                else:
                    log_link_display_text = "Log文件不存在"


                # 结果XLSX文件超链接的显示文本和location
                result_link_display_text = "打开结果XLSX"
                result_link_location = None
                if result_xlsx_abs_path and result_xlsx_abs_path.exists():
                    result_link_location = normalize_drive_letter(str(result_xlsx_abs_path)).replace("\\", "/")
                    if not sys.platform.startswith('win'): 
                        result_link_location = f'file://{result_link_location}'
                else:
                    result_link_display_text = "结果XLSX文件不存在"
                
                row_data = [
                    entry['scan_time'],
                    str(entry['folder_path']),
                    entry['total_files'],
                    entry['found_txt_count'],
                    entry['not_found_txt_count'],
                    normalize_drive_letter(str(log_file_abs_path)) if log_file_abs_path else "N/A",      # Log文件绝对路径
                    log_link_display_text,       # Log文件超链接显示文本
                    normalize_drive_letter(str(result_xlsx_abs_path)) if result_xlsx_abs_path else "N/A",   # 结果XLSX文件绝对路径
                    result_link_display_text    # 结果XLSX文件超链接显示文本
                ]
                ws.append(row_data)

                # 获取新添加的行的单元格，并设置超链接
                new_row_idx = ws.max_row
                
                # Log文件超链接单元格 (第7列)
                log_link_cell = ws.cell(row=new_row_idx, column=7)
                set_hyperlink_and_style(
                    log_link_cell, 
                    log_link_location, # 传入可能为None的location
                    log_link_display_text, 
                    self.log_manager, 
                    source_description=f"历史记录Log文件 (行: {new_row_idx})"
                )

                # 结果XLSX文件超链接单元格 (第9列)
                result_link_cell = ws.cell(row=new_row_idx, column=9)
                set_hyperlink_and_style(
                    result_link_cell, 
                    result_link_location, # 传入可能为None的location
                    result_link_display_text, 
                    self.log_manager, 
                    source_description=f"历史记录结果XLSX文件 (行: {new_row_idx})"
                )
            
            # 设置所有列宽
            set_fixed_column_widths(ws, FIXED_COLUMN_WIDTH, self.log_manager)
            
            wb.save(str(self.history_file_path))
            self.log_manager.write_log(f"成功将历史记录保存到Excel: {normalize_drive_letter(str(self.history_file_path))}", level="INFO")
            return True

        except Exception as e:
            self.log_manager.write_log(f"错误: 将历史记录保存到Excel失败 {normalize_drive_letter(str(self.history_file_path))}: {e}", level="ERROR")
            print(f"错误: 将历史记录保存到Excel失败 {self.history_file_path}. 错误: {e}")
            return False

# --- File Operations ---
def validate_directory(path: Path, log_manager: Optional[LogManager]) -> bool:
    """
    验证给定的路径是否是一个存在的目录。
    """
    if not path.is_dir():
        if log_manager:
            log_manager.write_log(f"验证失败: 目录不存在或不是一个目录: {normalize_drive_letter(str(path))}", level="WARNING")
        return False
    return True

def create_directory_if_not_exists(directory_path: Path, log_manager: Optional[LogManager]) -> bool:
    """
    如果指定目录不存在，则创建它。
    Args:
        directory_path (Path): 要创建的目录路径。
        log_manager (Optional[LogManager]): 日志管理器实例，可选。
    Returns:
        bool: 如果目录存在或成功创建，则返回True；否则返回False。
    """
    if not directory_path.exists():
        try:
            os.makedirs(directory_path)
            if log_manager:
                log_manager.write_log(f"已创建目录: {normalize_drive_letter(str(directory_path))}", level="INFO")
            return True
        except OSError as e:
            if log_manager:
                log_manager.write_log(f"错误: 无法创建目录 {normalize_drive_letter(str(directory_path))}: {e}", level="ERROR")
            print(f"错误: 无法创建文件夹 {directory_path}。错误: {e}")
            return False
    return True

def copy_file(source_path: Path, destination_path: Path, log_manager: Optional[LogManager]) -> bool:
    """
    复制文件从源路径到目标路径。
    增加对权限错误的捕获和提示。
    """
    if not source_path.exists():
        if log_manager:
            log_manager.write_log(f"错误: 源文件不存在，无法复制: {normalize_drive_letter(str(source_path))}", level="ERROR")
        print(f"错误: 源文件不存在，无法复制: {source_path}")
        return False

    try:
        shutil.copy2(str(source_path), str(destination_path)) 
        if log_manager:
            log_manager.write_log(f"已复制 '{normalize_drive_letter(str(source_path))}' 到 '{normalize_drive_letter(str(destination_path))}'", level="INFO")
        return True
    except PermissionError as e:
        if log_manager:
            log_manager.write_log(
                f"权限错误: 复制文件从 '{normalize_drive_letter(str(source_path))}' 到 '{normalize_drive_letter(str(destination_path))}' 失败: {e}. 请确保目标文件未被其他程序（如Excel）占用。", 
                level="CRITICAL"
            )
        print(f"错误: 权限拒绝！无法复制文件到 '{destination_path}'。请确保该文件未被其他程序（如Excel）打开。错误: {e}")
        return False
    except Exception as e:
        if log_manager:
            log_manager.write_log(f"错误: 复制文件从 '{normalize_drive_letter(str(source_path))}' 到 '{normalize_drive_letter(str(destination_path))}' 失败: {e}", level="ERROR")
        print(f"错误: 无法复制文件从 '{source_path}' 到 '{destination_path}'。错误: {e}")
        return False

def get_file_details(file_path: Path) -> Tuple[str, str]:
    """
    获取文件的名称（不含扩展名）和扩展名。
    """
    return file_path.stem, file_path.suffix

# --- Excel Utilities ---
def create_main_workbook():
    """
    创建主Excel工作簿，包含“匹配文件”和“未匹配文件”工作表。
    """
    wb = Workbook()
    
    # 移除默认创建的Sheet
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
        
    return wb

def setup_excel_sheets(wb: Workbook) -> Tuple[Worksheet, Worksheet, Worksheet]:
    """
    设置Excel工作表及其标题行。
    """
    # 匹配文件工作表
    ws_matched = wb.create_sheet("匹配文件", 0) # 插入到最前面
    ws_matched.append([
        "文件夹路径", "文件绝对路径", "文件链接", "文件扩展名", "TXT文件绝对路径",
        "TXT文件内容", "清洗后内容", "内容长度", "提示词类型", "找到TXT"
    ])

    # 未匹配文件工作表
    ws_no_txt = wb.create_sheet("未匹配文件", 1) # 插入到第二个
    ws_no_txt.append([
        "文件夹路径", "文件绝对路径", "文件链接", "文件扩展名", "找到TXT"
    ])

    # Tag词频统计工作表
    ws_tag_frequency = wb.create_sheet("Tag词频统计", 2) # 插入到第三个
    ws_tag_frequency.append(["Tag", "出现次数"])

    return ws_matched, ws_no_txt, ws_tag_frequency

# --- Data Processor (RESTORED FROM V4.0) ---
def detect_types(line: str, cleaned: str) -> str:
    """
    根据文本内容推断提示词类型。还原自 V4.0 版本。
    Args:
        line (str): 原始的txt文件内容。
        cleaned (str): 清洗后的txt文件内容。
    Returns:
        str: 识别到的提示词类型，用逗号分隔。
    """
    types = []
    lower_line = line.lower()
    # R18相关词汇
    if any(word in lower_line for word in [
        'sex', 'nude', 'pussy', 'penis', 'cum', 'nipples', 'vaginal',
        'cum_in_pussy', 'oral', 'rape', 'fellatio', 'facial', 'anus',
        'anal', 'ejaculation', 'gangbang', 'testicles', 'multiple_penises',
        'erection', 'handjob', 'cumdrip', 'pubic_hair', 'pussy_juice',
        'bukkake', 'clitoris', 'female_ejaculation', 'threesome',
        'doggystyle', 'sex_from_behind', 'cum_on_breasts', 'double_penetration',
        'anal_object_insertion', 'cunnilingus', 'triple_penetration',
        'paizuri', 'vaginal_object_insertion', 'imminent_rape', 'impregnation',
        'prone_bone', 'reverse_cowgirl_position', 'cum_inflation',
        'milking_machine', 'cumdump', 'anal_hair', 'futanari', 'glory_hole',
        'penis_on_face', 'licking_penis', 'breast_sucking', 'breast_squeeze', 'straddling'
    ]):
        types.append('R18')
    # boy类型
    if any(boy_word in lower_line for boy_word in ['1boy', '2boys', 'multiple_boys']):
        types.append('boy')
    # no_human类型
    if 'no_human' in lower_line:
        types.append('no_human')
    # furry 类型
    if any(word in lower_line for word in ['furry', 'animal_focus']):
        types.append('furry')
    # monochrome和greyscale类型
    if any(word in lower_line for word in ['monochrome', 'greyscale']):
        types.append('黑白原图')
    # 新增功能：检测"background"相关词汇并标记为“简单背景”类型
    if 'background' in lower_line:
        types.append('简单背景')
    
    # 如果没有检测到任何类型，返回 "N/A"
    if not types:
        return "N/A"
    return ','.join(types)

def clean_tags(line: str) -> Tuple[str, bool]:
    """
    清洗标签字符串。修改了对'censor'词的清理逻辑和'uncensored'的添加逻辑。
    Args:
        line (str): 原始的标签字符串。
    Returns:
        Tuple[str, bool]: 清洗后的字符串和是否含有敏感词的布尔值。
    """
    tags = [tag.strip() for tag in line.strip().split(',')]
    
    # 定义需要清洗掉的关键词，不包含 'uncensored'
    words_to_clean = ['censor', 'censored', 'monochrome', 'greyscale', 'furry', 'animal_focus', 'no_human', 'background']
    
    # 检查是否含有敏感词 (基于原始标签列表，因为这些词不应该被清洗掉，而是用于标记)
    has_sensitive = any(
        any(word in tag.lower() for word in [
            'nipple', 'pussy', 'penis', 'hetero', 'sex', 'anus', 'naked', 'explicit' # 增加一些常见的敏感词
        ])
        for tag in tags
    )
    
    # 过滤掉需要清洗的关键词
    cleaned_tags = []
    for tag in tags:
        # 如果是 'uncensored'，直接添加，不进行清洗
        if tag.lower() == 'uncensored':
            cleaned_tags.append(tag)
            continue
            
        # 只有当tag不包含任何words_to_clean中的词时才保留
        if not any(word in tag.lower() for word in words_to_clean):
            cleaned_tags.append(tag)

    # 如果检测到敏感词，则添加 'uncensored' 标记 
    # 确保只添加一次
    if has_sensitive and 'uncensored' not in [t.lower() for t in cleaned_tags]:
        cleaned_tags.append('uncensored')
    
    # 过滤掉空字符串，然后用逗号和空格连接
    cleaned_line = ', '.join([tag for tag in cleaned_tags if tag])
    return cleaned_line, has_sensitive


# --- Scanner ---
def scan_files_and_extract_data(
    base_folder_path: Path,
    ws_matched: Worksheet,
    ws_no_txt: Worksheet,
    log_manager: LogManager
) -> Tuple[int, int, int, Dict[str, int]]:
    """
    扫描指定文件夹下的文件，查找匹配的TXT文件，提取数据并写入Excel。
    """
    total_files_scanned = 0
    found_txt_count = 0
    not_found_txt_count = 0
    tag_counts = defaultdict(int)

    # 新增：定义要跳过的文件夹名称
    skip_scan_folders = {'.bf'} # 使用集合以便快速查找
    # 新增：定义要跳过的文件扩展名
    skip_scan_extensions = {'.txt', '.xlsx', '.json', '.ini', '.db'} # 新增 .db 到跳过列表
    # 使用集合以便快速查找

    # 新增：用于收集文件扩展名信息
    all_extensions: Set[str] = set()
    skipped_extensions: Set[str] = set()

    log_manager.write_log(f"开始扫描文件夹: {normalize_drive_letter(str(base_folder_path))}", level="INFO")

    try:
        for root_str, dirs, files in os.walk(base_folder_path):
            root = Path(root_str)

            should_skip_current_path = False
            for skip_folder_name in skip_scan_folders:
                if skip_folder_name in root.parts: 
                    should_skip_current_path = True
                    break
            
            if root.name in skip_scan_folders:
                should_skip_current_path = True

            if should_skip_current_path:
                log_manager.write_log(f"跳过扫描文件夹及其子文件夹，因为它包含要跳过的名称: {normalize_drive_letter(str(root))}", level="INFO", to_file_only=True)
                dirs[:] = [] 
                continue 

            current_txt_files = {os.path.splitext(f)[0].lower(): root / f for f in files if f.lower().endswith('.txt')}

            for file_name in files:
                file_path = root / file_name
                file_stem, file_ext = get_file_details(file_path)

                all_extensions.add(file_ext.lower()) 

                if file_ext.lower() in skip_scan_extensions:
                    skipped_extensions.add(file_ext.lower()) 
                    continue 

                total_files_scanned += 1
                
                file_abs_path = file_path.resolve()
                
                # --- 主要修改点：根据文件是否存在来设置超链接的 location 和 display_text ---
                file_link_location = None
                file_link_text = "文件不存在" # 默认显示文本

                if file_abs_path.exists():
                    file_link_location = normalize_drive_letter(str(file_abs_path)).replace("\\", "/") 
                    if not sys.platform.startswith('win'): 
                        file_link_location = f'file://{file_link_location}'
                    file_link_text = str(file_abs_path.name) # 超链接显示文本为文件名

                else:
                    # 这条日志保留，因为是文件系统层面的缺失，但级别可以低一些
                    log_manager.write_log(f"文件不存在，无法生成有效超链接: {normalize_drive_letter(str(file_abs_path))}", level="INFO", to_file_only=True) 
                    file_link_text = f"文件不存在: {file_abs_path.name}" # 提示文件不存在

                txt_content = ""
                cleaned_data = ""
                cleaned_data_length = 0
                prompt_type = "N/A"
                found_txt_flag = '否' 
                txt_absolute_path = "N/A" 

                if file_stem.lower() in current_txt_files:
                    txt_file_path = current_txt_files[file_stem.lower()]
                    txt_absolute_path = txt_file_path.resolve()
                    found_txt_flag = '是'
                    found_txt_count += 1
                    try:
                        with open(txt_file_path, 'r', encoding='utf-8') as f:
                            for line in f:
                                txt_content = line.strip()
                                cleaned_data, _ = clean_tags(txt_content) 
                                cleaned_data_length = len(cleaned_data)
                                prompt_type = detect_types(txt_content, cleaned_data) 
                                
                                for tag in cleaned_data.split(', '):
                                    if tag:
                                        tag_counts[tag.strip().lower()] += 1
                                break
                    except Exception as e:
                        log_manager.write_log(f"错误: 读取或处理TXT文件 {normalize_drive_letter(str(txt_file_path))} 失败: {e}", level="ERROR")
                        txt_content = f"Error reading TXT: {e}"
                        found_txt_flag = '否 (读取错误)'
                        not_found_txt_count += 1
                else:
                    log_manager.write_log(f"未找到匹配的TXT文件: {normalize_drive_letter(str(file_path))}", level="INFO", to_file_only=True)
                    not_found_txt_count += 1
                
                current_row_data = [
                    str(root.resolve()),
                    str(file_abs_path),
                    file_link_text, # 这里直接传入 file_link_text 作为显示文本
                    file_ext,
                    str(txt_absolute_path),
                    txt_content,
                    cleaned_data,
                    cleaned_data_length,
                    prompt_type,
                    found_txt_flag 
                ]

                if found_txt_flag == '是': 
                    ws_matched.append(current_row_data)
                    link_cell = ws_matched.cell(row=ws_matched.max_row, column=3)
                    set_hyperlink_and_style(
                        link_cell, 
                        file_link_location, # 传入可能为None的location
                        file_link_text, # 传入已准备好的显示文本
                        log_manager, 
                        source_description=f"匹配文件 (行: {ws_matched.max_row})"
                    )

                else:
                    current_row_data_no_txt = [
                        str(root.resolve()),
                        str(file_abs_path),
                        file_link_text, # 这里直接传入 file_link_text 作为显示文本
                        file_ext,
                        found_txt_flag
                    ]
                    ws_no_txt.append(current_row_data_no_txt)
                    link_cell = ws_no_txt.cell(row=ws_no_txt.max_row, column=3)
                    set_hyperlink_and_style(
                        link_cell, 
                        file_link_location, # 传入可能为None的location
                        file_link_text, # 传入已准备好的显示文本
                        log_manager, 
                        source_description=f"未匹配文件 (行: {ws_no_txt.max_row})"
                    )

    except Exception as e:
        log_manager.write_log(f"致命错误: 扫描文件过程中发生意外错误 for folder {normalize_drive_letter(str(base_folder_path))}: {e}", level="CRITICAL")
        print(f"致命错误: 扫描文件过程中发生意外错误 for folder {base_folder_path}: {e}")
    
    log_manager.write_log(f"文件夹 {normalize_drive_letter(str(base_folder_path))} 扫描完成. 总文件数: {total_files_scanned}, 找到TXT: {found_txt_count}, 未找到TXT: {not_found_txt_count}", level="INFO")
    
    log_manager.write_log(f"\n--- 扫描文件类型概览 ---", level="INFO")
    if all_extensions:
        for ext in sorted(list(all_extensions)):
            status = "已处理"
            if ext in skipped_extensions:
                status = "已跳过"
            log_manager.write_log(f"文件扩展名: '{ext}' - 状态: {status}", level="INFO")
    else:
        log_manager.write_log("未扫描到任何文件扩展名。", level="INFO")
    log_manager.write_log(f"--- 文件类型概览结束 ---\n", level="INFO")

    # 修正这里，返回 found_txt_count 和 not_found_txt_count
    return total_files_scanned, found_txt_count, not_found_txt_count, tag_counts


def open_output_files_automatically(files_to_open: List[Path], main_log_manager: LogManager):
    """
    自动打开指定的文件。
    Args:
        files_to_open (List[Path]): 要打开的文件路径列表。
        main_log_manager (LogManager): 主日志管理器实例。
    """
    try:
        # 定义一个正则表达式来匹配文件路径中的日期时间戳 (YYYYMMDD_HHMMSS)
        # 允许文件名在时间戳前后有其他字符，例如 scan_results_20240101_123045.xlsx 或 scan_history_backup_20240101_123045.xlsx
        timestamp_pattern = re.compile(r'\d{8}_\d{6}')

        for file_path_to_open in files_to_open:
            if not file_path_to_open.exists():
                main_log_manager.write_log(f"警告: 尝试打开不存在的文件: {normalize_drive_letter(str(file_path_to_open))}", level="WARNING")
                continue

            # 检查文件名是否包含时间戳，或者是否为明确允许打开的临时文件（例如本次扫描结果）
            # 我们只允许打开包含时间戳的文件作为“快照”，不打开作为“数据库”的源文件
            # 这里的逻辑是：如果文件名包含时间戳，或者文件名是本次扫描生成的xlsx或log文件，则允许打开
            file_name = file_path_to_open.name
            
            is_allowed_to_open = False
            
            # 检查是否包含时间戳（例如 cache 中的历史文件，或者备份的历史文件）
            if timestamp_pattern.search(file_name):
                is_allowed_to_open = True
            
            # 检查是否是本次扫描生成的 Excel 或 Log 文件（这些文件本身就包含时间戳和/或文件夹前缀）
            # 新的命名约定：[folder_prefix]_scan_results_YYYYMMDD_HHMMSS.xlsx
            #             [folder_prefix]_scan_log_YYYYMMDD_HHMMSS.txt
            if re.match(r'.*_scan_results_\d{8}_\d{6}\.xlsx', file_name):
                is_allowed_to_open = True
            elif re.match(r'.*_scan_log_\d{8}_\d{6}\.txt', file_name):
                is_allowed_to_open = True
            elif file_name.startswith("error_warning_log_") and file_name.endswith(".txt"): # 允许打开错误日志
                is_allowed_to_open = True
            # [RESTORE] 恢复对以“scan_history_cached_”开头的文件的特殊允许，因为现在需要缓存备份
            elif file_name.startswith("scan_history_cached_") and file_name.endswith(".xlsx"):
                is_allowed_to_open = True
            # [NEW] 允许打开因权限问题而生成的备用结果文件
            elif file_name.startswith("FALLBACK_") and file_name.endswith(".xlsx"): # 修改为更通用的匹配FALLBACK_
                is_allowed_to_open = True


            if not is_allowed_to_open:
                main_log_manager.write_log(f"拒绝自动打开没有时间戳或不符合命名约定的源文件: {normalize_drive_letter(str(file_path_to_open))}", level="WARNING")
                print(f"拒绝自动打开没有时间戳或不符合命名约定的源文件: {file_path_to_open}")
                continue


            if sys.platform.startswith('win'): 
                subprocess.Popen(f'start "" "{file_path_to_open}"', shell=True) 
            elif sys.platform == 'darwin': 
                subprocess.Popen(['open', str(file_path_to_open)])
            else: 
                subprocess.Popen(['xdg-open', str(file_path_to_open)])
            
            print(f"自动打开: {file_path_to_open}")
            main_log_manager.write_log(f"自动打开: {normalize_drive_letter(str(file_path_to_open))}", level="INFO")

    except Exception as e:
        main_log_manager.write_log(f"错误: 自动打开文件失败. 错误: {e}", level="ERROR")
        print(f"错误: 无法自动打开文件。请手动检查。错误: {e}")

# --- New Function: Read Batch Paths ---
def read_batch_paths(batch_file_path: Path, log_manager: LogManager) -> List[Path]:
    """
    从 batchPath.txt 文件中读取需要扫描的文件夹路径列表。
    Args:
        batch_file_path (Path): batchPath.txt 文件的路径。
        log_manager (LogManager): 日志管理器实例。
    Returns:
        List[Path]: 文件夹路径的列表。
    """
    folders = []
    if not batch_file_path.exists():
        log_manager.write_log(f"错误: 批量路径文件 '{normalize_drive_letter(str(batch_file_path))}' 不存在。", level="ERROR")
        print(f"错误: 批量路径文件 '{batch_file_path}' 不存在。")
        return folders
    try:
        with open(batch_file_path, 'r', encoding='utf-8') as f:
            for line in f:
                path_str = line.strip()
                if path_str and not path_str.startswith('#'): # 忽略空行和注释行
                    folder_path = Path(path_str)
                    if validate_directory(folder_path, log_manager):
                        folders.append(folder_path)
                    else:
                        log_manager.write_log(f"警告: 批量路径文件中的路径无效或不存在，已跳过: {normalize_drive_letter(str(folder_path))}", level="WARNING")
        if not folders:
            log_manager.write_log(f"警告: 批量路径文件 '{normalize_drive_letter(str(batch_file_path))}' 中没有找到有效的文件夹路径。", level="WARNING")
    except Exception as e:
        log_manager.write_log(f"错误: 读取批量路径文件 '{normalize_drive_letter(str(batch_file_path))}' 失败: {e}", level="CRITICAL")
        print(f"错误: 读取批量路径文件 '{batch_file_path}' 失败。错误: {e}")
    return folders


def main():
    script_dir = Path(os.path.dirname(os.path.abspath(__file__)))
    log_output_dir = script_dir / "logs"
    history_folder_path = script_dir / HISTORY_FOLDER_NAME
    output_base_dir = script_dir / OUTPUT_FOLDER_NAME # 这是反推记录文件夹
    final_history_excel_path = history_folder_path / HISTORY_EXCEL_NAME 
    
    # [RESTORE] 恢复对 CACHE_FOLDER_NAME 的定义和使用，因为现在需要缓存备份
    cache_folder_path = script_dir / CACHE_FOLDER_NAME

    # 新增：用于记录WARNING和ERROR的独立日志管理器
    error_warning_log_dir = script_dir / "logs" # 错误日志也放在logs文件夹
    error_warning_log_file_name = f"error_warning_log_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.txt"
    error_warning_log_manager = LogManager(error_warning_log_dir, log_file_name=error_warning_log_file_name)
    error_warning_log_manager.write_log("错误和警告日志文件已创建。", level="INFO", to_error_log=False) # 这条日志不需再转发

    # 修改：主日志管理器现在可以将错误日志转发给 error_warning_log_manager
    main_log_manager = LogManager(log_output_dir, error_log_manager=error_warning_log_manager)
    main_log_manager.write_log(f"程序启动. 脚本目录: {normalize_drive_letter(str(script_dir))}", level="INFO")
    main_log_manager.write_log(f"错误和警告日志将写入到: {normalize_drive_letter(str(error_warning_log_manager.log_file_path))}", level="INFO")


    if not create_directory_if_not_exists(history_folder_path, main_log_manager):
        main_log_manager.write_log("致命错误: 无法创建历史记录文件夹，程序退出。", level="CRITICAL")
        error_warning_log_manager.close() # 确保关闭错误日志
        sys.exit(1)

    # [RESTORE] 恢复缓存文件夹的创建逻辑
    if not create_directory_if_not_exists(cache_folder_path, main_log_manager):
        main_log_manager.write_log("致命错误: 无法创建缓存文件夹，程序退出。", level="CRITICAL")
        error_warning_log_manager.close() # 确保关闭错误日志
        sys.exit(1)
    
    # 新增：确保反推记录文件夹存在
    if not create_directory_if_not_exists(output_base_dir, main_log_manager):
        main_log_manager.write_log("致命错误: 无法创建输出文件夹 (反推记录)，程序退出。", level="CRITICAL")
        error_warning_log_manager.close()
        sys.exit(1)


    # 初始化历史管理器 (现在是Excel版本)
    history_manager = HistoryManager(final_history_excel_path, main_log_manager)

    batch_file_path = script_dir / "batchPath.txt"
    # 调用read_batch_paths函数
    folders_to_scan = read_batch_paths(batch_file_path, main_log_manager)
    
    # 新增的日志提示：显示从batchPath.txt读取到的有效路径数量
    main_log_manager.write_log(f"在 '{normalize_drive_letter(str(batch_file_path))}' 中检测到 {len(folders_to_scan)} 条有效地址。", level="INFO")

    if not folders_to_scan:
        main_log_manager.write_log("没有找到要扫描的文件夹路径，程序终止。", level="INFO")
        print("没有找到要扫描的文件夹路径，程序终止。")
        main_log_manager.close()
        error_warning_log_manager.close() # 确保关闭错误日志
        sys.exit(0)

    for folder_path in folders_to_scan:
        main_log_manager.write_log(f"\n--- 开始处理文件夹: {normalize_drive_letter(str(folder_path))} ---", level="INFO")
        print(f"\n--- 开始处理文件夹: {folder_path} ---")

        scan_timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        folder_prefix = generate_folder_prefix(folder_path) # [NEW] 生成文件夹前缀
        
        # 移除了 current_output_folder 的创建
        # [MODIFIED] 文件名中包含文件夹前缀
        current_scan_log_file = output_base_dir / f"{folder_prefix}_scan_log_{scan_timestamp}.txt" # 直接在 output_base_dir 下
        current_excel_file = output_base_dir / f"{folder_prefix}_scan_results_{scan_timestamp}.xlsx" # 直接在 output_base_dir 下
        
        # [NEW] 定义备用 Excel 文件路径
        fallback_excel_file = log_output_dir / f"FALLBACK_{folder_prefix}_scan_results_{scan_timestamp}.xlsx" # [MODIFIED] 备用文件名也包含文件夹前缀

        # 确保每个扫描日志使用独立的LogManager实例，现在log_directory直接指向 output_base_dir
        # [MODIFIED] log_file_name 包含文件夹前缀
        scan_log_manager = LogManager(output_base_dir, log_file_name=f"{folder_prefix}_scan_log_{scan_timestamp}.txt", error_log_manager=error_warning_log_manager)
        scan_log_manager.write_log(f"开始扫描 {normalize_drive_letter(str(folder_path))}", level="INFO")

        try:
            wb = create_main_workbook()
            ws_matched, ws_no_txt, ws_tag_frequency = setup_excel_sheets(wb)

            total_files, found_txt_count, not_found_txt_count, tag_counts = scan_files_and_extract_data( # 修正这里的变量名
                folder_path, ws_matched, ws_no_txt, scan_log_manager
            )

            sorted_tags = sorted(tag_counts.items(), key=lambda item: item[1], reverse=True)
            for tag, count in sorted_tags:
                ws_tag_frequency.append([tag, count])

            for worksheet in [ws_matched, ws_no_txt, ws_tag_frequency]:
                set_fixed_column_widths(worksheet, FIXED_COLUMN_WIDTH, scan_log_manager)
            
            # --- 主要修改点：尝试保存Excel文件，如果权限拒绝则保存到logs目录 ---
            save_successful = False
            actual_result_file_path = Path("N/A_SAVE_FAILED") # 默认标记为保存失败

            for attempt in range(MAX_SAVE_RETRIES):
                try:
                    wb.save(str(current_excel_file))
                    scan_log_manager.write_log(f"扫描结果已保存到: {normalize_drive_letter(str(current_excel_file))} (尝试 {attempt + 1}/{MAX_SAVE_RETRIES})", level="INFO")
                    print(f"扫描结果已保存到: {current_excel_file}")
                    actual_result_file_path = current_excel_file
                    save_successful = True
                    break # 成功保存，跳出重试循环
                except PermissionError as e:
                    scan_log_manager.write_log(
                        f"警告: 无法将扫描结果保存到 '{normalize_drive_letter(str(current_excel_file))}'，原因: 权限拒绝！请确保该文件未被其他程序（如Excel）打开。尝试 {attempt + 1}/{MAX_SAVE_RETRIES}。错误: {e}", 
                        level="WARNING"
                    )
                    print(f"警告: 无法保存结果到 {current_excel_file}！原因: 权限拒绝。尝试 {attempt + 1}/{MAX_SAVE_RETRIES}。等待 {RETRY_DELAY_SECONDS} 秒后重试...")
                    time.sleep(RETRY_DELAY_SECONDS) # 等待一段时间后重试
                except Exception as e: # 捕获其他保存错误
                    scan_log_manager.write_log(
                        f"错误: 将扫描结果保存到 '{normalize_drive_letter(str(current_excel_file))}' 失败: {e} (尝试 {attempt + 1}/{MAX_SAVE_RETRIES})", 
                        level="ERROR"
                    )
                    print(f"错误: 无法保存结果到 {current_excel_file}. 错误: {e}")
                    # 对于非权限错误，可能没有必要重试，直接跳出并转到备用
                    break 
            
            if not save_successful: # 如果所有重试都失败了
                scan_log_manager.write_log(
                    f"严重警告: 经过 {MAX_SAVE_RETRIES} 次尝试后，仍无法将扫描结果保存到 '{normalize_drive_letter(str(current_excel_file))}'。尝试保存到备用位置。", 
                    level="CRITICAL"
                )
                print(f"严重警告: 经过 {MAX_SAVE_RETRIES} 次尝试后，仍无法将扫描结果保存到 {current_excel_file}。尝试保存到备用位置...")
                try:
                    wb.save(str(fallback_excel_file))
                    scan_log_manager.write_log(f"成功将扫描结果保存到备用位置: {normalize_drive_letter(str(fallback_excel_file))}", level="WARNING")
                    print(f"成功将扫描结果保存到备用位置: {fallback_excel_file}")
                    actual_result_file_path = fallback_excel_file # 更新实际保存路径
                except Exception as fallback_e:
                    scan_log_manager.write_log(
                        f"致命错误: 尝试将扫描结果保存到备用位置 '{normalize_drive_letter(str(fallback_excel_file))}' 也失败了！错误: {fallback_e}", 
                        level="CRITICAL"
                    )
                    print(f"致命错误: 无法保存结果到任何位置！错误: {fallback_e}")
                    actual_result_file_path = Path("N/A_SAVE_FAILED") # 标记保存失败
            # --- 修改结束 ---

            # 将本次扫描结果添加到内存中的历史记录
            # 确保传递的是实际保存成功的路径，如果失败则传入None或标记
            history_manager.add_history_entry(
                folder_path, total_files, found_txt_count, not_found_txt_count, actual_result_file_path, current_scan_log_file
            )
            main_log_manager.write_log(f"本次扫描历史记录已成功添加至内存。", level="INFO")

            # 自动打开本次扫描的Excel、Log文件 
            files_to_open_this_scan = [current_scan_log_file]
            if actual_result_file_path.exists(): # 只有当实际结果文件存在时才尝试打开
                files_to_open_this_scan.append(actual_result_file_path)
            
            open_output_files_automatically(files_to_open_this_scan, main_log_manager)

        except Exception as e:
            main_log_manager.write_log(f"处理文件夹 {normalize_drive_letter(str(folder_path))} 时发生错误: {e}", level="ERROR")
            print(f"处理文件夹 {folder_path} 时发生错误。错误: {e}")
        finally:
            scan_log_manager.close() 
            main_log_manager.write_log(f"--- 完成处理文件夹: {normalize_drive_letter(str(folder_path))} ---\n", level="INFO")
            print(f"--- 完成处理文件夹: {folder_path} ---\n")

    # 所有文件夹处理完毕后，将内存中的历史记录保存到最终的Excel文件
    main_log_manager.write_log(f"所有扫描任务完成，开始将历史记录保存到最终的Excel文件: {normalize_drive_letter(str(final_history_excel_path))}", level="INFO")
    main_log_manager.write_log(f"准备保存 {len(history_manager.history_data)} 条历史记录到Excel。", level="INFO") # 确认即将保存的数量
    main_log_manager.write_log(f"最初在 '{normalize_drive_letter(str(batch_file_path))}' 中检测到 {len(folders_to_scan)} 条有效地址。", level="INFO") # 再次提示最初检测到的有效地址数量

    save_history_success = history_manager.save_history_to_excel()

    # 最终需要自动打开的文件列表
    final_files_to_open_at_end = []

    if save_history_success:
        main_log_manager.write_log("历史记录已成功保存到Excel。", level="INFO")
        
        # [RESTORE] 恢复将历史记录复制到缓存文件夹的逻辑
        if create_directory_if_not_exists(cache_folder_path, main_log_manager):
            current_timestamp_for_cache = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            cached_history_file_name = f"scan_history_cached_{current_timestamp_for_cache}.xlsx" # 区分命名
            cached_history_file_path = cache_folder_path / cached_history_file_name

            main_log_manager.write_log(f"开始复制历史记录到缓存文件夹: {normalize_drive_letter(str(cached_history_file_path))}", level="INFO")
            copy_cache_success = copy_file(final_history_excel_path, cached_history_file_path, main_log_manager)
            
            if copy_cache_success:
                main_log_manager.write_log("历史记录已成功复制到缓存文件夹。", level="INFO")
                final_files_to_open_at_end.append(cached_history_file_path) # 如果复制成功，则打开缓存文件
            else:
                main_log_manager.write_log("历史记录复制到缓存文件夹失败。", level="ERROR")
        else:
            main_log_manager.write_log(f"无法创建缓存文件夹: {normalize_drive_letter(str(cache_folder_path))}，将无法复制历史记录。", level="ERROR")
            
    else:
        main_log_manager.write_log("历史记录保存到Excel失败，将不会自动打开历史Excel文件。", level="ERROR")

    # 将错误和警告日志文件添加到需要自动打开的列表
    if error_warning_log_manager.log_file_path and error_warning_log_manager.log_file_path.exists():
        final_files_to_open_at_end.append(error_warning_log_manager.log_file_path)
    else:
        main_log_manager.write_log("警告: 错误和警告日志文件不存在或路径无效，无法自动打开。", level="WARNING")

    # 最终的自动打开操作
    open_output_files_automatically(final_files_to_open_at_end, main_log_manager)

    main_log_manager.write_log("所有文件夹处理完毕，程序即将退出。", level="INFO")
    print("所有文件夹处理完毕，程序即将退出。")
    main_log_manager.close() # 关闭主日志文件句柄
    error_warning_log_manager.close() # 关闭错误日志文件句柄

if __name__ == "__main__":
    main()