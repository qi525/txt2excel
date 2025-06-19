# InterrogateText2Xlsx7.0_utils.py
import os
import sys
import datetime
import shutil
import subprocess
import re # 导入re模块用于正则表达式
import time # 新增：导入time模块用于延迟重试
from pathlib import Path
from collections import defaultdict
from typing import Tuple, Dict, Optional, Set, List, Any
import hashlib # 新增：导入hashlib用于生成文件夹名的哈希值

# openpyxl 相关的导入
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill
from openpyxl.styles.colors import Color # 导入Color
from openpyxl.worksheet.worksheet import Worksheet # 用于类型提示

# ... (保留其他原有的导入，如 os, sys, shutil, subprocess, re, time, pathlib.Path, collections.defaultdict, typing 相关的 Tuple, Dict, Optional, Set, List, Any, hashlib, openpyxl 相关的导入) ...

# 移除 normalize_drive_letter 函数定义（因为它已移至 my_logger.py）

# 新增：从 my_logger 模块导入 logger (如果 utils.py 中的函数需要 logger 作为类型提示)
# 如果 utils.py 中没有函数直接使用 logger 作为类型提示，可以移除这行
# from my_logger import logger 

from loguru import logger # 新增行 1

# 新增：从 my_logger 导入 normalize_drive_letter
from my_logger import normalize_drive_letter # 新增行 2
# --- Configuration ---
HISTORY_FOLDER_NAME = "运行历史记录"
HISTORY_EXCEL_NAME = "scan_history.xlsx"
OUTPUT_FOLDER_NAME = "反推记录"
CACHE_FOLDER_NAME = "cache"

# 全局或常量定义超链接字体样式
from openpyxl.styles import Font # 确保这里导入了Font
HYPERLINK_FONT = Font(color="0000FF", underline="single")
# 全局常量：所有Excel列的固定宽度
FIXED_COLUMN_WIDTH = 20

# 新增：文件保存重试参数
MAX_SAVE_RETRIES = 5
RETRY_DELAY_SECONDS = 2

# ... (其他辅助函数和常量，除了 normalize_drive_letter) ...

# 例如，如果 validate_directory 函数仍然在 utils.py 中，并且它也需要 normalize_drive_letter
# 那么它也需要从 my_logger.py 中导入 normalize_drive_letter
# 或者，更好的做法是，如果 normalize_drive_letter 是一个非常通用的工具，可以放到一个独立的 'common_utils.py' 文件中
# 然后 utils.py 和 my_logger.py 都从 'common_utils.py' 导入它。
# 但为了快速解决当前问题，移动到 my_logger.py 是最直接的。

# 新增：文件保存重试参数
MAX_SAVE_RETRIES = 5  # 最大重试次数
RETRY_DELAY_SECONDS = 0.5  # 每次重试之间的延迟 (秒)

# --- Utility Function to Normalize Drive Letter ---
def normalize_drive_letter(path_str: str) -> str:
    """
    如果路径以驱动器号开头，将其转换为大写。
    例如: c:\\test -> C:\\test
    """
    if sys.platform.startswith('win') and len(path_str) >= 2 and path_str[1] == ':':
        return path_str[0].upper() + path_str[1:]
    return path_str

# --- NEW FUNCTION: Generate a safe and identifiable folder prefix for filenames ---
def generate_folder_prefix(folder_path: Path) -> str:
    """
    根据文件夹路径生成一个安全且可识别的前缀，用于文件名。
    原理：
        1. 获取文件夹的basename（即文件夹本身的名称）。
        2. 如果 basename 包含中文或特殊字符，为了确保文件名在各种文件系统中的兼容性，
           我们使用该 basename 的MD5哈希值的前8位作为唯一标识。
        3. 如果 basename 只包含ASCII字符（数字、字母、下划线、短横线），
           则直接使用 basename。
        4. 最终前缀会限制长度，避免文件名过长。
    Args:
        folder_path (Path): 文件夹的Path对象。
    Returns:
        str: 一个安全且短小的字符串，用于作为文件名前缀。
    """
    folder_name = folder_path.name
    # 检查是否包含非ASCII字符（例如中文），或者其他不适合作为文件名的字符
    if not re.fullmatch(r'[\w.-]+', folder_name): # 允许字母、数字、下划线、点、短横线
        # 如果包含特殊字符或中文，则使用哈希值
        return hashlib.md5(folder_name.encode('utf-8')).hexdigest()[:8]
    else:
        # 否则，使用文件夹名，并限制长度，防止文件名过长
        return folder_name[:30] # 限制为30个字符，避免过长

# --- logger Class ---
# class logger:
    """
    负责程序的日志记录。
    """
    def __init__(self, log_directory: Path, log_file_name: str = None, 
                 error_logger_obj: Optional['logger'] = None,
                 is_error_logger_obj: bool = False): # 新增参数：标记是否为错误日志管理器
        self.log_directory = log_directory
        self.log_file_path = None
        self.file_handle = None
        self.error_logger_obj = error_logger_obj
        self.is_error_logger_obj = is_error_logger_obj # 新增：标记是否为错误日志管理器
        self._is_initialized = False # 新增：标记日志文件是否已实际打开

        # 尝试创建日志目录（对于所有logger实例，目录都应该存在）
        try:
            if not self.log_directory.exists():
                os.makedirs(self.log_directory)
                # 只有非错误日志管理器才打印创建目录的INFO日志到控制台
                if not self.is_error_logger_obj:
                    print(f"已创建日志文件夹: {normalize_drive_letter(str(self.log_directory))}")
                    # 对于主日志，可以在目录创建后立即写入日志，但不使用to_error_log避免循环
                    self.write_log(f"已创建日志文件夹: {normalize_drive_letter(str(self.log_directory))}", level="INFO", to_error_log=False)
        except Exception as e:
            print(f"关键错误: 无法创建日志文件夹 {normalize_drive_letter(str(self.log_directory))}. 日志将仅打印到控制台. 错误: {e}")
            # 对于主日志，可以在目录创建失败后立即写入日志，但不使用to_error_log避免循环
            self.write_log(f"关键错误: 无法创建日志文件夹 {normalize_drive_letter(str(self.log_directory))}. 日志将仅打印到控制台. 错误: {e}", level="CRITICAL", to_error_log=False)
            self.log_directory = None

        if self.log_directory: # 如果目录创建成功
            if log_file_name is None:
                self.log_file_path = self.log_directory / f"main_program_log_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.txt"
            else:
                self.log_file_path = self.log_directory / log_file_name
        else: # 目录创建失败，所有日志都将仅打印到控制台
            self.log_file_path = None

    def _ensure_log_file_open(self):
        """
        确保日志文件已打开。如果未打开，则尝试打开并设置 _is_initialized。
        """
        if self.file_handle is None and self.log_file_path:
            try:
                self.file_handle = open(self.log_file_path, 'a', encoding='utf-8')
                self._is_initialized = True # 标记为已初始化
                # 只有非错误日志管理器才打印文件打开的INFO日志到控制台和自身文件
                if not self.is_error_logger_obj:
                    print(f"日志文件已打开: {normalize_drive_letter(str(self.log_file_path))}")
                    # 避免写入自身，否则会进入无限循环
                    self.file_handle.write(f"[{datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] [INFO] 日志文件已打开: {normalize_drive_letter(str(self.log_file_path))}\n")
                    self.file_handle.flush() # 立即写入
            except Exception as e:
                print(f"关键错误: 无法打开日志文件 {normalize_drive_letter(str(self.log_file_path))}. 所有后续日志将仅打印到控制台. 错误: {e}")
                self.file_handle = None # 无法打开，设为None
                # 对于自身打开失败，不应该再通过self.write_log去写入，否则可能陷入循环
                # 这里直接打印到控制台并通知调用方（如果有的话）
                # 原始代码中的自我写入会导致无限递归，所以这里直接依赖print

    def write_log(self, message: str, level: str = "INFO", to_file_only: bool = False, to_error_log: bool = True):
        """
        写入日志信息到文件，如果文件句柄无效则打印到控制台。
        Args:
            message (str): 日志消息。
            level (str): 日志级别 (INFO, WARNING, ERROR, CRITICAL).
            to_file_only (bool): 如果为True，则只写入文件，不打印到控制台。
            to_error_log (bool): 如果为True且存在error_logger_obj，则将WARNING/ERROR/CRITICAL日志写入错误日志。
        """
        timestamp = datetime.datetime.now().strftime("[%Y-%m-%d %H:%M:%S]")
        log_message = f"{timestamp} [{level}] {message}"
        
        # 将WARNING, ERROR, CRITICAL 级别的日志写入单独的错误日志文件
        # 并且确保不是错误日志管理器自己在给自己写（避免循环）
        if to_error_log and self.error_logger_obj and self.error_logger_obj is not self and level in ["WARNING", "ERROR", "CRITICAL"]:
            # 只有当错误日志管理器接收到实际的错误或警告时才触发其文件的打开和写入
            self.error_logger_obj._ensure_log_file_open() # 确保错误日志文件已打开
            if self.error_logger_obj.file_handle: # 再次检查句柄是否有效
                try:
                    self.error_logger_obj.file_handle.write(log_message + "\n")
                    self.error_logger_obj.file_handle.flush()
                except Exception as e:
                    # 如果错误日志管理器写入失败，则打印到控制台，不应再次尝试通过logger写入
                    print(f"关键错误: 写入错误日志文件 {normalize_drive_letter(str(self.error_logger_obj.log_file_path))} 失败. 消息: {message}. 错误: {e}")

        # 对于当前logger实例，只有在文件被实际打开后才写入文件
        # 或者如果是错误日志管理器，并且是WARNING/ERROR/CRITICAL级别才尝试打开并写入
        should_write_to_this_file = False
        if not self.is_error_logger_obj: # 对于主日志管理器和扫描日志管理器
            self._ensure_log_file_open() # 确保文件已打开
            should_write_to_this_file = True
        elif self.is_error_logger_obj and level in ["WARNING", "ERROR", "CRITICAL"]: # 对于错误日志管理器，只有这些级别才写入
            self._ensure_log_file_open() # 确保文件已打开
            should_write_to_this_file = True

        if should_write_to_this_file and self.file_handle:
            try:
                self.file_handle.write(log_message + "\n")
                self.file_handle.flush() # 立即将缓冲区内容写入文件
                if not to_file_only:
                    print(log_message)
            except Exception as e:
                # 写入失败，打印到控制台
                print(f"关键错误: 写入日志文件 {normalize_drive_letter(str(self.log_file_path))} 失败. 消息: {message}. 错误: {e}")
                if self.file_handle:
                    self.file_handle.close()
                self.file_handle = None
                print(f"日志消息重定向到控制台: {log_message}") # 失败后打印到控制台
        else:
            # 如果不应该写入文件（例如错误日志管理器未收到错误级别日志），或者文件句柄无效，则打印到控制台
            if not to_file_only:
                print(log_message)

    def close(self):
        """
        关闭日志文件句柄。
        """
        if self.file_handle:
            try:
                self.file_handle.close()
                self.file_handle = None
                self._is_initialized = False # 重置初始化状态
                # 只有非错误日志管理器才打印关闭信息
                if not self.is_error_logger_obj:
                    print(f"日志文件已关闭: {normalize_drive_letter(str(self.log_file_path))}")
            except Exception as e:
                print(f"关闭日志文件 {normalize_drive_letter(str(self.log_file_path))} 失败. 错误: {e}")
                # 避免循环调用，直接打印
                # self.write_log(f"关闭日志文件 {normalize_drive_letter(str(self.log_file_path))} 失败. 错误: {e}", level="ERROR", to_file_only=True, to_error_log=False)

    def __del__(self):
        """
        析构函数，确保在对象被销毁时关闭文件句柄。
        """
        self.close()

# --- NEW FUNCTION: Set Hyperlink and Style ---
def set_hyperlink_and_style(
    cell, 
    location: Optional[str], # location 现在可以是 Optional[str]
    display_text: str, 
    logger_obj: logger, 
    source_description: str = "未知源"
):
    """
    封装设置单元格超链接和样式的逻辑。
    Args:
        cell: openpyxl 单元格对象。
        location (Optional[str]): 超链接指向的实际位置（文件路径或URL）。如果为None或空字符串，则不设置超链接。
        display_text (str): 在单元格中显示的文本。
        logger_obj (logger): 日志管理器实例。
        source_description (str): 描述超链接来源，用于日志记录。
    """
    try:
        cell.value = display_text # 首先设置单元格显示文本
        
        # 只有当 location 不为 None 且不为空时才设置超链接
        if location: # 检查 location 是否有效
            cell.hyperlink = location # 然后设置超链接目标
            cell.font = HYPERLINK_FONT # 最后应用预定义的超链接字体样式
            logger_obj.info(
                f"成功设置超链接和样式 for '{display_text}' (Location: '{location}', Source: {source_description})", 
                level="DEBUG", to_file_only=True
            )
        else:
            # 如果没有 location，确保不设置超链接，并移除可能的超链接样式
            cell.hyperlink = None 
            cell.font = Font(color="000000") # 恢复默认字体颜色，去除下划线
            # 这条日志保留，因为仍然是提示没有设置超链接，但级别可以低一些
            logger_obj.info(
                f"未为 '{display_text}' (Source: {source_description}) 设置超链接，因为location无效或为空。", 
                level="INFO", to_file_only=True
            )

    except Exception as e:
        logger_obj.error(
            f"错误: 无法为单元格设置超链接或样式 for '{display_text}' (Location: '{location}', Source: {source_description}). 错误: {e}"
        )
        # 即使出错，也要确保单元格值被设置，即使没有超链接样式
        cell.value = display_text

# --- NEW FUNCTION: Set Fixed Column Widths for a Worksheet ---
def set_fixed_column_widths(worksheet: Worksheet, width: int, logger_obj: logger):
    """
    为给定工作表的所有列设置固定宽度。
    Args:
        worksheet (Worksheet): openpyxl 工作表对象。
        width (int): 要设置的列宽。
        logger_obj (logger): 日志管理器实例。
    """
    try:
        for col_idx in range(1, worksheet.max_column + 1): # 从1开始遍历所有列
            column_letter = get_column_letter(col_idx)
            worksheet.column_dimensions[column_letter].width = width
        #logger_obj.info(f"已为工作表 '{worksheet.title}' 设置所有列宽度为 {width}.", level="INFO", to_file_only=True)
        logger_obj.info(f"已为工作表 '{worksheet.title}' 设置所有列宽度为 {width}.") # 替换为 Loguru 的 info 方法，to_file_only 行为 Loguru 默认在 setup 时配置
    except Exception as e:
        #logger_obj.info(f"错误: 无法为工作表 '{worksheet.title}' 设置列宽: {e}")#error
        logger_obj.error(f"错误: 无法为工作表 '{worksheet.title}' 设置列宽: {e}") # 替换为 Loguru 的 error 方法
        print(f"错误: 无法为工作表 '{worksheet.title}' 设置列宽. 错误: {e}")

# --- HistoryManager (Excel Version) ---
class HistoryManager:
    """
    负责程序扫描历史记录的Excel文件存储。
    """
    def __init__(self, history_file_path: Path, logger_obj: logger):
        self.history_file_path = history_file_path
        self.logger_obj = logger_obj
        self.history_data: List[Dict[str, Any]] = [] # 存储内存中的历史记录
        self._load_history_from_excel()

    def _load_history_from_excel(self):
        """
        从Excel文件加载历史记录到内存。
        """
        self.history_data = []
        if not self.history_file_path.exists():
            #self.logger_obj.info(f"历史记录Excel文件不存在: {normalize_drive_letter(str(self.history_file_path))}. 将创建新文件。"
            self.logger_obj.info(f"历史记录Excel文件不存在: {normalize_drive_letter(str(self.history_file_path))}. 将创建新文件。") # 替换为 Loguru 的 info 方法
            return

        try:
            wb = load_workbook(str(self.history_file_path))
            if "扫描历史" in wb.sheetnames:
                ws = wb["扫描历史"]
                headers = [cell.value for cell in ws[1]] # 获取表头
                if not headers:
                    self.logger_obj.warning(f"历史记录Excel文件 '{normalize_drive_letter(str(self.history_file_path))}' 的 '扫描历史' 工作表为空，无历史记录可加载。")#warning
                    return

                # 确保表头符合预期，避免因旧文件格式导致的问题
                expected_headers = [
                    "扫描时间",
                    "文件夹路径",
                    "总文件数",
                    "找到TXT文件数",
                    "未找到TXT文件数",
                    "Log文件绝对路径",
                    "Log文件超链接", # 此列不再是实际数据，而是超链接显示文本
                    "结果XLSX文件绝对路径",
                    "结果XLSX文件超链接" # 此列不再是实际数据，而是超链接显示文本
                ]
                # 简化检查，只需要检查前5列和两个绝对路径列是否存在，因为超链接列是动态生成的
                if not all(h in headers for h in expected_headers[:6] + [expected_headers[7]]): # 检查前6个和第8个（结果XLSX绝对路径）
                    self.logger_obj.warning(f"历史记录Excel文件 '{normalize_drive_letter(str(self.history_file_path))}' 表头不匹配预期，可能无法完全加载所有历史记录。")#warning
                    # 继续尝试加载，但可能不完整

                for row_idx in range(2, ws.max_row + 1): # 从第二行开始读取数据
                    row_values = [cell.value for cell in ws[row_idx]]
                    # 将行数据映射到字典
                    entry = {}
                    for i, header in enumerate(headers):
                        if i < len(row_values):
                            entry[header] = row_values[i]
                        else:
                            entry[header] = None # 如果某些列没有值，则设置为None

                    # 仅保留我们需要的字段，并确保路径是Path对象
                    self.history_data.append({
                        'scan_time': entry.get("扫描时间"),
                        'folder_path': Path(entry.get("文件夹路径")) if entry.get("文件夹路径") else None,
                        'total_files': entry.get("总文件数"),
                        'found_txt_count': entry.get("找到TXT文件数"),
                        'not_found_txt_count': entry.get("未找到TXT文件数"),
                        'log_file_abs_path': Path(entry.get("Log文件绝对路径")) if entry.get("Log文件绝对路径") else None,
                        'result_xlsx_abs_path': Path(entry.get("结果XLSX文件绝对路径")) if entry.get("结果XLSX文件绝对路径") else None
                    })
            self.logger_obj.info(f"成功从历史记录Excel文件加载 {len(self.history_data)} 条历史记录。")
        except Exception as e:
            self.logger_obj.error(f"错误: 从历史记录Excel文件 {normalize_drive_letter(str(self.history_file_path))} 加载历史记录失败: {e}")#error
            self.history_data = [] # 加载失败则清空内存数据，避免脏数据

    def add_history_entry(self, folder_path: Path, total_scanned: int, found_txt_count: int,
                          not_found_txt_count: int, result_file_path: Path, log_file_path: Path):
        """
        向内存中的历史记录列表添加一条新的扫描历史记录。
        """
        scan_time = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        entry = {
            'scan_time': scan_time,
            'folder_path': folder_path,
            'total_files': total_scanned,
            'found_txt_count': found_txt_count,
            'not_found_txt_count': not_found_txt_count,
            'log_file_abs_path': log_file_path,
            'result_xlsx_abs_path': result_file_path
        }
        self.history_data.append(entry)
        self.logger_obj.info(f"历史记录成功添加至内存: 文件夹'{folder_path.name}'")

    def save_history_to_excel(self) -> bool:
        """
        将内存中的所有历史记录保存到Excel文件。
        Returns:
            bool: 如果保存成功返回True，否则返回False。
        """
        self.logger_obj.info(f"开始将内存中的历史记录保存到Excel: {normalize_drive_letter(str(self.history_file_path))}")

        # 尝试删除旧的历史Excel文件，以便重新写入
        if self.history_file_path.exists():
            try:
                os.remove(str(self.history_file_path))
                self.logger_obj.info(f"已删除旧的历史记录Excel文件: {normalize_drive_letter(str(self.history_file_path))}")
            except PermissionError as e: # 明确捕获权限错误
                self.logger_obj.warning(f"警告: 无法删除旧的历史记录Excel文件 {normalize_drive_letter(str(self.history_file_path))}. 可能文件被占用. 错误: {e}")#warning
                print(f"警告: 无法删除旧的历史记录Excel文件 {self.history_file_path}. 可能文件被占用. 错误: {e}")
                self.logger_obj.critical("无法覆盖旧的历史文件，历史记录将无法保存。请关闭Excel中打开的历史文件。")#critical
                return False # 删除失败，返回False
            except Exception as e:
                self.logger_obj.warning(f"警告: 删除旧的历史记录Excel文件时发生未知错误 {normalize_drive_letter(str(self.history_file_path))}. 错误: {e}")#warning
                print(f"警告: 删除旧的历史记录Excel文件时发生未知错误 {self.history_file_path}. 错误: {e}")
                self.logger_obj.critical("删除旧文件失败，历史记录将无法保存。")#critical
                return False # 删除失败，返回False


        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "扫描历史"

            # 定义Excel表头
            excel_headers = [
                "扫描时间",
                "文件夹路径",
                "总文件数",
                "找到TXT文件数",
                "未找到TXT文件数",
                "Log文件绝对路径",
                "Log文件超链接",
                "结果XLSX文件绝对路径",
                "结果XLSX文件超链接"
            ]
            ws.append(excel_headers)

            for entry in self.history_data:
                log_file_abs_path = entry['log_file_abs_path']
                result_xlsx_abs_path = entry['result_xlsx_abs_path']

                # Log文件超链接的显示文本和location
                log_link_display_text = "打开Log" 
                log_link_location = None
                if log_file_abs_path and log_file_abs_path.exists():
                    log_link_location = normalize_drive_letter(str(log_file_abs_path)).replace("\\", "/")
                    if not sys.platform.startswith('win'): 
                        log_link_location = f'file://{log_link_location}'
                else:
                    log_link_display_text = "Log文件不存在"


                # 结果XLSX文件超链接的显示文本和location
                result_link_display_text = "打开结果XLSX"
                result_link_location = None
                if result_xlsx_abs_path and result_xlsx_abs_path.exists():
                    result_link_location = normalize_drive_letter(str(result_xlsx_abs_path)).replace("\\", "/")
                    if not sys.platform.startswith('win'): 
                        result_link_location = f'file://{result_link_location}'
                else:
                    result_link_display_text = "结果XLSX文件不存在"
                
                row_data = [
                    entry['scan_time'],
                    str(entry['folder_path']),
                    entry['total_files'],
                    entry['found_txt_count'],
                    entry['not_found_txt_count'],
                    normalize_drive_letter(str(log_file_abs_path)) if log_file_abs_path else "N/A",      # Log文件绝对路径
                    log_link_display_text,       # Log文件超链接显示文本
                    normalize_drive_letter(str(result_xlsx_abs_path)) if result_xlsx_abs_path else "N/A",   # 结果XLSX文件绝对路径
                    result_link_display_text    # 结果XLSX文件超链接显示文本
                ]
                ws.append(row_data)

                # 获取新添加的行的单元格，并设置超链接
                new_row_idx = ws.max_row
                
                # Log文件超链接单元格 (第7列)
                log_link_cell = ws.cell(row=new_row_idx, column=7)
                set_hyperlink_and_style(
                    log_link_cell, 
                    log_link_location, # 传入可能为None的location
                    log_link_display_text, 
                    self.logger_obj, 
                    source_description=f"历史记录Log文件 (行: {new_row_idx})"
                )

                # 结果XLSX文件超链接单元格 (第9列)
                result_link_cell = ws.cell(row=new_row_idx, column=9)
                set_hyperlink_and_style(
                    result_link_cell, 
                    result_link_location, # 传入可能为None的location
                    result_link_display_text, 
                    self.logger_obj, 
                    source_description=f"历史记录结果XLSX文件 (行: {new_row_idx})"
                )
            
            # 设置所有列宽
            set_fixed_column_widths(ws, FIXED_COLUMN_WIDTH, self.logger_obj)
            
            wb.save(str(self.history_file_path))
            self.logger_obj.info(f"成功将历史记录保存到Excel: {normalize_drive_letter(str(self.history_file_path))}")
            return True

        except Exception as e:
            self.logger_obj.error(f"错误: 将历史记录保存到Excel失败 {normalize_drive_letter(str(self.history_file_path))}: {e}")#error
            print(f"错误: 将历史记录保存到Excel失败 {self.history_file_path}. 错误: {e}")
            return False

# --- File Operations ---
def validate_directory(path: Path, logger_obj: logger) -> bool:
    """
    验证给定的路径是否是一个存在的目录。
    """
    if not path.is_dir():
        if logger_obj:
            logger_obj.warning(f"验证失败: 目录不存在或不是一个目录: {normalize_drive_letter(str(path))}")#warning
        return False
    return True

# def create_directory_if_not_exists(directory_path: Path, logger_obj: Optional[logger]) -> bool:
def create_directory_if_not_exists(directory_path: Path, logger_obj) -> bool: # 更改参数名为 logger_obj，并移除 logger 类型提示
    """
    如果指定目录不存在，则创建它。
    Args:
        directory_path (Path): 要创建的目录路径。
        logger_obj (Optional[logger]): 日志管理器实例，可选。
    Returns:
        bool: 如果目录存在或成功创建，则返回True；否则返回False。
    """
    if not directory_path.exists():
        try:
            os.makedirs(directory_path)
            if logger_obj:
                #logger_obj.info(f"已创建目录: {normalize_drive_letter(str(directory_path))}"
                logger_obj.info(f"已创建目录: {normalize_drive_letter(str(directory_path))}") # 替换为 Loguru 的 info 方法
            return True
        except OSError as e:
            if logger_obj:
                #logger_obj.info(f"错误: 无法创建目录 {normalize_drive_letter(str(directory_path))}: {e}")#error
                logger_obj.error(f"创建目录失败 {normalize_drive_letter(str(directory_path))}: {e}") # 替换为 Loguru 的 error 方法
            print(f"错误: 无法创建文件夹 {directory_path}。错误: {e}")
            return False
    return True

def copy_file(source_path: Path, destination_path: Path, logger_obj:logger) -> bool:
    """
    复制文件从源路径到目标路径。
    增加对权限错误的捕获和提示。
    """
    if not source_path.exists():
        if logger_obj:
            logger_obj.error(f"错误: 源文件不存在，无法复制: {normalize_drive_letter(str(source_path))}")#error
        print(f"错误: 源文件不存在，无法复制: {source_path}")
        return False

    try:
        shutil.copy2(str(source_path), str(destination_path)) 
        if logger_obj:
            logger_obj.info(f"已复制 '{normalize_drive_letter(str(source_path))}' 到 '{normalize_drive_letter(str(destination_path))}'")
        return True
    except PermissionError as e:
        if logger_obj:
            logger_obj.info(
                f"权限错误: 复制文件从 '{normalize_drive_letter(str(source_path))}' 到 '{normalize_drive_letter(str(destination_path))}' 失败: {e}. 请确保目标文件未被其他程序（如Excel）占用。", 
                level="CRITICAL"
            )
        print(f"错误: 权限拒绝！无法复制文件到 '{destination_path}'。请确保该文件未被其他程序（如Excel）打开。错误: {e}")
        return False
    except Exception as e:
        if logger_obj:
            logger_obj.error(f"错误: 复制文件从 '{normalize_drive_letter(str(source_path))}' 到 '{normalize_drive_letter(str(destination_path))}' 失败: {e}")#error
        print(f"错误: 无法复制文件从 '{source_path}' 到 '{destination_path}'。错误: {e}")
        return False

def get_file_details(file_path: Path) -> Tuple[str, str]:
    """
    获取文件的名称（不含扩展名）和扩展名。
    """
    return file_path.stem, file_path.suffix

# --- Excel Utilities ---
def create_main_workbook():
    """
    创建主Excel工作簿，包含“匹配文件”和“未匹配文件”工作表。
    """
    wb = Workbook()
    
    # 移除默认创建的Sheet
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
        
    return wb

def setup_excel_sheets(wb: Workbook) -> Tuple[Worksheet, Worksheet, Worksheet]:
    """
    设置Excel工作表及其标题行。
    """
    # 匹配文件工作表
    ws_matched = wb.create_sheet("匹配文件", 0) # 插入到最前面
    ws_matched.append([
        "文件夹路径", "文件绝对路径", "文件链接", "文件扩展名", "TXT文件绝对路径",
        "TXT文件内容", "清洗后内容", "内容长度", "提示词类型", "找到TXT"
    ])

    # 未匹配文件工作表
    ws_no_txt = wb.create_sheet("未匹配文件", 1) # 插入到第二个
    ws_no_txt.append([
        "文件夹路径", "文件绝对路径", "文件链接", "文件扩展名", "找到TXT"
    ])

    # Tag词频统计工作表
    ws_tag_frequency = wb.create_sheet("Tag词频统计", 2) # 插入到第三个
    ws_tag_frequency.append(["Tag", "出现次数"])

    return ws_matched, ws_no_txt, ws_tag_frequency

# --- Data Processor (RESTORED FROM V4.0) ---
def detect_types(line: str, cleaned: str) -> str:
    """
    根据文本内容推断提示词类型。还原自 V4.0 版本。
    Args:
        line (str): 原始的txt文件内容。
        cleaned (str): 清洗后的txt文件内容。
    Returns:
        str: 识别到的提示词类型，用逗号分隔。
    """
    types = []
    lower_line = line.lower()
    # R18相关词汇
    if any(word in lower_line for word in [
        'sex', 'nude', 'pussy', 'penis', 'cum', 'nipples', 'vaginal',
        'cum_in_pussy', 'oral', 'rape', 'fellatio', 'facial', 'anus',
        'anal', 'ejaculation', 'gangbang', 'testicles', 'multiple_penises',
        'erection', 'handjob', 'cumdrip', 'pubic_hair', 'pussy_juice',
        'bukkake', 'clitoris', 'female_ejaculation', 'threesome',
        'doggystyle', 'sex_from_behind', 'cum_on_breasts', 'double_penetration',
        'anal_object_insertion', 'cunnilingus', 'triple_penetration',
        'paizuri', 'vaginal_object_insertion', 'imminent_rape', 'impregnation',
        'prone_bone', 'reverse_cowgirl_position', 'cum_inflation',
        'milking_machine', 'cumdump', 'anal_hair', 'futanari', 'glory_hole',
        'penis_on_face', 'licking_penis', 'breast_sucking', 'breast_squeeze', 'straddling'
    ]):
        types.append('R18')
    # boy类型
    if any(boy_word in lower_line for boy_word in ['1boy', '2boys', 'multiple_boys']):
        types.append('boy')
    # no_human类型
    if 'no_human' in lower_line:
        types.append('no_human')
    # furry 类型
    if any(word in lower_line for word in ['furry', 'animal_focus']):
        types.append('furry')
    # monochrome和greyscale类型
    if any(word in lower_line for word in ['monochrome', 'greyscale']):
        types.append('黑白原图')
    # 新增功能：检测"background"相关词汇并标记为“简单背景”类型
    if 'background' in lower_line:
        types.append('简单背景')
    
    # 如果没有检测到任何类型，返回 "N/A"
    if not types:
        return "N/A"
    return ','.join(types)

def clean_tags(line: str) -> Tuple[str, bool]:
    """
    清洗标签字符串。修改了对'censor'词的清理逻辑和'uncensored'的添加逻辑。
    Args:
        line (str): 原始的标签字符串。
    Returns:
        Tuple[str, bool]: 清洗后的字符串和是否含有敏感词的布尔值。
    """
    tags = [tag.strip() for tag in line.strip().split(',')]
    
    # 定义需要清洗掉的关键词，不包含 'uncensored'
    words_to_clean = ['censor', 'censored', 'monochrome', 'greyscale', 'furry', 'animal_focus', 'no_human', 'background']
    
    # 检查是否含有敏感词 (基于原始标签列表，因为这些词不应该被清洗掉，而是用于标记)
    has_sensitive = any(
        any(word in tag.lower() for word in [
            'nipple', 'pussy', 'penis', 'hetero', 'sex', 'anus', 'naked', 'explicit' # 增加一些常见的敏感词
        ])
        for tag in tags
    )
    
    # 过滤掉需要清洗的关键词
    cleaned_tags = []
    for tag in tags:
        # 如果是 'uncensored'，直接添加，不进行清洗
        if tag.lower() == 'uncensored':
            cleaned_tags.append(tag)
            continue
            
        # 只有当tag不包含任何words_to_clean中的词时才保留
        if not any(word in tag.lower() for word in words_to_clean):
            cleaned_tags.append(tag)

    # 如果检测到敏感词，则添加 'uncensored' 标记 
    # 确保只添加一次
    if has_sensitive and 'uncensored' not in [t.lower() for t in cleaned_tags]:
        cleaned_tags.append('uncensored')
    
    # 过滤掉空字符串，然后用逗号和空格连接
    cleaned_line = ', '.join([tag for tag in cleaned_tags if tag])
    return cleaned_line, has_sensitive


# --- Scanner ---
def scan_files_and_extract_data(
    base_folder_path: Path,
    ws_matched: Worksheet,
    ws_no_txt: Worksheet,
    logger_obj: logger#原logger

) -> Tuple[int, int, int, Dict[str, int]]:
    """
    扫描指定文件夹下的文件，查找匹配的TXT文件，提取数据并写入Excel。
    """
    total_files_scanned = 0
    found_txt_count = 0
    not_found_txt_count = 0
    tag_counts = defaultdict(int)

    # 新增：定义要跳过的文件夹名称
    skip_scan_folders = {'.bf'} # 使用集合以便快速查找
    # 新增：定义要跳过的文件扩展名
    skip_scan_extensions = {'.txt', '.xlsx', '.json', '.ini', '.db'} # 新增 .db 到跳过列表
    # 使用集合以便快速查找

    # 新增：用于收集文件扩展名信息
    all_extensions: Set[str] = set()
    skipped_extensions: Set[str] = set()

    logger_obj.info(f"开始扫描文件夹: {normalize_drive_letter(str(base_folder_path))}")

    try:
        for root_str, dirs, files in os.walk(base_folder_path):
            root = Path(root_str)

            should_skip_current_path = False
            for skip_folder_name in skip_scan_folders:
                if skip_folder_name in root.parts: 
                    should_skip_current_path = True
                    break
            
            if root.name in skip_scan_folders:
                should_skip_current_path = True

            if should_skip_current_path:
                logger_obj.info(f"跳过扫描文件夹及其子文件夹，因为它包含要跳过的名称: {normalize_drive_letter(str(root))}", level="INFO", to_file_only=True)
                dirs[:] = [] 
                continue 

            current_txt_files = {os.path.splitext(f)[0].lower(): root / f for f in files if f.lower().endswith('.txt')}

            for file_name in files:
                file_path = root / file_name
                file_stem, file_ext = get_file_details(file_path)

                all_extensions.add(file_ext.lower()) 

                if file_ext.lower() in skip_scan_extensions:
                    skipped_extensions.add(file_ext.lower()) 
                    continue 

                total_files_scanned += 1
                
                file_abs_path = file_path.resolve()
                
                # --- 主要修改点：根据文件是否存在来设置超链接的 location 和 display_text ---
                file_link_location = None
                file_link_text = "文件不存在" # 默认显示文本

                if file_abs_path.exists():
                    file_link_location = normalize_drive_letter(str(file_abs_path)).replace("\\", "/") 
                    if not sys.platform.startswith('win'): 
                        file_link_location = f'file://{file_link_location}'
                    file_link_text = str(file_abs_path.name) # 超链接显示文本为文件名

                else:
                    # 这条日志保留，因为是文件系统层面的缺失，但级别可以低一些
                    logger_obj.info(f"文件不存在，无法生成有效超链接: {normalize_drive_letter(str(file_abs_path))}", level="INFO", to_file_only=True) 
                    file_link_text = f"文件不存在: {file_abs_path.name}" # 提示文件不存在

                txt_content = ""
                cleaned_data = ""
                cleaned_data_length = 0
                prompt_type = "N/A"
                found_txt_flag = '否' 
                txt_absolute_path = "N/A" 

                if file_stem.lower() in current_txt_files:
                    txt_file_path = current_txt_files[file_stem.lower()]
                    txt_absolute_path = txt_file_path.resolve()
                    found_txt_flag = '是'
                    found_txt_count += 1
                    try:
                        with open(txt_file_path, 'r', encoding='utf-8') as f:
                            for line in f:
                                txt_content = line.strip()
                                cleaned_data, _ = clean_tags(txt_content) 
                                cleaned_data_length = len(cleaned_data)
                                prompt_type = detect_types(txt_content, cleaned_data) 
                                
                                for tag in cleaned_data.split(', '):
                                    if tag:
                                        tag_counts[tag.strip().lower()] += 1
                                break
                    except Exception as e:
                        logger_obj.error(f"错误: 读取或处理TXT文件 {normalize_drive_letter(str(txt_file_path))} 失败: {e}")#error
                        txt_content = f"Error reading TXT: {e}"
                        found_txt_flag = '否 (读取错误)'
                        not_found_txt_count += 1
                else:
                    logger_obj.info(f"未找到匹配的TXT文件: {normalize_drive_letter(str(file_path))}", level="INFO", to_file_only=True)
                    not_found_txt_count += 1
                
                current_row_data = [
                    str(root.resolve()),
                    str(file_abs_path),
                    file_link_text, # 这里直接传入 file_link_text 作为显示文本
                    file_ext,
                    str(txt_absolute_path),
                    txt_content,
                    cleaned_data,
                    cleaned_data_length,
                    prompt_type,
                    found_txt_flag 
                ]

                if found_txt_flag == '是': 
                    ws_matched.append(current_row_data)
                    link_cell = ws_matched.cell(row=ws_matched.max_row, column=3)
                    set_hyperlink_and_style(
                        link_cell, 
                        file_link_location, # 传入可能为None的location
                        file_link_text, # 传入已准备好的显示文本
                        logger_obj, 
                        source_description=f"匹配文件 (行: {ws_matched.max_row})"
                    )

                else:
                    current_row_data_no_txt = [
                        str(root.resolve()),
                        str(file_abs_path),
                        file_link_text, # 这里直接传入 file_link_text 作为显示文本
                        file_ext,
                        found_txt_flag
                    ]
                    ws_no_txt.append(current_row_data_no_txt)
                    link_cell = ws_no_txt.cell(row=ws_no_txt.max_row, column=3)
                    set_hyperlink_and_style(
                        link_cell, 
                        file_link_location, # 传入可能为None的location
                        file_link_text, # 传入已准备好的显示文本
                        logger_obj, 
                        source_description=f"未匹配文件 (行: {ws_no_txt.max_row})"
                    )

    except Exception as e:
        logger_obj.critical(f"致命错误: 扫描文件过程中发生意外错误 for folder {normalize_drive_letter(str(base_folder_path))}: {e}")#critical
        print(f"致命错误: 扫描文件过程中发生意外错误 for folder {base_folder_path}: {e}")
    
    logger_obj.info(f"文件夹 {normalize_drive_letter(str(base_folder_path))} 扫描完成. 总文件数: {total_files_scanned}, 找到TXT: {found_txt_count}, 未找到TXT: {not_found_txt_count}")
    
    logger_obj.info(f"\n--- 扫描文件类型概览 ---")
    if all_extensions:
        for ext in sorted(list(all_extensions)):
            status = "已处理"
            if ext in skipped_extensions:
                status = "已跳过"
            logger_obj.info(f"文件扩展名: '{ext}' - 状态: {status}")
    else:
        logger_obj.info("未扫描到任何文件扩展名。")
    logger_obj.info(f"--- 文件类型概览结束 ---\n")

    # 修正这里，返回 found_txt_count 和 not_found_txt_count
    return total_files_scanned, found_txt_count, not_found_txt_count, tag_counts

# --- New Function: Read Batch Paths ---
def read_batch_paths(batch_file_path: Path, logger_obj: logger#原logger
) -> List[Path]:
    """
    从 batchPath.txt 文件中读取需要扫描的文件夹路径列表。
    Args:
        batch_file_path (Path): batchPath.txt 文件的路径。
        logger_obj (logger): 日志管理器实例。
    Returns:
        List[Path]: 文件夹路径的列表。
    """
    folders = []
    if not batch_file_path.exists():
        logger_obj.error(f"错误: 批量路径文件 '{normalize_drive_letter(str(batch_file_path))}' 不存在。")#error
        print(f"错误: 批量路径文件 '{batch_file_path}' 不存在。")
        return folders
    try:
        with open(batch_file_path, 'r', encoding='utf-8') as f:
            for line in f:
                path_str = line.strip()
                if path_str and not path_str.startswith('#'): # 忽略空行和注释行
                    folder_path = Path(path_str)
                    if validate_directory(folder_path, logger_obj):
                        folders.append(folder_path)
                    else:
                        logger_obj.warning(f"警告: 批量路径文件中的路径无效或不存在，已跳过: {normalize_drive_letter(str(folder_path))}")#warning
        if not folders:
            logger_obj.warning(f"警告: 批量路径文件 '{normalize_drive_letter(str(batch_file_path))}' 中没有找到有效的文件夹路径。")#warning
    except Exception as e:
        logger_obj.critical(f"错误: 读取批量路径文件 '{normalize_drive_letter(str(batch_file_path))}' 失败: {e}")#critical
        print(f"错误: 读取批量路径文件 '{batch_file_path}' 失败。错误: {e}")
    return folders