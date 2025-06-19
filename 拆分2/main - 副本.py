# InterrogateText2Xlsx7.0_main.py

import os
import sys
import datetime
import shutil
import subprocess
import re
import time
from pathlib import Path
from collections import defaultdict
from typing import Tuple, Dict, Optional, Set, List, Any
import hashlib


# openpyxl 相关的导入
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill
from openpyxl.styles.colors import Color
from openpyxl.worksheet.worksheet import Worksheet

# 从 utils 导入除了 normalize_drive_letter 的其他常量和函数
from utils import (
    generate_folder_prefix,
    HistoryManager,
    create_directory_if_not_exists,
    copy_file,
    create_main_workbook,
    setup_excel_sheets,
    scan_files_and_extract_data,
    read_batch_paths,
    HISTORY_FOLDER_NAME,
    HISTORY_EXCEL_NAME,
    OUTPUT_FOLDER_NAME,
    CACHE_FOLDER_NAME,
    MAX_SAVE_RETRIES,
    RETRY_DELAY_SECONDS,
    FIXED_COLUMN_WIDTH
)

# --- BEGIN MODIFICATION 1/X (解决 NameError) ---
# 新增 Loguru 和 my_logger 的导入
# [MODIFIED] 添加 get_error_log_file_path 到导入列表
from my_logger import setup_logger, get_error_log_file_path 
from loguru import logger
from my_logger import normalize_drive_letter # 导入 normalize_drive_letter

# 将 script_dir 和 log_output_dir 移到全局作用域
script_dir = Path(os.path.dirname(os.path.abspath(__file__)))
log_output_dir = script_dir / "logs"

# 初始化 Loguru 日志系统
setup_logger(log_output_dir)
# --- END MODIFICATION 1/X ---


# --- Configuration ---
# ... (这一部分在你的原始main.py中可能存在，请保留原有内容) ...
# HYPERLINK_FONT = Font(color="0000FF", underline="single") # 如果Font不是从openpyxl导入并全局定义，则可能需要在这里定义
def open_output_files_automatically(files_to_open: List[Path], logger):
    """
    自动打开指定的文件。
    Args:
        files_to_open (List[Path]): 要打开的文件路径列表。
        logger (LogManager): 主日志管理器实例。
    """
    try:
        # 定义一个正则表达式来匹配文件路径中的日期时间戳 (YYYYMMDD_HHMMSS)
        # 允许文件名在时间戳前后有其他字符，例如 scan_results_20240101_123045.xlsx 或 scan_history_backup_20240101_123045.xlsx
        timestamp_pattern = re.compile(r'\d{8}_\d{6}')

        for file_path_to_open in files_to_open:
            if not file_path_to_open.exists():
                logger.warning(f"警告: 尝试打开不存在的文件: {normalize_drive_letter(str(file_path_to_open))}")
                continue

            # 检查文件名是否包含时间戳，或者是否为明确允许打开的临时文件（例如本次扫描结果）
            # 我们只允许打开包含时间戳的文件作为“快照”，不打开作为“数据库”的源文件
            # 这里的逻辑是：如果文件名包含时间戳，或者文件名是本次扫描生成的xlsx或log文件，则允许打开
            file_name = file_path_to_open.name
            
            is_allowed_to_open = False
            
            # 检查是否包含时间戳（例如 cache 中的历史文件，或者备份的历史文件）
            if timestamp_pattern.search(file_name):
                is_allowed_to_open = True
            
            # 检查是否是本次扫描生成的 Excel 或 Log 文件（这些文件本身就包含时间戳和/或文件夹前缀）
            # 新的命名约定：[folder_prefix]_scan_results_YYYYMMDD_HHMMSS.xlsx
            #             [folder_prefix]_scan_log_YYYYMMDD_HHMMSS.txt
            if re.match(r'.*_scan_results_\d{8}_\d{6}\.xlsx', file_name):
                is_allowed_to_open = True
            elif re.match(r'.*_scan_log_\d{8}_\d{6}\.txt', file_name):
                is_allowed_to_open = True
            elif file_name.startswith("error_warning_log_") and file_name.endswith(".txt"): # 允许打开错误日志
                is_allowed_to_open = True
            # [RESTORE] 恢复对以“scan_history_cached_”开头的文件的特殊允许，因为现在需要缓存备份
            elif file_name.startswith("scan_history_cached_") and file_name.endswith(".xlsx"):
                is_allowed_to_open = True
            # [NEW] 允许打开因权限问题而生成的备用结果文件
            elif file_name.startswith("FALLBACK_") and file_name.endswith(".xlsx"): # 修改为更通用的匹配FALLBACK_
                is_allowed_to_open = True


            if not is_allowed_to_open:
                logger.warning(f"拒绝自动打开没有时间戳或不符合命名约定的源文件: {normalize_drive_letter(str(file_path_to_open))}")
                print(f"拒绝自动打开没有时间戳或不符合命名约定的源文件: {file_path_to_open}")
                continue


            if sys.platform.startswith('win'): 
                subprocess.Popen(f'start "" "{file_path_to_open}"', shell=True) 
            elif sys.platform == 'darwin': 
                subprocess.Popen(['open', str(file_path_to_open)])
            else: 
                subprocess.Popen(['xdg-open', str(file_path_to_open)])
            
            print(f"自动打开: {file_path_to_open}")
            logger.info(f"自动打开: {normalize_drive_letter(str(file_path_to_open))}")

    except Exception as e:
        logger.error(f"错误: 自动打开文件失败. 错误: {e}")
        print(f"错误: 无法自动打开文件。请手动检查。错误: {e}")

# main.py

# ... (保留原有导入，包括 from loguru import logger 和 from my_logger import setup_logger, get_error_log_file_path) ...

# main.py

# ... (imports and global definitions) ...

def main():
    script_dir = Path(os.path.dirname(os.path.abspath(__file__)))
    log_output_dir = script_dir / "logs"
    history_folder_path = script_dir / HISTORY_FOLDER_NAME
    output_base_dir = script_dir / OUTPUT_FOLDER_NAME
    final_history_excel_path = history_folder_path / HISTORY_EXCEL_NAME
    cache_folder_path = script_dir / CACHE_FOLDER_NAME

    # 新增：用于存储错误和警告日志文件路径
    error_warning_log_file_path = get_error_log_file_path()

    logger.info(f"程序启动. 脚本目录: {normalize_drive_letter(str(script_dir))}")
    logger.info(f"日志文件将保存到: {normalize_drive_letter(str(log_output_dir))}")
    logger.info(f"错误和警告日志将写入到独立的日志文件 (仅当出现警告或错误时创建)。")

    if not create_directory_if_not_exists(history_folder_path, logger):
        logger.critical("致命错误: 无法创建历史记录文件夹，程序退出。")
        sys.exit(1)

    if not create_directory_if_not_exists(cache_folder_path, logger):
        logger.critical("致命错误: 无法创建缓存文件夹，程序退出。")
        sys.exit(1)

    if not create_directory_if_not_exists(output_base_dir, logger):
        logger.critical("致命错误: 无法创建输出文件夹 (反推记录)，程序退出。")
        sys.exit(1)

    history_manager = HistoryManager(final_history_excel_path, logger)

    batch_file_path = script_dir / "batchPath.txt"
    folders_to_scan = read_batch_paths(batch_file_path, logger)

    logger.info(f"在 '{normalize_drive_letter(str(batch_file_path))}' 中检测到 {len(folders_to_scan)} 条有效地址。")

    if not folders_to_scan:
        logger.info("没有找到要扫描的文件夹路径，程序终止。")
        print("没有找到要扫描的文件夹路径，程序终止。")
        logger.close()
        sys.exit(0)

    current_folder_log_sink_id: Optional[int] = None

    # --- START MODIFICATION ---
    # Initialize final_files_to_open_at_end here, at the top level of main()
    final_files_to_open_at_end = []
    # --- END MODIFICATION ---

    for folder_path in folders_to_scan:
        # ... (rest of the loop content, no changes needed here) ...
        logger.info(f"\n--- 开始处理文件夹: {normalize_drive_letter(str(folder_path))} ---")
        print(f"\n--- 开始处理文件夹: {folder_path} ---")

        scan_timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        folder_prefix = generate_folder_prefix(folder_path)

        current_scan_log_file = output_base_dir / f"{folder_prefix}_scan_log_{scan_timestamp}.txt"
        current_excel_file = output_base_dir / f"{folder_prefix}_scan_results_{scan_timestamp}.xlsx"

        fallback_excel_file = log_output_dir / f"FALLBACK_{folder_prefix}_scan_results_{scan_timestamp}.xlsx"

        try:
            current_folder_log_sink_id = logger.add(
                str(current_scan_log_file),
                level="INFO",
                rotation="5 MB",
                compression="zip",
                enqueue=True,
                encoding="utf-8",
                format="{time:YYYY-MM-DD HH:mm:ss.SSS} | {level: <8} | {message}"
            )
            logger.info(f"针对当前文件夹的扫描日志将写入: {normalize_drive_letter(str(current_scan_log_file))}")
        except Exception as e:
            logger.error(f"无法为文件夹 '{normalize_drive_letter(str(folder_path))}' 添加扫描日志文件 sink: {e}")
            current_folder_log_sink_id = None

        logger.info(f"开始扫描 {normalize_drive_letter(str(folder_path))}")

        try:
            wb = create_main_workbook()
            ws_matched, ws_no_txt, ws_tag_frequency = setup_excel_sheets(wb)

            total_files, found_txt_count, not_found_txt_count, tag_counts = scan_files_and_extract_data(
                folder_path, ws_matched, ws_no_txt, logger
            )

            sorted_tags = sorted(tag_counts.items(), key=lambda item: item[1], reverse=True)
            for tag, count in sorted_tags:
                ws_tag_frequency.append([tag, count])

            for worksheet in [ws_matched, ws_no_txt, ws_tag_frequency]:
                from utils import set_fixed_column_widths
                set_fixed_column_widths(worksheet, FIXED_COLUMN_WIDTH, logger)

            save_successful = False
            actual_result_file_path = Path("N/A_SAVE_FAILED")

            for attempt in range(MAX_SAVE_RETRIES):
                try:
                    wb.save(str(current_excel_file))
                    logger.info(f"扫描结果已保存到: {normalize_drive_letter(str(current_excel_file))} (尝试 {attempt + 1}/{MAX_SAVE_RETRIES})")
                    print(f"扫描结果已保存到: {current_excel_file}")
                    actual_result_file_path = current_excel_file
                    save_successful = True
                    break
                except PermissionError as e:
                    logger.warning(
                        f"警告: 无法将扫描结果保存到 '{normalize_drive_letter(str(current_excel_file))}'，原因: 权限拒绝！请确保该文件未被其他程序（如Excel）打开。尝试 {attempt + 1}/{MAX_SAVE_RETRIES}。错误: {e}"
                    )
                    print(f"警告: 无法保存结果到 {current_excel_file}！原因: 权限拒绝。尝试 {attempt + 1}/{MAX_SAVE_RETRIES}。等待 {RETRY_DELAY_SECONDS} 秒后重试...")
                    time.sleep(RETRY_DELAY_SECONDS)
                except Exception as e:
                    logger.error(
                        f"错误: 将扫描结果保存到 '{normalize_drive_letter(str(current_excel_file))}' 失败: {e} (尝试 {attempt + 1}/{MAX_SAVE_RETRIES})",
                        level="ERROR"
                    )
                    print(f"错误: 无法保存结果到 {current_excel_file}. 错误: {e}")
                    break

            if not save_successful:
                logger.critical(
                    f"严重警告: 经过 {MAX_SAVE_RETRIES} 次尝试后，仍无法将扫描结果保存到 '{normalize_drive_letter(str(current_excel_file))}'。尝试保存到备用位置。"
                )
                print(f"严重警告: 经过 {MAX_SAVE_RETRIES} 次尝试后，仍无法将扫描结果保存到 {current_excel_file}。尝试保存到备用位置...")
                try:
                    wb.save(str(fallback_excel_file))
                    logger.warning(f"成功将扫描结果保存到备用位置: {normalize_drive_letter(str(fallback_excel_file))}")
                    print(f"成功将扫描结果保存到备用位置: {fallback_excel_file}")
                    actual_result_file_path = fallback_excel_file
                except Exception as fallback_e:
                    logger.critical(
                        f"致命错误: 尝试将扫描结果保存到备用位置 '{normalize_drive_letter(str(fallback_excel_file))}' 也失败了！错误: {fallback_e}"
                    )
                    print(f"致命错误: 无法保存结果到任何位置！错误: {fallback_e}")
                    actual_result_file_path = Path("N/A_SAVE_FAILED")

            history_manager.add_history_entry(
                folder_path, total_files, found_txt_count, not_found_txt_count, actual_result_file_path, current_scan_log_file
            )
            logger.info(f"本次扫描历史记录已成功添加至内存。")

            files_to_open_this_scan = [current_scan_log_file]
            if actual_result_file_path.exists():
                files_to_open_this_scan.append(actual_result_file_path)

            open_output_files_automatically(files_to_open_this_scan, logger)

        except Exception as e:
            logger.error(f"处理文件夹 {normalize_drive_letter(str(folder_path))} 时发生错误: {e}")
            print(f"处理文件夹 {folder_path} 时发生错误。错误: {e}")
        finally:
            if current_folder_log_sink_id is not None:
                logger.remove(current_folder_log_sink_id)
                current_folder_log_sink_id = None
            logger.info(f"--- 完成处理文件夹: {normalize_drive_letter(str(folder_path))} ---\n")
            print(f"--- 完成处理文件夹: {folder_path} ---\n")

    logger.info(f"所有扫描任务完成，开始将历史记录保存到最终的Excel文件: {normalize_drive_letter(str(final_history_excel_path))}")
    logger.info(f"准备保存 {len(history_manager.history_data)} 条历史记录到Excel。")
    logger.info(f"最初在 '{normalize_drive_letter(str(batch_file_path))}' 中检测到 {len(folders_to_scan)} 条有效地址。")

    save_history_success = history_manager.save_history_to_excel()

    # The list is now initialized at the beginning of main()
    # final_files_to_open_at_end = [] # This line is moved/removed

    if save_history_success:
        logger.info("历史记录已成功保存到Excel。")
        if create_directory_if_not_exists(cache_folder_path, logger):
            current_timestamp_for_cache = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            cached_history_file_name = f"scan_history_cached_{current_timestamp_for_cache}.xlsx"
            cached_history_file_path = cache_folder_path / cached_history_file_name

            logger.info(f"开始复制历史记录到缓存文件夹: {normalize_drive_letter(str(cached_history_file_path))}")
            copy_cache_success = copy_file(final_history_excel_path, cached_history_file_path, logger)

            if copy_cache_success:
                logger.info("历史记录已成功复制到缓存文件夹。")
                final_files_to_open_at_end.append(cached_history_file_path)
            else:
                logger.error("历史记录复制到缓存文件夹失败。")
        else:
            logger.error(f"无法创建缓存文件夹: {normalize_drive_letter(str(cache_folder_path))}，将无法复制历史记录。")
    else:
        logger.error("历史记录保存到Excel失败，将不会自动打开历史Excel文件。")

    # Add error/warning log file to the list
    if error_warning_log_file_path and error_warning_log_file_path.exists():
        final_files_to_open_at_end.append(error_warning_log_file_path)
    else:
        logger.warning("警告: 错误和警告日志文件不存在或路径无效，无法自动打开。")

    open_output_files_automatically(final_files_to_open_at_end, logger)

    logger.info("所有文件夹处理完毕，程序即将退出。")
    print("所有文件夹处理完毕，程序即将退出。")
    logger.close()

if __name__ == "__main__":
    main()
