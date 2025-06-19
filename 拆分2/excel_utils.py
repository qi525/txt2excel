# excel_utils.py

from openpyxl.styles import Font, colors
from openpyxl.worksheet.worksheet import Worksheet
from loguru import logger
from my_logger import normalize_drive_letter
from typing import Optional # <--- 添加这一行
from openpyxl.utils import get_column_letter

# --- 常量 ---
# 定义超链接字体样式
HYPERLINK_FONT = Font(color="0000FF", underline="single")
# 定义固定列宽（以字符为单位）
FIXED_COLUMN_WIDTH = 20

# --- 辅助函数 ---
# 将 set_hyperlink_and_style 函数粘贴到这里
def set_hyperlink_and_style(
    cell, 
    location: Optional[str], # location 现在可以是 Optional[str]
    display_text: str, 
    logger_obj: logger, 
    source_description: str = "未知源"
):
    """
    封装设置单元格超链接和样式的逻辑。
    Args:
        cell: openpyxl 单元格对象。
        location (Optional[str]): 超链接指向的实际位置（文件路径或URL）。如果为None或空字符串，则不设置超链接。
        display_text (str): 在单元格中显示的文本。
        logger_obj (logger): 日志管理器实例。
        source_description (str): 描述超链接来源，用于日志记录。
    """
    try:
        cell.value = display_text # 首先设置单元格显示文本
        
        # 只有当 location 不为 None 且不为空时才设置超链接
        if location: # 检查 location 是否有效
            cell.hyperlink = location # 然后设置超链接目标
            cell.font = HYPERLINK_FONT # 最后应用预定义的超链接字体样式
            logger_obj.info(
                f"成功设置超链接和样式 for '{display_text}' (Location: '{location}', Source: {source_description})"
            )
        else:
            # 如果没有 location，确保不设置超链接，并移除可能的超链接样式
            cell.hyperlink = None 
            cell.font = Font(color="000000") # 恢复默认字体颜色，去除下划线
            # 这条日志保留，因为仍然是提示没有设置超链接，但级别可以低一些
            logger_obj.info(f"未为 '{display_text}' (Source: {source_description}) 设置超链接，因为location无效或为空。")

    except Exception as e:
        logger_obj.error(
            f"错误: 无法为单元格设置超链接或样式 for '{display_text}' (Location: '{location}', Source: {source_description}). 错误: {e}"
        )
        # 即使出错，也要确保单元格值被设置，即使没有超链接样式
        cell.value = display_text



# --- NEW FUNCTION: Set Fixed Column Widths for a Worksheet ---
def set_fixed_column_widths(worksheet: Worksheet, width: int, logger_obj: logger):
    """
    为给定工作表的所有列设置固定宽度。
    Args:
        worksheet (Worksheet): openpyxl 工作表对象。
        width (int): 要设置的列宽。
        logger_obj (logger): 日志管理器实例。
    """
    try:
        for col_idx in range(1, worksheet.max_column + 1): # 从1开始遍历所有列
            column_letter = get_column_letter(col_idx)
            worksheet.column_dimensions[column_letter].width = width
        logger_obj.info(f"已为工作表 '{worksheet.title}' 设置所有列宽度为 {width}.") # 替换为 Loguru 的 info 方法，to_file_only 行为 Loguru 默认在 setup 时配置
    except Exception as e:
        #logger_obj.info(f"错误: 无法为工作表 '{worksheet.title}' 设置列宽: {e}")#error
        logger_obj.error(f"错误: 无法为工作表 '{worksheet.title}' 设置列宽: {e}") # 替换为 Loguru 的 error 方法
        print(f"错误: 无法为工作表 '{worksheet.title}' 设置列宽. 错误: {e}")


# Scanner.py
# ... (原有导入不变)
from typing import List, Dict, Any, Tuple # 确保这些类型提示可用
# --- Scanner ---
def scan_files_and_extract_data(
    base_folder_path: Path,
    ws_matched: Worksheet,
    ws_no_txt: Worksheet,
    logger_obj: logger#原logger

) -> Tuple[int, int, int, Dict[str, int]]:
    """
    扫描指定文件夹下的文件，查找匹配的TXT文件，提取数据并写入Excel。
    """
    total_files_scanned = 0
    found_txt_count = 0
    not_found_txt_count = 0
    tag_counts = defaultdict(int)

    # 新增：定义要跳过的文件夹名称
    skip_scan_folders = {'.bf'} # 使用集合以便快速查找
    # 新增：定义要跳过的文件扩展名
    skip_scan_extensions = {'.txt', '.xlsx', '.json', '.ini', '.db'} # 新增 .db 到跳过列表
    # 使用集合以便快速查找

    # 新增：用于收集文件扩展名信息
    all_extensions: Set[str] = set()
    skipped_extensions: Set[str] = set()

    logger_obj.info(f"开始扫描文件夹: {normalize_drive_letter(str(base_folder_path))}")

    try:
        for root_str, dirs, files in os.walk(base_folder_path):
            root = Path(root_str)

            should_skip_current_path = False
            for skip_folder_name in skip_scan_folders:
                if skip_folder_name in root.parts: 
                    should_skip_current_path = True
                    break
            
            if root.name in skip_scan_folders:
                should_skip_current_path = True

            if should_skip_current_path:
                logger_obj.info(f"跳过扫描文件夹及其子文件夹，因为它包含要跳过的名称: {normalize_drive_letter(str(root))}")
                dirs[:] = [] 
                continue 

            current_txt_files = {os.path.splitext(f)[0].lower(): root / f for f in files if f.lower().endswith('.txt')}

            for file_name in files:
                file_path = root / file_name
                file_stem, file_ext = get_file_details(file_path)

                all_extensions.add(file_ext.lower()) 

                if file_ext.lower() in skip_scan_extensions:
                    skipped_extensions.add(file_ext.lower()) 
                    continue 

                total_files_scanned += 1
                
                file_abs_path = file_path.resolve()
                
                # --- 主要修改点：根据文件是否存在来设置超链接的 location 和 display_text ---
                file_link_location = None
                file_link_text = "文件不存在" # 默认显示文本

                if file_abs_path.exists():
                    file_link_location = normalize_drive_letter(str(file_abs_path)).replace("\\", "/") 
                    if not sys.platform.startswith('win'): 
                        file_link_location = f'file://{file_link_location}'
                    file_link_text = str(file_abs_path.name) # 超链接显示文本为文件名

                else:
                    # 这条日志保留，因为是文件系统层面的缺失，但级别可以低一些
                    logger_obj.info(f"文件不存在，无法生成有效超链接: {normalize_drive_letter(str(file_abs_path))}") 
                    file_link_text = f"文件不存在: {file_abs_path.name}" # 提示文件不存在

                txt_content = ""
                cleaned_data = ""
                cleaned_data_length = 0
                prompt_type = "N/A"
                found_txt_flag = '否' 
                txt_absolute_path = "N/A" 

                if file_stem.lower() in current_txt_files:
                    txt_file_path = current_txt_files[file_stem.lower()]
                    txt_absolute_path = txt_file_path.resolve()
                    found_txt_flag = '是'
                    found_txt_count += 1
                    try:
                        with open(txt_file_path, 'r', encoding='utf-8') as f:
                            for line in f:
                                txt_content = line.strip()
                                cleaned_data, _ = clean_tags(txt_content) 
                                cleaned_data_length = len(cleaned_data)
                                prompt_type = detect_types(txt_content, cleaned_data) 
                                
                                for tag in cleaned_data.split(', '):
                                    if tag:
                                        tag_counts[tag.strip().lower()] += 1
                                break
                    except Exception as e:
                        logger_obj.error(f"错误: 读取或处理TXT文件 {normalize_drive_letter(str(txt_file_path))} 失败: {e}")#error
                        txt_content = f"Error reading TXT: {e}"
                        found_txt_flag = '否 (读取错误)'
                        not_found_txt_count += 1
                else:
                    logger_obj.info(f"未找到匹配的TXT文件: {normalize_drive_letter(str(file_path))}")
                    not_found_txt_count += 1
                
                current_row_data = [
                    str(root.resolve()),
                    str(file_abs_path),
                    file_link_text, # 这里直接传入 file_link_text 作为显示文本
                    file_ext,
                    str(txt_absolute_path),
                    txt_content,
                    cleaned_data,
                    cleaned_data_length,
                    prompt_type,
                    found_txt_flag 
                ]

                if found_txt_flag == '是': 
                    ws_matched.append(current_row_data)
                    link_cell = ws_matched.cell(row=ws_matched.max_row, column=3)
                    set_hyperlink_and_style(
                        link_cell, 
                        file_link_location, # 传入可能为None的location
                        file_link_text, # 传入已准备好的显示文本
                        logger_obj, 
                        source_description=f"匹配文件 (行: {ws_matched.max_row})"
                    )

                else:
                    current_row_data_no_txt = [
                        str(root.resolve()),
                        str(file_abs_path),
                        file_link_text, # 这里直接传入 file_link_text 作为显示文本
                        file_ext,
                        found_txt_flag
                    ]
                    ws_no_txt.append(current_row_data_no_txt)
                    link_cell = ws_no_txt.cell(row=ws_no_txt.max_row, column=3)
                    set_hyperlink_and_style(
                        link_cell, 
                        file_link_location, # 传入可能为None的location
                        file_link_text, # 传入已准备好的显示文本
                        logger_obj, 
                        source_description=f"未匹配文件 (行: {ws_no_txt.max_row})"
                    )

    except Exception as e:
        logger_obj.critical(f"致命错误: 扫描文件过程中发生意外错误 for folder {normalize_drive_letter(str(base_folder_path))}: {e}")#critical
        print(f"致命错误: 扫描文件过程中发生意外错误 for folder {base_folder_path}: {e}")
    
    logger_obj.info(f"文件夹 {normalize_drive_letter(str(base_folder_path))} 扫描完成. 总文件数: {total_files_scanned}, 找到TXT: {found_txt_count}, 未找到TXT: {not_found_txt_count}")
    
    logger_obj.info(f"\n--- 扫描文件类型概览 ---")
    if all_extensions:
        for ext in sorted(list(all_extensions)):
            status = "已处理"
            if ext in skipped_extensions:
                status = "已跳过"
            logger_obj.info(f"文件扩展名: '{ext}' - 状态: {status}")
    else:
        logger_obj.info("未扫描到任何文件扩展名。")
    logger_obj.info(f"--- 文件类型概览结束 ---\n")

    # 修正这里，返回 found_txt_count 和 not_found_txt_count
    return total_files_scanned, found_txt_count, not_found_txt_count, tag_counts
