import os
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from openpyxl.worksheet.hyperlink import Hyperlink
import datetime
import shutil # 导入shutil模块用于文件复制
import subprocess # 用于跨平台打开文件
from collections import defaultdict # 新增：导入defaultdict用于词频统计

# 通过input询问用户要处理的文件夹路径
folder_path = input("请输入您要处理的文件夹路径: ")

# 验证路径是否存在
if not os.path.isdir(folder_path):
    print(f"错误: 您输入的路径 '{folder_path}' 不是一个有效的文件夹。程序将退出。")
    exit() # 直接退出程序

# 获取Python运行文件的目录
python_script_dir = os.path.dirname(os.path.abspath(__file__))

# 定义Python运行目录中的“反推历史记录”子文件夹路径
python_history_folder = os.path.join(python_script_dir, "反推历史记录")

# 确保“反推历史记录”文件夹存在
if not os.path.exists(python_history_folder):
    os.makedirs(python_history_folder)
    print(f"已创建Python运行目录下的“反推历史记录”文件夹: {python_history_folder}")


current_time = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")

# 定义主要保存位置 (Python运行目录下的“反推历史记录”子文件夹)
main_output_xlsx = os.path.join(python_history_folder, f'scan_results_{current_time}.xlsx')
main_log_file = os.path.join(python_history_folder, f'scan_log_{current_time}.txt') # log文件路径

# 定义历史记录文件路径 - 修改点：将scan_history.xlsx的存放位置改到python运行文件同级文件夹
history_xlsx_path = os.path.join(python_script_dir, 'scan_history.xlsx')

# 定义反推记录子文件夹路径 (作为次要保存位置)
output_folder = os.path.join(folder_path, "反推记录")
# 定义目标文件夹中的文件路径
target_output_xlsx = os.path.join(output_folder, f'scan_results_{current_time}.xlsx')
target_log_file = os.path.join(output_folder, f'scan_log_{current_time}.txt')

# 创建一个日志文件函数，现在它将主要写入到运行目录的log文件
def write_log(message):
    with open(main_log_file, 'a', encoding='utf-8') as f:
        f.write(f"{datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')} - {message}\n")

write_log("程序开始运行...") # 记录程序启动

wb = Workbook()
ws = wb.active
ws.title = "已匹配TXT文件" # 设置第一个工作表的标题

# 主工作表的列头
ws.append([
    '文件夹绝对路径',
    '文件绝对路径',
    '文件超链接',
    '文件后缀', # 新增列：文件后缀
    'TXT绝对路径',
    'TXT内容',
    '清洗后的数据',
    '清洗后的数据字数', # 新增列：清洗后的数据字数
    '提示词类型',
    '是否找到匹配TXT'
])

# 创建第二个工作表用于存放未找到TXT的文件
ws_no_txt = wb.create_sheet("未匹配TXT文件", 1) # 在第一个工作表后创建
ws_no_txt.append([
    '文件夹绝对路径',
    '文件绝对路径',
    '文件超链接',
    '文件后缀', # 新增列：文件后缀
    '是否找到匹配TXT'
])

# 新增：创建第三个工作表用于存放Tag词频统计
ws_tag_frequency = wb.create_sheet("Tag 词频统计", 2)
ws_tag_frequency.append([
    'Tag',
    '出现次数'
])

def detect_types(line, cleaned):
    types = []
    lower_line = line.lower()
    # R18相关词汇
    if any(word in lower_line for word in [
        'sex',  # 性行为
        'nude',  # 裸体
        'pussy',  # 阴户
        'penis',  # 阴茎
        'cum',  # 精液
        'nipples',  # 乳头
        'vaginal',  # 阴道的
        'cum_in_pussy',  # 精液入阴
        'oral',  # 口交
        'rape',  # 强奸
        'fellatio',  # 口交（舔阴茎）
        'facial',  # 面部射精
        'anus',  # 肛门
        'anal',  # 肛交
        'ejaculation',  # 射精
        'gangbang',  # 群交
        'testicles',  # 睾丸
        'multiple_penises',  # 多阴茎
        'erection',  # 勃起
        'handjob',  # 手淫
        'cumdrip',  # 精液滴落
        'pubic_hair',  # 阴毛
        'pussy_juice',  # 阴道分泌液
        'bukkake',  # 颜射
        'clitoris',  # 阴蒂
        'female_ejaculation',  # 女性射精
        'threesome',  # 三人性行为
        'doggystyle',  # 狗爬式（性交姿势）
        'sex_from_behind',  # 从后性交
        'cum_on_breasts',  # 精液在乳房上
        'double_penetration',  # 双重插入
        'anal_object_insertion',  # 肛门异物插入
        'cunnilingus',  # 口交（舔阴）
        'triple_penetration',  # 三重插入
        'paizuri',  # 乳交
        'vaginal_object_insertion',  # 阴道异物插入
        'imminent_rape',  # 即将强奸
        'impregnation',  # 受孕
        'prone_bone',  # 俯卧性交
        'reverse_cowgirl_position',  # 反向女上位
        'cum_inflation',  # 精液灌注
        'milking_machine',  # 榨精机
        'cumdump',  # 精液倾泻对象
        'anal_hair',  # 肛毛
        'futanari',  # 扶他
        'glory_hole',  # 窥淫孔
        'penis_on_face',  # 阴茎在脸上
        'licking_penis',  # 舔阴茎
        'breast_sucking',  # 吸吮乳房
        'breast_squeeze',  # 挤压乳房
        'straddling'  # 跨骑
    ]):
        types.append('R18')
    # boy类型
    if any(boy_word in lower_line for boy_word in ['1boy', '2boys', 'multiple_boys']):
        types.append('boy')
    # no_human类型
    if 'no_human' in lower_line:
        types.append('no_human')
    # furry 类型
    if any(word in lower_line for word in ['furry', 'animal_focus']):
        types.append('furry')
    # monochrome和greyscale类型
    if any(word in lower_line for word in ['monochrome', 'greyscale']):
        types.append('黑白原图')
    # 新增功能：检测"background"相关词汇并标记为“简单背景”类型
    if 'background' in lower_line:
        types.append('简单背景')
    return ','.join(types)

def clean_tags(line):
    tags = [tag.strip() for tag in line.strip().split(',')]
    # 定义需要清洗掉的关键词
    words_to_clean = ['censor', 'monochrome', 'greyscale', 'furry', 'animal_focus', 'no_human', 'background']
    
    # 根据words_to_clean列表过滤tag
    cleaned_tags = [
        tag for tag in tags 
        if not any(word in tag.lower() for word in words_to_clean)
    ]

    # 检查是否含有敏感词
    has_sensitive = any(
        any(word in tag.lower() for word in [
            'censor', 'nipple', 'pussy', 'penis', 'hetero', 'sex', 'anus'
        ])
        for tag in tags
    )
    if has_sensitive:
        cleaned_tags.append('uncensored')
    cleaned_line = ', '.join([tag for tag in cleaned_tags if tag])
    return cleaned_line, has_sensitive

# 新增计数器
total_files_scanned = 0
found_txt_count = 0
not_found_txt_count = 0
tag_counts = defaultdict(int) # 新增：用于存储tag词频

# 遍历所有文件，而不仅仅是图片
for root, dirs, files in os.walk(folder_path):
    # 将当前文件夹下的所有txt文件收集起来，方便查找
    current_txt_files = {os.path.splitext(f)[0].lower(): os.path.join(root, f) for f in files if f.lower().endswith('.txt')}

    for f_name in files:
        # 跳过txt文件本身，因为我们主要关注其他文件类型以及它们是否关联了txt
        if f_name.lower().endswith('.txt'):
            continue

        total_files_scanned += 1 # 统计总文件量

        file_path = os.path.join(root, f_name)
        file_name_without_ext, file_ext = os.path.splitext(f_name)
        file_ext = file_ext.lower() # 统一小写文件后缀

        txt_content = ''
        cleaned_data = ''
        prompt_type = ''
        txt_absolute_path = ''
        found_txt = '否' # 默认没有找到匹配的TXT文件
        cleaned_data_length = 0 # 新增：初始化清洗后的数据字数

        # 添加文件超链接
        file_abs_path = os.path.abspath(file_path)
        hyperlink_formula = f'=HYPERLINK("{file_abs_path}", "打开文件")'

        # 尝试查找同名的txt文件
        if file_name_without_ext.lower() in current_txt_files:
            txt_file_path = current_txt_files[file_name_without_ext.lower()]
            try:
                with open(txt_file_path, 'r', encoding='utf-8') as f:
                    # 假设一个txt文件只有一行内容，如果有多行，您可以根据需求修改
                    # 这里只读取第一行作为txt内容进行处理，如果txt文件有多个描述，需要调整逻辑
                    for line in f: # 遍历txt文件中的每一行
                        txt_content = line.strip() # 获取原始txt内容
                        cleaned_data, _ = clean_tags(txt_content) # 清洗数据
                        cleaned_data_length = len(cleaned_data) # 计算清洗后的数据字数
                        prompt_type = detect_types(txt_content, cleaned_data) # 检测提示词类型
                        txt_absolute_path = os.path.abspath(txt_file_path)
                        found_txt = '是'
                        found_txt_count += 1 # 统计成功匹配数量

                        # 新增：统计Tag词频
                        for tag in cleaned_data.split(', '):
                            if tag: # 确保tag不为空
                                tag_counts[tag.strip().lower()] += 1 # 统一小写并去除首尾空格

                        break # 只处理txt文件中的第一行，如果需要处理所有行，请移除此行
            except Exception as e:
                write_log(f"Error reading TXT file {txt_file_path}: {e}")
                txt_content = f"Error reading TXT: {e}" # 记录读取错误信息
                found_txt = '否 (读取错误)'
                not_found_txt_count += 1 # 统计失败匹配数量（读取错误也算失败）
        else:
            write_log(f"No matching TXT file found for: {file_path}")
            not_found_txt_count += 1 # 统计失败匹配数量

        # 根据是否找到TXT文件，将数据写入不同的工作表
        if found_txt == '是':
            ws.append([
                os.path.abspath(root),
                file_abs_path,
                hyperlink_formula,
                file_ext, # 新增列数据
                txt_absolute_path,
                txt_content,
                cleaned_data,
                cleaned_data_length, # 新增列数据：清洗后的数据字数
                prompt_type,
                found_txt
            ])
        else:
            ws_no_txt.append([
                os.path.abspath(root),
                file_abs_path,
                hyperlink_formula,
                file_ext, # 新增列数据
                found_txt
            ])

# 设置主工作表超链接列为蓝色字体
# 注意：超链接列现在是第4列（索引3）
for row in ws.iter_rows(min_row=2, min_col=3, max_col=3):
    for cell in row:
        cell.font = Font(color="0000FF", underline="single")

# 设置未匹配TXT工作表超链接列为蓝色字体
# 注意：超链接列现在是第3列（索引2）
for row in ws_no_txt.iter_rows(min_row=2, min_col=3, max_col=3):
    for cell in row:
        cell.font = Font(color="0000FF", underline="single")

# 新增：将Tag词频数据写入新的工作表
sorted_tags = sorted(tag_counts.items(), key=lambda item: item[1], reverse=True)
for tag, count in sorted_tags:
    ws_tag_frequency.append([tag, count])

# ====================================================================
# 优先保存到Python运行目录下的“反推历史记录”文件夹
# ====================================================================
try:
    wb.save(main_output_xlsx)
    print('合并完成，已保存至Python运行目录下的“反推历史记录”文件夹:', main_output_xlsx)
    write_log(f"Processing complete. Results saved to Python script history directory: {main_output_xlsx}")
    write_log(f"Log saved to Python script history directory: {main_log_file}")
except Exception as e:
    print(f"错误: 无法保存结果到Python运行目录下的“反推历史记录”文件夹 {main_output_xlsx}。错误: {e}")
    write_log(f"Error: Could not save results to Python script history directory {main_output_xlsx}. Error: {e}")
    print("程序将退出，请检查Python运行目录的写入权限。", )
    exit() # 如果连运行目录都无法写入，则直接退出

# ====================================================================
# 尝试保存一份到目标文件夹
# ====================================================================
try:
    # 检查并创建反推记录文件夹（如果之前没有创建的话，确保在尝试写入前存在）
    if not os.path.exists(output_folder):
        os.makedirs(output_folder)
        print(f"已创建输出文件夹: {output_folder}")

    # 复制已经生成在Python运行目录的XLSX文件到目标文件夹
    shutil.copy2(main_output_xlsx, target_output_xlsx)
    print(f"一份副本已保存至目标文件夹: {target_output_xlsx}")
    write_log(f"A copy of XLSX saved to target folder: {target_output_xlsx}")

    # 复制已经生成在Python运行目录的log文件到目标文件夹
    shutil.copy2(main_log_file, target_log_file)
    print(f"一份log副本已保存至目标文件夹: {target_log_file}")
    write_log(f"A copy of log saved to target folder: {target_log_file}")

except Exception as e:
    print(f"警告: 无法保存文件到目标文件夹 '{output_folder}'。错误: {e}")
    print("请检查目标文件夹的写入权限。Python运行目录已保存一份结果。")
    write_log(f"Warning: Could not save files to target folder '{output_folder}'. Error: {e}")
    write_log("Python script directory already contains a copy of results.")

# ====================================================================
# 写入历史记录
# ====================================================================
# 修改 update_history 函数，增加接收 main_output_xlsx 和 main_log_file 参数
def update_history(history_file_path, folder_path, total_scanned, found_count, not_found_count, xlsx_file_path, log_file_path):
    try:
        if os.path.exists(history_file_path):
            history_wb = load_workbook(history_file_path)
            history_ws = history_wb.active
            # 检查现有列头，如果少了，则可能需要处理（这里简化为直接添加新行，假定第一次运行会创建完整列头）
            if history_ws['F1'].value != 'Log文件绝对路径': # 简单检查是否存在新列头
                print("警告：检测到旧版历史记录文件格式，将追加新列头。建议手动检查或删除旧文件以生成完整新格式。")
                # 如果是旧文件，可以考虑插入列或者给出提示，这里为了简单直接追加新的表头行
                # 但是append会直接加到最后，这里更合理的是在初始化时统一
                # 由于默认是active sheet，所以如果文件存在且没有这些列，可能需要更复杂的处理
                # For simplicity, we assume the new columns will be added correctly or history.xlsx is new.
                pass 
        else:
            history_wb = Workbook()
            history_ws = history_wb.active
            history_ws.title = "扫描历史记录"
            # 增加新的列头
            history_ws.append([
                '运行时间',
                '分析目录',
                '总文件量',
                '成功匹配TXT数量',
                '失败匹配TXT数量',
                'Log文件绝对路径',        # 新增
                'Log文件超链接',         # 新增
                '结果XLSX文件绝对路径',    # 新增
                '结果XLSX文件超链接'       # 新增
            ])

        current_run_time = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        # 生成超链接公式
        log_hyperlink_formula = f'=HYPERLINK("{log_file_path}", "打开Log")'
        xlsx_hyperlink_formula = f'=HYPERLINK("{xlsx_file_path}", "打开结果XLSX")'

        # 增加新的数据列
        history_ws.append([
            current_run_time,
            folder_path,
            total_scanned,
            found_count,
            not_found_count,
            log_file_path,
            log_hyperlink_formula,
            xlsx_file_path,
            xlsx_hyperlink_formula
        ])

        # 设置新添加的超链接列的样式
        new_row_idx = history_ws.max_row
        # Log文件超链接列 (G列, 索引6)
        history_ws.cell(row=new_row_idx, column=7).font = Font(color="0000FF", underline="single")
        # 结果XLSX文件超链接列 (I列, 索引8)
        history_ws.cell(row=new_row_idx, column=9).font = Font(color="0000FF", underline="single")


        history_wb.save(history_file_path)
        print(f"历史记录已更新到: {history_file_path}")
        write_log(f"History record updated: {history_file_path}")
    except Exception as e:
        print(f"错误: 无法更新历史记录文件 {history_file_path}。错误: {e}")
        write_log(f"Error: Could not update history file {history_file_path}. Error: {e}")

# 调用更新历史记录函数，传入新参数
update_history(history_xlsx_path, folder_path, total_files_scanned, found_txt_count, not_found_txt_count, main_output_xlsx, main_log_file)


# ====================================================================
# 自动打开文件
# ====================================================================
# 定义缓存文件夹路径
cache_folder = r'C:\个人数据\pythonCode\反推图片信息\cache' # 使用原始字符串，避免反斜杠转义

# 确保缓存文件夹存在
if not os.path.exists(cache_folder):
    try:
        os.makedirs(cache_folder)
        print(f"已创建缓存文件夹: {cache_folder}")
        write_log(f"Created cache folder: {cache_folder}")
    except Exception as e:
        print(f"错误: 无法创建缓存文件夹 '{cache_folder}'。错误: {e}")
        write_log(f"Error: Could not create cache folder '{cache_folder}'. Error: {e}")

# 生成带时间戳的scan_history.xlsx缓存文件路径
history_cache_file_name = f"scan_history_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
history_cache_file_path = os.path.join(cache_folder, history_cache_file_name)


try:
    # 复制 scan_history.xlsx 到缓存文件夹
    if os.path.exists(history_xlsx_path):
        shutil.copy2(history_xlsx_path, history_cache_file_path)
        print(f"scan_history.xlsx 已复制到缓存: {history_cache_file_path}")
        write_log(f"scan_history.xlsx copied to cache: {history_cache_file_path}")
    else:
        print(f"警告: 原始历史记录文件不存在，无法复制到缓存: {history_xlsx_path}")
        write_log(f"Warning: Original history file does not exist, cannot copy to cache: {history_xlsx_path}")
        # 如果原始文件不存在，则直接尝试打开主输出文件和log文件，跳过历史记录缓存打开
        history_cache_file_path = None # 将路径设为None，后续不尝试打开

    # 使用subprocess.Popen以兼容更多系统
    if os.name == 'nt': # Windows
        os.startfile(main_output_xlsx)
        os.startfile(main_log_file)
        if history_cache_file_path: # 只有当成功复制了缓存文件时才尝试打开
            os.startfile(history_cache_file_path)
    elif os.uname().sysname == 'Darwin': # macOS
        subprocess.Popen(['open', main_output_xlsx])
        subprocess.Popen(['open', main_log_file])
        if history_cache_file_path:
            subprocess.Popen(['open', history_cache_file_path])
    else: # Linux/Unix
        subprocess.Popen(['xdg-open', main_output_xlsx])
        subprocess.Popen(['xdg-open', main_log_file])
        if history_cache_file_path:
            subprocess.Popen(['xdg-open', history_cache_file_path])

    print(f"Automatically opened: {main_output_xlsx}")
    print(f"Automatically opened: {main_log_file}")
    if history_cache_file_path:
        print(f"Automatically opened cached history: {history_cache_file_path}")
except Exception as e:
    print(f"Could not automatically open files or cache history. Please check manually. Error: {e}")
    write_log(f"Failed to auto-open files or cache history: {main_output_xlsx}, {main_log_file}. Error: {e}")
    if history_cache_file_path:
        write_log(f"Failed to auto-open cached history: {history_cache_file_path}. Error: {e}")