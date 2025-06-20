import os
import sys
from pathlib import Path
from openpyxl.worksheet.worksheet import Worksheet
from typing import Tuple, Dict, Optional, Set, List, Any

from collections import defaultdict

from file_system_utils import normalize_drive_letter,get_file_details
from tag_processing import clean_tags, detect_types
from excel_utilities import set_hyperlink_and_style

# --- 模块级别常量 ---
SKIP_SCAN_FOLDERS = {'.bf'}
SKIP_SCAN_EXTENSIONS = {'.txt', '.xlsx', '.json', '.ini', '.db'}

# 新增辅助函数：判断是否跳过路径
def _should_skip_path(path: Path, skip_folders: Set[str]) -> bool:
    """
    判断一个给定的路径是否应该被跳过扫描。
    如果路径的任何部分包含在 skip_folders 中，或者路径本身的名称在 skip_folders 中，则跳过。
    """
    for skip_folder_name in skip_folders:
        if skip_folder_name in path.parts:
            return True
        if path.name == skip_folder_name:
            return True
    return False

# 新增辅助函数：处理单个文件并提取数据
def _process_single_file(
    file_path: Path,
    current_txt_files: Dict[str, Path],
    logger_obj, # 移除类型提示
    tag_counts: defaultdict
) -> Dict[str, Any]:
    """
    处理单个文件，包括生成超链接信息、检查匹配的TXT文件、读取TXT内容、清洗数据和检测类型。
    返回一个包含所有处理后数据的字典。
    """
    file_stem, file_ext = get_file_details(file_path)
    file_abs_path = file_path.resolve()

    file_link_location = None
    file_link_text = "文件不存在"

    if file_abs_path.exists():
        file_link_location = normalize_drive_letter(str(file_abs_path)).replace("\\", "/")
        if not sys.platform.startswith('win'):
            file_link_location = f'file://{file_link_location}'
        file_link_text = str(file_abs_path.name)
    else:
        logger_obj.info(f"文件不存在，无法生成有效超链接: {normalize_drive_letter(str(file_abs_path))}")
        file_link_text = f"文件不存在: {file_abs_path.name}"

    txt_content = ""
    cleaned_data = ""
    cleaned_data_length = 0
    prompt_type = "N/A"
    found_txt_flag = '否'
    txt_absolute_path_str = "N/A"

    if file_stem.lower() in current_txt_files:
        txt_file_path = current_txt_files[file_stem.lower()]
        txt_absolute_path_str = str(txt_file_path.resolve())
        found_txt_flag = '是'
        try:
            with open(txt_file_path, 'r', encoding='utf-8') as f:
                for line in f:
                    txt_content = line.strip()
                    cleaned_data, _ = clean_tags(txt_content)
                    cleaned_data_length = len(cleaned_data)
                    prompt_type = detect_types(txt_content, cleaned_data)

                    for tag in cleaned_data.split(', '):
                        if tag:
                            tag_counts[tag.strip().lower()] += 1
                    break # 只读取TXT文件的第一行，如果需要读取所有行请移除此行
        except Exception as e:
            logger_obj.error(f"错误: 读取或处理TXT文件 {normalize_drive_letter(str(txt_file_path))} 失败: {e}")
            txt_content = f"Error reading TXT: {e}"
            found_txt_flag = '否 (读取错误)'
    else:
        logger_obj.info(f"未找到匹配的TXT文件: {normalize_drive_letter(str(file_path))}")

    return {
        "root_resolved_path": str(file_path.parent.resolve()),
        "file_absolute_path": str(file_abs_path),
        "file_link_text": file_link_text,
        "file_link_location": file_link_location,
        "file_extension": file_ext,
        "txt_absolute_path": txt_absolute_path_str,
        "txt_content": txt_content,
        "cleaned_data": cleaned_data,
        "cleaned_data_length": cleaned_data_length,
        "prompt_type": prompt_type,
        "found_txt_flag": found_txt_flag
    }

# --- Scanner ---
def scan_files_and_extract_data(
    base_folder_path: Path,
    ws_matched: Worksheet,
    ws_no_txt: Worksheet,
    logger_obj # 移除类型提示
) -> Tuple[int, int, int, Dict[str, int]]:
    """
    扫描指定文件夹下的文件，查找匹配的TXT文件，提取数据并写入Excel。
    """
    total_files_scanned = 0
    found_txt_count = 0
    not_found_txt_count = 0
    tag_counts = defaultdict(int)

    all_extensions: Set[str] = set()
    skipped_extensions: Set[str] = set()

    logger_obj.info(f"开始扫描文件夹: {normalize_drive_letter(str(base_folder_path))}")

    try:
        for root_str, dirs, files in os.walk(base_folder_path):
            root = Path(root_str)

            if _should_skip_path(root, SKIP_SCAN_FOLDERS):
                logger_obj.info(f"跳过扫描文件夹及其子文件夹: {normalize_drive_letter(str(root))}")
                dirs[:] = []
                continue

            current_txt_files = {os.path.splitext(f)[0].lower(): root / f for f in files if f.lower().endswith('.txt')}

            for file_name in files:
                file_path = root / file_name
                _file_stem, file_ext = get_file_details(file_path)
                file_ext_lower = file_ext.lower()

                all_extensions.add(file_ext_lower)

                if file_ext_lower in SKIP_SCAN_EXTENSIONS:
                    skipped_extensions.add(file_ext_lower)
                    continue

                total_files_scanned += 1

                # 调用辅助函数处理单个文件
                processed_data = _process_single_file(file_path, current_txt_files, logger_obj, tag_counts)

                # 根据处理结果填充 Excel
                if processed_data["found_txt_flag"] == '是':
                    found_txt_count += 1
                    current_row_data = [
                        processed_data["root_resolved_path"],
                        processed_data["file_absolute_path"],
                        processed_data["file_link_text"],
                        processed_data["file_extension"],
                        processed_data["txt_absolute_path"],
                        processed_data["txt_content"],
                        processed_data["cleaned_data"],
                        processed_data["cleaned_data_length"],
                        processed_data["prompt_type"],
                        processed_data["found_txt_flag"]
                    ]
                    ws_matched.append(current_row_data)
                    link_cell = ws_matched.cell(row=ws_matched.max_row, column=3)
                    set_hyperlink_and_style(
                        link_cell,
                        processed_data["file_link_location"],
                        processed_data["file_link_text"],
                        logger_obj,
                        source_description=f"匹配文件 (行: {ws_matched.max_row})"
                    )
                else:
                    not_found_txt_count += 1
                    current_row_data_no_txt = [
                        processed_data["root_resolved_path"],
                        processed_data["file_absolute_path"],
                        processed_data["file_link_text"],
                        processed_data["file_extension"],
                        processed_data["found_txt_flag"]
                    ]
                    ws_no_txt.append(current_row_data_no_txt)
                    link_cell = ws_no_txt.cell(row=ws_no_txt.max_row, column=3)
                    set_hyperlink_and_style(
                        link_cell,
                        processed_data["file_link_location"],
                        processed_data["file_link_text"],
                        logger_obj,
                        source_description=f"未匹配文件 (行: {ws_no_txt.max_row})"
                    )

    except Exception as e:
        logger_obj.critical(f"致命错误: 扫描文件过程中发生意外错误 for folder {normalize_drive_letter(str(base_folder_path))}: {e}")


    logger_obj.info(f"文件夹 {normalize_drive_letter(str(base_folder_path))} 扫描完成. 总文件数: {total_files_scanned}, 找到TXT: {found_txt_count}, 未找到TXT: {not_found_txt_count}")

    logger_obj.info(f"\n--- 扫描文件类型概览 ---")
    if all_extensions:
        for ext in sorted(list(all_extensions)):
            status = "已处理"
            if ext in skipped_extensions:
                status = "已跳过"
            logger_obj.info(f"文件扩展名: '{ext}' - 状态: {status}")
    else:
        logger_obj.info("未扫描到任何文件扩展名。")
    logger_obj.info(f"\n--- 文件类型概览结束 ---")

    return total_files_scanned, found_txt_count, not_found_txt_count, tag_counts