import os
import sys
import datetime
from pathlib import Path
from typing import  Dict,List,Any

# openpyxl 相关的导入
from openpyxl import Workbook, load_workbook

# 新增：从 my_logger 导入 normalize_drive_letter
from file_system_utils import normalize_drive_letter # 新增行 2

from excel_utilities import set_hyperlink_and_style, set_fixed_column_widths,HYPERLINK_FONT # 导入辅助函数
from excel_utilities import FIXED_COLUMN_WIDTH # 导入固定列宽常量


# --- Configuration ---
HISTORY_FOLDER_NAME = "运行历史记录"
HISTORY_EXCEL_NAME = "scan_history.xlsx"

class HistoryManager:
    """
    负责程序扫描历史记录的Excel文件存储。
    """
    def __init__(self, history_file_path: Path, logger_obj):
        self.history_file_path = history_file_path
        self.logger_obj = logger_obj
        self.history_data: List[Dict[str, Any]] = [] # 存储内存中的历史记录
        self._load_history_from_excel()

    def _load_history_from_excel(self):
        """
        从Excel文件加载历史记录到内存。
        """
        self.history_data = []
        if not self.history_file_path.exists():
            self.logger_obj.info(f"历史记录Excel文件不存在: {normalize_drive_letter(str(self.history_file_path))}. 将创建新文件。") # 替换为 Loguru 的 info 方法
            return

        try:
            wb = load_workbook(str(self.history_file_path))
            if "扫描历史" in wb.sheetnames:
                ws = wb["扫描历史"]
                headers = [cell.value for cell in ws[1]] # 获取表头
                if not headers:
                    self.logger_obj.warning(f"历史记录Excel文件 '{normalize_drive_letter(str(self.history_file_path))}' 的 '扫描历史' 工作表为空，无历史记录可加载。")#warning
                    return

                # 确保表头符合预期，避免因旧文件格式导致的问题
                expected_headers = [
                    "扫描时间",
                    "文件夹路径",
                    "总文件数",
                    "找到TXT文件数",
                    "未找到TXT文件数",
                    "Log文件绝对路径",
                    "Log文件超链接", # 此列不再是实际数据，而是超链接显示文本
                    "结果XLSX文件绝对路径",
                    "结果XLSX文件超链接" # 此列不再是实际数据，而是超链接显示文本
                ]
                # 简化检查，只需要检查前5列和两个绝对路径列是否存在，因为超链接列是动态生成的
                if not all(h in headers for h in expected_headers[:6] + [expected_headers[7]]): # 检查前6个和第8个（结果XLSX绝对路径）
                    self.logger_obj.warning(f"历史记录Excel文件 '{normalize_drive_letter(str(self.history_file_path))}' 表头不匹配预期，可能无法完全加载所有历史记录。")#warning
                    # 继续尝试加载，但可能不完整

                for row_idx in range(2, ws.max_row + 1): # 从第二行开始读取数据
                    row_values = [cell.value for cell in ws[row_idx]]
                    # 将行数据映射到字典
                    entry = {}
                    for i, header in enumerate(headers):
                        if i < len(row_values):
                            entry[header] = row_values[i]
                        else:
                            entry[header] = None # 如果某些列没有值，则设置为None

                    # 仅保留我们需要的字段，并确保路径是Path对象
                    self.history_data.append({
                        'scan_time': entry.get("扫描时间"),
                        'folder_path': Path(entry.get("文件夹路径")) if entry.get("文件夹路径") else None,
                        'total_files': entry.get("总文件数"),
                        'found_txt_count': entry.get("找到TXT文件数"),
                        'not_found_txt_count': entry.get("未找到TXT文件数"),
                        'log_file_abs_path': Path(entry.get("Log文件绝对路径")) if entry.get("Log文件绝对路径") else None,
                        'result_xlsx_abs_path': Path(entry.get("结果XLSX文件绝对路径")) if entry.get("结果XLSX文件绝对路径") else None
                    })
            self.logger_obj.info(f"成功从历史记录Excel文件加载 {len(self.history_data)} 条历史记录。")
        except Exception as e:
            self.logger_obj.error(f"错误: 从历史记录Excel文件 {normalize_drive_letter(str(self.history_file_path))} 加载历史记录失败: {e}")#error
            self.history_data = [] # 加载失败则清空内存数据，避免脏数据

    def add_history_entry(self, folder_path: Path, total_scanned: int, found_txt_count: int,
                          not_found_txt_count: int, result_file_path: Path, log_file_path: Path):
        """
        向内存中的历史记录列表添加一条新的扫描历史记录。
        """
        scan_time = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        entry = {
            'scan_time': scan_time,
            'folder_path': folder_path,
            'total_files': total_scanned,
            'found_txt_count': found_txt_count,
            'not_found_txt_count': not_found_txt_count,
            'log_file_abs_path': log_file_path,
            'result_xlsx_abs_path': result_file_path
        }
        self.history_data.append(entry)
        self.logger_obj.info(f"历史记录成功添加至内存: 文件夹'{folder_path.name}'")

    def save_history_to_excel(self) -> bool:
        """
        将内存中的所有历史记录保存到Excel文件。
        Returns:
            bool: 如果保存成功返回True，否则返回False。
        """
        self.logger_obj.info(f"开始将内存中的历史记录保存到Excel: {normalize_drive_letter(str(self.history_file_path))}")

        # 尝试删除旧的历史Excel文件，以便重新写入
        if self.history_file_path.exists():
            try:
                os.remove(str(self.history_file_path))
                self.logger_obj.info(f"已删除旧的历史记录Excel文件: {normalize_drive_letter(str(self.history_file_path))}")
            except PermissionError as e: # 明确捕获权限错误
                self.logger_obj.critical(f"致命错误: 无法删除旧的历史记录Excel文件 '{normalize_drive_letter(str(self.history_file_path))}'，可能文件被占用。请关闭Excel中打开的历史文件。错误详情: {e}")
                return False # 删除失败，返回False
            except Exception as e:
                self.logger_obj.critical(f"致命错误: 删除旧的历史记录Excel文件 '{normalize_drive_letter(str(self.history_file_path))}' 时发生未知错误。历史记录将无法保存。错误详情: {e}")
                return False # 删除失败，返回False

        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "扫描历史"

            # 定义Excel表头
            excel_headers = [
                "扫描时间",
                "文件夹路径",
                "总文件数",
                "找到TXT文件数",
                "未找到TXT文件数",
                "Log文件绝对路径",
                "Log文件超链接",
                "结果XLSX文件绝对路径",
                "结果XLSX文件超链接"
            ]
            ws.append(excel_headers)

            for entry in self.history_data:
                log_file_abs_path = entry['log_file_abs_path']
                result_xlsx_abs_path = entry['result_xlsx_abs_path']

                # Log文件超链接的显示文本和location
                log_link_display_text = "打开Log" 
                log_link_location = None
                if log_file_abs_path and log_file_abs_path.exists():
                    log_link_location = normalize_drive_letter(str(log_file_abs_path)).replace("\\", "/")
                    if not sys.platform.startswith('win'): 
                        log_link_location = f'file://{log_link_location}'
                else:
                    log_link_display_text = "Log文件不存在"


                # 结果XLSX文件超链接的显示文本和location
                result_link_display_text = "打开结果XLSX"
                result_link_location = None
                if result_xlsx_abs_path and result_xlsx_abs_path.exists():
                    result_link_location = normalize_drive_letter(str(result_xlsx_abs_path)).replace("\\", "/")
                    if not sys.platform.startswith('win'): 
                        result_link_location = f'file://{result_link_location}'
                else:
                    result_link_display_text = "结果XLSX文件不存在"
                
                row_data = [
                    entry['scan_time'],
                    str(entry['folder_path']),
                    entry['total_files'],
                    entry['found_txt_count'],
                    entry['not_found_txt_count'],
                    normalize_drive_letter(str(log_file_abs_path)) if log_file_abs_path else "N/A",      # Log文件绝对路径
                    log_link_display_text,       # Log文件超链接显示文本
                    normalize_drive_letter(str(result_xlsx_abs_path)) if result_xlsx_abs_path else "N/A",   # 结果XLSX文件绝对路径
                    result_link_display_text    # 结果XLSX文件超链接显示文本
                ]
                ws.append(row_data)

                # 获取新添加的行的单元格，并设置超链接
                new_row_idx = ws.max_row
                
                # Log文件超链接单元格 (第7列)
                log_link_cell = ws.cell(row=new_row_idx, column=7)
                set_hyperlink_and_style(
                    log_link_cell, 
                    log_link_location, # 传入可能为None的location
                    log_link_display_text, 
                    self.logger_obj, 
                    source_description=f"历史记录Log文件 (行: {new_row_idx})"
                )

                # 结果XLSX文件超链接单元格 (第9列)
                result_link_cell = ws.cell(row=new_row_idx, column=9)
                set_hyperlink_and_style(
                    result_link_cell, 
                    result_link_location, # 传入可能为None的location
                    result_link_display_text, 
                    self.logger_obj, 
                    source_description=f"历史记录结果XLSX文件 (行: {new_row_idx})"
                )
            
            # 设置所有列宽
            set_fixed_column_widths(ws, FIXED_COLUMN_WIDTH, self.logger_obj)
            
            wb.save(str(self.history_file_path))
            self.logger_obj.info(f"成功将历史记录保存到Excel: {normalize_drive_letter(str(self.history_file_path))}")
            return True

        except Exception as e:
            self.logger_obj.error(f"错误: 将历史记录保存到Excel失败 {normalize_drive_letter(str(self.history_file_path))}: {e}")#error
            return False


