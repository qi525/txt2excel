# excel_utilities.py
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from typing import Tuple, Optional, List, Dict # 导入 List 和 Dict
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

HYPERLINK_FONT = Font(color="0000FF", underline="single")

# 定义固定列宽（以字符为单位）
FIXED_COLUMN_WIDTH = 20

# --- Excel Utilities (Modified for more generality) ---

def create_empty_workbook() -> Workbook: # 重命名，更清晰
    """
    创建一个空的Excel工作簿，并移除默认创建的Sheet。
    通用性：高
    """
    wb = Workbook()
    # 移除默认创建的Sheet
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    return wb

def create_sheet_with_headers(
    workbook: Workbook,
    sheet_name: str,
    headers: List[str],
    index: Optional[int] = None
) -> Worksheet:
    """
    在给定的工作簿中创建一个新工作表，并设置其标题行。
    Args:
        workbook (Workbook): openpyxl 工作簿对象。
        sheet_name (str): 新工作表的名称。
        headers (List[str]): 包含所有列标题的列表。
        index (Optional[int]): 工作表插入的位置索引。如果为 None，则添加到最后。
    Returns:
        Worksheet: 新创建的工作表对象。
    通用性：高
    """
    ws = workbook.create_sheet(sheet_name, index)
    ws.append(headers)
    return ws


# --- 辅助函数 ---
# 将 set_hyperlink_and_style 函数粘贴到这里
# --- MODIFIED FUNCTION: 设置单元格超链接和样式 ---
def set_hyperlink_and_style(
    cell, 
    location: Optional[str], 
    display_text: str, 
    logger_obj, 
    source_description: str = "未知来源"
) -> bool:
    """
    为openpyxl单元格设置超链接和样式。
    原理：
        此函数现在专注于设置超链接和样式。它不再在内部记录所有非异常情况下的日志。
        如果由于location无效而未能设置超链接，它会静默处理，但会在发生实际错误时记录ERROR级别的日志。
        同时，函数返回一个布尔值来告知调用者操作是否成功，让调用者可以根据返回值进行后续处理或日志记录。
    Args:
        cell: openpyxl单元格对象。
        location (Optional[str]): 超链接的目标路径，如果为None或空字符串则不设置超链接。
        display_text (str): 单元格显示的文本。
        logger_obj: 日志管理器实例。
        source_description (str): 用于日志记录的来源描述，方便定位问题。
    Returns:
        bool: 如果成功设置超链接则返回True，否则返回False。
    """
    cell.value = display_text # 总是设置单元格的显示文本
    try:
        if location: # 检查 location 是否存在且非空
            cell.hyperlink = location
            cell.font = HYPERLINK_FONT # 设置超链接字体样式
            return True
        else:
            # 如果没有 location，确保不设置超链接，并移除可能的超链接样式
            cell.hyperlink = None 
            cell.font = Font(color="000000") # 恢复默认字体颜色，去除下划线
            return False # 未设置超链接
    except Exception as e:
        logger_obj.error(
            f"错误: 无法为单元格设置超链接或样式 for '{display_text}' (Location: '{location}', Source: {source_description}). 错误: {e}"
        )
        # 即使出错，也要确保单元格值被设置，即使没有超链接样式
        cell.value = display_text
        return False # 发生错误，未能设置超链接


# excel_utilities.py
# ... (前面的导入和常量保持不变) ...

def set_column_widths(
    worksheet: Worksheet,
    column_widths: Optional[Dict[str, int]] = None,
    default_width: Optional[int] = None,
    logger_obj=None
):
    """
    为给定工作表的列设置宽度。
    可以为所有列设置一个默认宽度，或者为特定列设置指定宽度。
    Args:
        worksheet (Worksheet): openpyxl 工作表对象。
        column_widths (Optional[Dict[str, int]]): 字典，键为列字母（如'A'）或列索引（从1开始），值为宽度。
        default_width (Optional[int]): 如果指定，所有未在 column_widths 中明确设置的列将使用此默认宽度。
        logger_obj (logger): 日志管理器实例。
    通用性：高
    """
    if logger_obj is None:
        # 避免未传入logger_obj时报错，可以使用一个哑日志器或直接pass
        class DummyLogger:
            def debug(self, msg): pass
            def info(self, msg): pass
            def warning(self, msg): pass
            def error(self, msg): pass
            def critical(self, msg): pass
        logger_obj = DummyLogger()

    try:
        if default_width is not None:
            for col_idx in range(1, worksheet.max_column + 1):
                column_letter = get_column_letter(col_idx)
                if column_widths is None or column_letter not in column_widths and col_idx not in column_widths:
                    worksheet.column_dimensions[column_letter].width = default_width

        if column_widths:
            for key, width in column_widths.items():
                if isinstance(key, str) and len(key) == 1 and key.isalpha():
                    col_letter = key.upper()
                elif isinstance(key, int):
                    col_letter = get_column_letter(key)
                else:
                    logger_obj.warning(f"警告: 无效的列宽度键 '{key}'，请使用列字母（如'A'）或列索引（如1）。")
                    continue
                worksheet.column_dimensions[col_letter].width = width
        logger_obj.debug(f"已为工作表 '{worksheet.title}' 设置列宽。")
    except Exception as e:
        logger_obj.error(f"错误: 无法为工作表 '{worksheet.title}' 设置列宽。错误: {e}")

# 将原有的 set_fixed_column_widths 修改为调用新的 set_column_widths
def set_fixed_column_widths(worksheet: Worksheet, width: int, logger_obj):
    """
    为给定工作表的所有列设置固定宽度。
    这是一个兼容旧接口的函数，内部调用 set_column_widths。
    Args:
        worksheet (Worksheet): openpyxl 工作表对象。
        width (int): 要设置的列宽。
        logger_obj (logger): 日志管理器实例。
    通用性：高 (作为旧接口的兼容层)
    """
    set_column_widths(worksheet, default_width=width, logger_obj=logger_obj)