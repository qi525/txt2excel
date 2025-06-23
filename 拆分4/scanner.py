import os
import sys
import datetime
from pathlib import Path
from openpyxl.worksheet.worksheet import Worksheet
from typing import Tuple, Dict, Optional, Set, List, Any, Generator, Protocol, runtime_checkable
import logging
from dataclasses import dataclass, field
from enum import Enum # 新增导入 Enum

from collections import defaultdict

from file_system_utils import normalize_drive_letter, get_file_details
from tag_processing import clean_tags, detect_types
from excel_utilities import set_hyperlink_and_style

# --- 模块级别常量 ---
class ScannerConstants:
    # 直接作为类属性，而不是嵌套类
    SKIP_SCAN_FOLDERS: Set[str] = {'.bf'}
    SKIP_SCAN_EXTENSIONS: Set[str] = {'.txt', '.xlsx', '.json', '.ini', '.db'}

    # 使用 Enum 类定义文件状态
    class FileStatus(Enum):
        FOUND_TXT_FLAG_YES = '是'
        FOUND_TXT_FLAG_NO = '否'
        FOUND_TXT_FLAG_ERROR = '否 (读取错误)'
        PROMPT_TYPE_NA = "N/A"
        FILE_NOT_EXISTS_TEXT = "文件不存在"

    # 使用 Enum 类定义错误类型
    class ErrorTypes(Enum):
        FILE_NOT_FOUND = "文件不存在"
        READ_TXT_FAILED = "读取TXT文件失败"
        TAG_PROCESSING_FAILED = "标签处理失败"
        INVALID_RETURN_TYPE = "返回非预期类型"
        DIRECTORY_ACCESS_FAILED = "目录访问失败"
        UNEXPECTED_SCAN_ERROR = "意外扫描错误"
        UNKNOWN_ERROR = "未知错误"

    class ExcelConfig:
        EXCEL_FILE_LINK_COLUMN = 3

# Scanner 配置类保持不变
@dataclass
class ScannerConfig:
    """
    封装 Scanner 的配置参数。
    """
    skip_folders: Set[str] = field(default_factory=lambda: ScannerConstants.SKIP_SCAN_FOLDERS.copy())
    skip_extensions: Set[str] = field(default_factory=lambda: ScannerConstants.SKIP_SCAN_EXTENSIONS.copy())

# 结构化错误记录保持不变
@dataclass
class ErrorRecord:
    """
    定义一个结构化的错误记录。
    """
    error_type: str
    message: str
    file_path: Optional[str] = None
    timestamp: datetime.datetime = field(default_factory=datetime.datetime.now)
    details: Optional[str] = None

@dataclass
class ProcessedFileData:
    """
    定义处理后文件数据的结构，提高代码可读性和类型安全性，并提供更像对象的访问方式。
    新增了结构化的错误记录列表。
    """
    root_resolved_path: str
    file_absolute_path: str
    file_link_text: str
    file_link_location: Optional[str]
    file_extension: str
    txt_absolute_path: str
    txt_content: str
    cleaned_data: str
    cleaned_data_length: int
    prompt_type: str
    found_txt_flag: str
    processing_errors: List[ErrorRecord] = field(default_factory=list)

    # _is_matched_flag 被替换为一个属性
    @property
    def is_matched_flag(self) -> bool:
        """
        根据 found_txt_flag 动态判断是否匹配成功。
        """
        return self.found_txt_flag == ScannerConstants.FileStatus.FOUND_TXT_FLAG_YES.value


# 元数据处理器接口和实现保持不变
@runtime_checkable
class MetadataProcessor(Protocol):
    """
    元数据处理器的抽象接口。
    定义了如何从文件中提取和处理元数据。
    """
    def process(self, file_path: Path, logger_obj: logging.Logger) -> Tuple[str, str, int, str, str, List[ErrorRecord]]:
        """
        处理指定文件，提取其元数据。
        返回 (txt_absolute_path, txt_content, cleaned_data, cleaned_data_length, prompt_type, errors)
        """
        ...

class TxtMetadataProcessor:
    """
    TXT文件元数据处理器的具体实现。
    """
    def process(self, txt_file_path: Path, logger_obj: logging.Logger) -> Tuple[str, str, int, str, str, List[ErrorRecord]]:
        txt_absolute_path = normalize_drive_letter(str(txt_file_path.resolve()))
        txt_content = ""
        cleaned_data = ""
        cleaned_data_length = 0
        prompt_type = ScannerConstants.FileStatus.PROMPT_TYPE_NA.value # 使用 .value
        errors: List[ErrorRecord] = []

        try_encodings = ['utf-8', 'gbk', 'latin-1']
        txt_read_success = False

        for encoding in try_encodings:
            try:
                with open(txt_file_path, 'r', encoding=encoding) as f:
                    txt_content = f.readline().strip()
                    txt_read_success = True
                    break
            except UnicodeDecodeError as e:
                msg = f"TXT文件 {normalize_drive_letter(str(txt_file_path))} 无法使用 {encoding} 解码，尝试其他编码。"
                logger_obj.warning(f"警告: {msg}")
                errors.append(ErrorRecord(ScannerConstants.ErrorTypes.READ_TXT_FAILED.value, msg, file_path=normalize_drive_letter(str(txt_file_path)), details=str(e))) # 使用 .value
                continue
            except Exception as e:
                msg = f"读取TXT文件 {normalize_drive_letter(str(txt_file_path))} 失败 (编码尝试 {encoding}): {e}"
                logger_obj.error(f"错误: {msg}")
                errors.append(ErrorRecord(ScannerConstants.ErrorTypes.READ_TXT_FAILED.value, msg, file_path=normalize_drive_letter(str(txt_file_path)), details=str(e))) # 使用 .value
                txt_read_success = False
                break

        if not txt_read_success:
            if not any(err.error_type == ScannerConstants.ErrorTypes.READ_TXT_FAILED.value for err in errors): # 使用 .value
                 errors.append(ErrorRecord(ScannerConstants.ErrorTypes.READ_TXT_FAILED.value, # 使用 .value
                                            "无法通过任何尝试的编码解码或发生其他读取错误",
                                            file_path=normalize_drive_letter(str(txt_file_path))))
            return txt_absolute_path, txt_content, cleaned_data, cleaned_data_length, prompt_type, errors

        if txt_read_success:
            try:
                temp_cleaned_data, _ = clean_tags(txt_content)

                if not isinstance(temp_cleaned_data, str):
                    msg = f"cleaned_data {ScannerConstants.ErrorTypes.INVALID_RETURN_TYPE.value}. 实际类型: {type(temp_cleaned_data).__name__}" # 使用 .value
                    logger_obj.warning(f"警告: {ScannerConstants.ErrorTypes.TAG_PROCESSING_FAILED.value}: {msg} for {normalize_drive_letter(str(txt_file_path))}") # 使用 .value
                    errors.append(ErrorRecord(ScannerConstants.ErrorTypes.TAG_PROCESSING_FAILED.value, msg, file_path=normalize_drive_letter(str(txt_file_path)))) # 使用 .value
                    temp_cleaned_data = ""
                cleaned_data = temp_cleaned_data
                cleaned_data_length = len(cleaned_data) if isinstance(cleaned_data, str) else 0

                temp_prompt_type = detect_types(txt_content, cleaned_data)

                if not isinstance(temp_prompt_type, str):
                    msg = f"prompt_type {ScannerConstants.ErrorTypes.INVALID_RETURN_TYPE.value}. 实际类型: {type(temp_prompt_type).__name__}" # 使用 .value
                    logger_obj.warning(f"警告: {ScannerConstants.ErrorTypes.TAG_PROCESSING_FAILED.value}: {msg} for {normalize_drive_letter(str(txt_file_path))}") # 使用 .value
                    errors.append(ErrorRecord(ScannerConstants.ErrorTypes.TAG_PROCESSING_FAILED.value, msg, file_path=normalize_drive_letter(str(txt_file_path)))) # 使用 .value
                    temp_prompt_type = ScannerConstants.FileStatus.PROMPT_TYPE_NA.value # 使用 .value
                prompt_type = temp_prompt_type

            except Exception as e:
                msg = f"标签处理失败: {e}"
                logger_obj.error(f"错误: {msg} for TXT文件 {normalize_drive_letter(str(txt_file_path))}")
                errors.append(ErrorRecord(ScannerConstants.ErrorTypes.TAG_PROCESSING_FAILED.value, msg, file_path=normalize_drive_letter(str(txt_file_path)), details=str(e))) # 使用 .value
                cleaned_data = ""
                cleaned_data_length = 0
                prompt_type = ScannerConstants.FileStatus.PROMPT_TYPE_NA.value # 使用 .value

        return txt_absolute_path, txt_content, cleaned_data, cleaned_data_length, prompt_type, errors

# 数据写入器接口和实现保持不变
@runtime_checkable
class DataWriter(Protocol):
    """
    数据写入器的抽象接口。
    定义了如何将处理后的文件数据写入到目标。
    """
    def write_matched_data(self, data: ProcessedFileData):
        """写入匹配文件的数据。"""
        ...

    def write_no_txt_data(self, data: ProcessedFileData):
        """写入未找到TXT文件的数据。"""
        ...

class ExcelDataWriter:
    """
    Excel 数据写入器的具体实现。
    将数据写入到两个 openpyxl 工作表。
    """
    def __init__(self, ws_matched: Worksheet, ws_no_txt: Worksheet, logger_obj: logging.Logger):
        self.ws_matched = ws_matched
        self.ws_no_txt = ws_no_txt
        self.logger_obj = logger_obj

    def write_matched_data(self, processed_data: ProcessedFileData):
        if processed_data.processing_errors:
            full_error_message = "; ".join([f"{err.error_type}: {err.message}" for err in processed_data.processing_errors])
            self.logger_obj.warning(f"文件 '{normalize_drive_letter(processed_data.file_absolute_path)}' 处理中遇到错误：{full_error_message}")

        current_row_data = [
            processed_data.root_resolved_path,
            processed_data.file_absolute_path,
            processed_data.file_link_text,
            processed_data.file_extension,
            processed_data.txt_absolute_path,
            processed_data.txt_content,
            processed_data.cleaned_data,
            processed_data.cleaned_data_length,
            processed_data.prompt_type,
            processed_data.found_txt_flag
        ]
        self.ws_matched.append(current_row_data)
        link_cell = self.ws_matched.cell(row=self.ws_matched.max_row, column=ScannerConstants.ExcelConfig.EXCEL_FILE_LINK_COLUMN)
        set_hyperlink_and_style(
            link_cell,
            processed_data.file_link_location,
            processed_data.file_link_text,
            self.logger_obj,
            source_description=f"匹配文件 (行: {self.ws_matched.max_row})"
        )

    def write_no_txt_data(self, processed_data: ProcessedFileData):
        if processed_data.processing_errors:
            full_error_message = "; ".join([f"{err.error_type}: {err.message}" for err in processed_data.processing_errors])
            self.logger_obj.warning(f"文件 '{normalize_drive_letter(processed_data.file_absolute_path)}' 处理中遇到错误：{full_error_message}")

        current_row_data_no_txt = [
            processed_data.root_resolved_path,
            processed_data.file_absolute_path,
            processed_data.file_link_text,
            processed_data.file_extension,
            processed_data.found_txt_flag
        ]
        self.ws_no_txt.append(current_row_data_no_txt)
        link_cell = self.ws_no_txt.cell(row=self.ws_no_txt.max_row, column=ScannerConstants.ExcelConfig.EXCEL_FILE_LINK_COLUMN)
        set_hyperlink_and_style(
            link_cell,
            processed_data.file_link_location,
            processed_data.file_link_text,
            self.logger_obj,
            source_description=f"未匹配文件 (行: {self.ws_no_txt.max_row})"
        )

# 标签聚合器协议和实现保持不变
@runtime_checkable
class TagAggregator(Protocol):
    """
    标签聚合器的抽象接口。
    定义了如何收集和获取标签统计结果。
    """
    def add_tags(self, tags: List[str]):
        """添加一批标签进行聚合。"""
        ...

    def get_counts(self) -> Dict[str, int]:
        """获取聚合后的标签计数。"""
        ...

class DefaultTagAggregator:
    """
    默认的标签聚合器实现，使用 defaultdict 进行计数。
    """
    def __init__(self):
        self._tag_counts = defaultdict(int)

    def add_tags(self, tags: List[str]):
        for tag in tags:
            if tag: # 确保标签不为空
                self._tag_counts[tag] += 1

    def get_counts(self) -> Dict[str, int]:
        return dict(self._tag_counts) # 返回字典的副本

class Scanner:
    def __init__(self, logger_obj: logging.Logger,
                 data_writer: DataWriter,
                 config: ScannerConfig = ScannerConfig(),
                 tag_aggregator: TagAggregator = DefaultTagAggregator()):
        self.logger_obj = logger_obj
        self.data_writer = data_writer
        self.config = config
        self.tag_aggregator = tag_aggregator
        self.all_extensions: Set[str] = set()
        self.skipped_extensions: Set[str] = set()
        self.all_scan_errors: List[ErrorRecord] = []
        self.metadata_processors: Dict[str, MetadataProcessor] = {
            '.txt': TxtMetadataProcessor()
        }

    def _generate_file_link_info(self, file_path: Path) -> Tuple[Optional[str], str, Optional[ErrorRecord]]:
        file_abs_path = file_path.resolve()
        file_link_location: Optional[str] = None
        file_link_text: str = ScannerConstants.FileStatus.FILE_NOT_EXISTS_TEXT.value # 使用 .value
        error_record: Optional[ErrorRecord] = None

        if file_abs_path.exists():
            file_link_location = normalize_drive_letter(str(file_abs_path)).replace("\\", "/")
            if not sys.platform.startswith('win'):
                file_link_location = f'file://{file_link_location}'
            file_link_text = file_abs_path.name
        else:
            msg = f"文件不存在: {normalize_drive_letter(str(file_abs_path))}"
            self.logger_obj.info(f"文件不存在，无法生成有效超链接: {msg}")
            error_record = ErrorRecord(ScannerConstants.ErrorTypes.FILE_NOT_FOUND.value, msg, file_path=normalize_drive_letter(str(file_abs_path))) # 使用 .value

        return file_link_location, file_link_text, error_record

    def _process_file_metadata(
        self,
        file_path: Path,
        matched_txt_path: Optional[Path]
    ) -> ProcessedFileData:
        file_stem, file_ext = get_file_details(file_path)

        file_link_location, file_link_text, file_exist_error = self._generate_file_link_info(file_path)

        result_data = ProcessedFileData(
            root_resolved_path=normalize_drive_letter(str(file_path.parent.resolve())),
            file_absolute_path=normalize_drive_letter(str(file_path.resolve())),
            file_link_text=file_link_text,
            file_link_location=file_link_location,
            file_extension=file_ext,
            txt_absolute_path=ScannerConstants.FileStatus.PROMPT_TYPE_NA.value, # 使用 .value
            txt_content="",
            cleaned_data="",
            cleaned_data_length=0,
            prompt_type=ScannerConstants.FileStatus.PROMPT_TYPE_NA.value, # 使用 .value
            found_txt_flag=ScannerConstants.FileStatus.FOUND_TXT_FLAG_NO.value # 使用 .value
        )
        # _is_matched_flag 现在通过属性自动计算，无需在这里设置

        if file_exist_error:
            result_data.processing_errors.append(file_exist_error)
            self.all_scan_errors.append(file_exist_error)

        if matched_txt_path:
            processor = self.metadata_processors.get('.txt')
            if processor:
                txt_absolute_path, txt_content, cleaned_data, cleaned_data_length, prompt_type, errors = \
                    processor.process(matched_txt_path, self.logger_obj)

                result_data.txt_absolute_path = txt_absolute_path
                result_data.txt_content = txt_content
                result_data.cleaned_data = cleaned_data
                result_data.cleaned_data_length = cleaned_data_length
                result_data.prompt_type = prompt_type
                result_data.processing_errors.extend(errors)
                self.all_scan_errors.extend(errors)

                if not errors and txt_content:
                     result_data.found_txt_flag = ScannerConstants.FileStatus.FOUND_TXT_FLAG_YES.value # 使用 .value
                else:
                    result_data.found_txt_flag = ScannerConstants.FileStatus.FOUND_TXT_FLAG_ERROR.value # 使用 .value

                if result_data.cleaned_data and isinstance(result_data.cleaned_data, str):
                    tags = [t.strip().lower() for t in result_data.cleaned_data.split(',') if t.strip()]
                    self.tag_aggregator.add_tags(tags)
            else:
                msg = f"未找到处理 {matched_txt_path.suffix} 文件的元数据处理器。"
                self.logger_obj.error(msg)
                err_record = ErrorRecord(ScannerConstants.ErrorTypes.UNKNOWN_ERROR.value, msg, file_path=normalize_drive_letter(str(matched_txt_path))) # 使用 .value
                result_data.processing_errors.append(err_record)
                self.all_scan_errors.append(err_record)
                result_data.found_txt_flag = ScannerConstants.FileStatus.FOUND_TXT_FLAG_ERROR.value # 使用 .value
        else:
            self.logger_obj.info(f"未找到匹配的TXT文件: {normalize_drive_letter(str(file_path))}")

        # result_data._is_matched_flag = (result_data.found_txt_flag == ScannerConstants.FileStatus.FOUND_TXT_FLAG_YES) # 移除此行
        return result_data

    def _scan_directory_recursive(self, current_dir: Path, all_files_to_scan: List[Path], all_txt_files_map: Dict[str, Path]):
        try:
            with os.scandir(current_dir) as entries:
                for entry in entries:
                    entry_path = Path(entry.path)

                    if entry.is_dir():
                        if any(sf in entry_path.parts or entry_path.name == sf for sf in self.config.skip_folders):
                            self.logger_obj.info(f"跳过扫描文件夹及其子文件夹: {normalize_drive_letter(str(entry_path))}")
                            continue
                        self._scan_directory_recursive(entry_path, all_files_to_scan, all_txt_files_map)
                    elif entry.is_file():
                        _file_stem, file_ext = get_file_details(entry_path)
                        file_ext_lower = file_ext.lower()

                        self.all_extensions.add(file_ext_lower)

                        if file_ext_lower == '.txt':
                            all_txt_files_map[_file_stem.lower()] = entry_path
                            continue

                        if file_ext_lower in self.config.skip_extensions:
                            self.skipped_extensions.add(file_ext_lower)
                            continue

                        all_files_to_scan.append(entry_path)
        except PermissionError as e:
            msg = f"权限不足，无法访问目录 '{normalize_drive_letter(str(current_dir))}': {e}"
            self.logger_obj.warning(f"警告: {msg}")
            self.all_scan_errors.append(ErrorRecord(ScannerConstants.ErrorTypes.DIRECTORY_ACCESS_FAILED.value, msg, file_path=normalize_drive_letter(str(current_dir)), details=str(e))) # 使用 .value
        except FileNotFoundError as e:
            msg = f"目录不存在或已被删除 '{normalize_drive_letter(str(current_dir))}': {e}"
            self.logger_obj.warning(f"警告: {msg}")
            self.all_scan_errors.append(ErrorRecord(ScannerConstants.ErrorTypes.DIRECTORY_ACCESS_FAILED.value, msg, file_path=normalize_drive_letter(str(current_dir)), details=str(e))) # 使用 .value
        except OSError as e:
            msg = f"遍历目录 '{normalize_drive_letter(str(current_dir))}' 时发生操作系统错误: {e}"
            self.logger_obj.error(f"错误: {msg}")
            self.all_scan_errors.append(ErrorRecord(ScannerConstants.ErrorTypes.DIRECTORY_ACCESS_FAILED.value, msg, file_path=normalize_drive_letter(str(current_dir)), details=str(e))) # 使用 .value
        except Exception as e:
            msg = f"遍历目录 '{normalize_drive_letter(str(current_dir))}' 时发生意外错误: {e}"
            self.logger_obj.error(f"错误: {msg}")
            self.all_scan_errors.append(ErrorRecord(ScannerConstants.ErrorTypes.UNEXPECTED_SCAN_ERROR.value, msg, file_path=normalize_drive_letter(str(current_dir)), details=str(e))) # 使用 .value


    def _collect_files(self, base_folder_path: Path) -> Tuple[List[Path], Dict[str, Path]]:
        all_files_to_scan: List[Path] = []
        all_txt_files_map: Dict[str, Path] = {}

        self._scan_directory_recursive(base_folder_path, all_files_to_scan, all_txt_files_map)

        return all_files_to_scan, all_txt_files_map

    def scan_files_and_extract_data(
        self,
        base_folder_path: Path,
    ) -> Tuple[int, int, int, Dict[str, int]]:
        total_files_scanned = 0
        found_txt_count = 0
        not_found_txt_count = 0

        self.all_scan_errors.clear()

        self.logger_obj.info(f"开始扫描文件夹: {normalize_drive_letter(str(base_folder_path))}")

        try:
            all_files_to_scan, all_txt_files_map = self._collect_files(base_folder_path)

            total_files_scanned = len(all_files_to_scan)

            for file_path in all_files_to_scan:
                processed_data = self._process_file_metadata(file_path, all_txt_files_map.get(get_file_details(file_path)[0].lower()))

                # 使用 processed_data.is_matched_flag 属性
                if processed_data.is_matched_flag:
                    self.data_writer.write_matched_data(processed_data)
                    found_txt_count += 1
                else:
                    self.data_writer.write_no_txt_data(processed_data)
                    not_found_txt_count += 1

        except Exception as e:
            msg = f"致命错误: {ScannerConstants.ErrorTypes.UNEXPECTED_SCAN_ERROR.value} for folder {normalize_drive_letter(str(base_folder_path))}: {e}" # 使用 .value
            self.logger_obj.critical(msg)
            self.all_scan_errors.append(ErrorRecord(ScannerConstants.ErrorTypes.UNEXPECTED_SCAN_ERROR.value, msg, file_path=normalize_drive_letter(str(base_folder_path)), details=str(e))) # 使用 .value

        self.logger_obj.info(
            f"文件夹 {normalize_drive_letter(str(base_folder_path))} 扫描完成. "
            f"总文件数: {total_files_scanned}, 找到TXT: {found_txt_count}, 未找到TXT: {not_found_txt_count}"
        )

        self.logger_obj.info(f"\n--- 扫描文件类型概览 ---")
        if self.all_extensions:
            for ext in sorted(list(self.all_extensions)):
                status = "已处理"
                if ext in self.skipped_extensions:
                    status = "已跳过"
                self.logger_obj.info(f"文件扩展名: '{ext}' - 状态: {status}")
        else:
            self.logger_obj.info("未扫描到任何文件扩展名。")
        self.logger_obj.info(f"\n--- 文件类型概览结束 ---")

        if self.all_scan_errors:
            self.logger_obj.warning("\n--- 扫描过程错误汇总 ---")
            reported_errors = set()
            for error in self.all_scan_errors:
                error_key = (error.error_type, error.message, error.file_path)
                if error_key not in reported_errors:
                    # 优化日志输出，更清晰地展示错误详情
                    details_str = f", 详情: {error.details}" if error.details else ""
                    self.logger_obj.warning(f"- 类型: {error.error_type}, 消息: {error.message}, 文件: {error.file_path or 'N/A'}{details_str}")
                    reported_errors.add(error_key)
            self.logger_obj.warning("--- 扫描过程错误汇总结束 ---\n")
        else:
            self.logger_obj.info("\n扫描过程中未发现明显错误。")

        return total_files_scanned, found_txt_count, not_found_txt_count, self.tag_aggregator.get_counts()


def scan_files_and_extract_data(
    base_folder_path: Path,
    data_writer: DataWriter,
    logger_obj: logging.Logger
) -> Tuple[int, int, int, Dict[str, int]]:
    """
    扫描指定文件夹下的文件，查找匹配的TXT文件，提取数据并写入。
    此函数现在是 main.py 的适配层，它实例化 Scanner 类并调用其方法。
    """
    scanner_config = ScannerConfig()
    tag_aggregator_instance = DefaultTagAggregator()
    scanner = Scanner(logger_obj=logger_obj, data_writer=data_writer,
                      config=scanner_config, tag_aggregator=tag_aggregator_instance)
    return scanner.scan_files_and_extract_data(base_folder_path)