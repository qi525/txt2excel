# history_execution.py

import os
import sys
import datetime
from pathlib import Path
from typing import Dict, List, Any, Optional, Tuple

# openpyxl 相关的导入
from openpyxl import Workbook, load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from file_system_utils import normalize_drive_letter, create_directory_if_not_exists, copy_file

# 导入辅助函数和常量
from excel_utilities import set_hyperlink_and_style, set_fixed_column_widths
from excel_utilities import FIXED_COLUMN_WIDTH

# --- Configuration ---
# 这些常量可以移到主配置文件中，这里保留是为了模块内部可见性
HISTORY_FOLDER_NAME = "操作记录" # 可以考虑移除，因为路径是传入的
HISTORY_EXCEL_NAME = "operation_records.xlsx" # 可以考虑移除，因为路径是传入的

# 移除 _handle_history_caching 函数，其逻辑将移入 HistoryManager 类中
# 从这里删除了原 _handle_history_caching 函数


class HistoryManager:
    """
    负责存储应用程序操作记录或活动日志的Excel文件。
    现在完全通用化，通过传入 field_definitions 配置数据结构。
    """
    def __init__(self,
                 history_file_path: Path,
                 logger_obj,
                 field_definitions: List[Dict[str, Any]],
                 sheet_name: str = "操作记录",
                 cache_folder_path: Optional[Path] = None, # 新增参数
                 files_to_open_at_end: Optional[List[Path]] = None # 新增参数
                ):
        """
        初始化HistoryManager。
        Args:
            history_file_path (Path): 记录文件（Excel格式）的完整路径。
            logger_obj: 用于记录日志的对象。
            field_definitions (List[Dict[str, Any]]): 字段定义列表。
                每个字典应包含:
                - "internal_key" (str): 内部数据字典中使用的键名。
                - "excel_header" (str): 在Excel中显示的列头名称。
                - "is_path" (bool, optional): 是否是文件/文件夹路径，默认为False。
                - "hyperlink_display_text" (str, optional): 如果是路径且文件存在，超链接的显示文本。
                - "hyperlink_not_exist_text" (str, optional): 如果是路径但文件不存在，超链接的显示文本。
            sheet_name (str, optional): Excel工作表的名称，默认为"操作记录"。
            cache_folder_path (Optional[Path]): 缓存历史记录Excel文件的目录。如果为None则不进行缓存。
            files_to_open_at_end (Optional[List[Path]]): 引用外部列表，用于存储最终需要自动打开的文件路径。
        """
        self.history_file_path = history_file_path
        self.logger_obj = logger_obj
        self.history_data: List[Dict[str, Any]] = []

        self.field_definitions = field_definitions
        self.sheet_name = sheet_name

        self.internal_to_excel_headers_map: Dict[str, str] = {
            fd["internal_key"]: fd["excel_header"] for fd in field_definitions
        }
        self.excel_to_internal_headers_map: Dict[str, str] = {
            fd["excel_header"]: fd["internal_key"] for fd in field_definitions
        }
        self.path_field_definitions: Dict[str, Dict[str, Any]] = {
            fd["internal_key"]: fd for fd in field_definitions if fd.get("is_path", False)
        }

        # 新增：缓存相关属性
        self.cache_folder_path = cache_folder_path
        self.files_to_open_at_end = files_to_open_at_end if files_to_open_at_end is not None else []

        self._load_history_from_excel()

    def _get_normalized_path_string(self, file_path: Optional[Path]) -> Optional[str]:
        """
        辅助方法：将Path对象转换为规范化的字符串路径，并处理超链接前缀。
        Returns:
            Optional[str]: 规范化的路径字符串，如果file_path为None或文件不存在则返回None。
        """
        if not file_path or not file_path.exists():
            return None
        normalized_path = normalize_drive_letter(str(file_path)).replace("\\", "/")
        if not sys.platform.startswith('win'):
            return f'file://{normalized_path}'
        return normalized_path


    def _create_hyperlink_info(self, file_path: Optional[Path], default_display_text: str, not_exist_text: str) -> Dict[str, Any]:
        """
        辅助方法：生成超链接所需的显示文本和location。
        Returns:
            Dict[str, Any]: 包含 'display_text' 和 'location' 的字典。
        """
        location = self._get_normalized_path_string(file_path)
        if location:
            display_text = default_display_text
        else:
            display_text = not_exist_text
        return {"display_text": display_text, "location": location}

    def _load_history_from_excel(self):
        """
        从Excel文件加载历史记录到内存。
        主要改动：动态读取表头，并使用映射关系将Excel表头转换为内部键名。
        """
        self.history_data = []
        if not self.history_file_path.exists():
            self.logger_obj.info(f"记录文件不存在: {normalize_drive_letter(str(self.history_file_path))}. 将创建新文件。")
            return

        try:
            wb = load_workbook(str(self.history_file_path))
            if self.sheet_name in wb.sheetnames:
                ws = wb[self.sheet_name]
                headers = [cell.value for cell in ws[1]]
                if not headers:
                    self.logger_obj.warning(f"记录文件 '{normalize_drive_letter(str(self.history_file_path))}' 的 '{self.sheet_name}' 工作表为空，无数据记录可加载。")
                    return

                for row_idx in range(2, ws.max_row + 1):
                    row_values = [cell.value for cell in ws[row_idx]]
                    entry = {}
                    for i, header_from_excel in enumerate(headers):
                        if i < len(row_values):
                            internal_key = self.excel_to_internal_headers_map.get(header_from_excel, None)

                            if internal_key is None and header_from_excel.endswith("超链接"):
                                continue
                            if internal_key is None:
                                internal_key = header_from_excel

                            value = row_values[i]
                            if self.path_field_definitions.get(internal_key) and value:
                                try:
                                    entry[internal_key] = Path(value)
                                except Exception as e:
                                    self.logger_obj.warning(f"加载数据记录时，将 '{value}' 转换为路径对象失败，字段: '{internal_key}'。错误: {e}")
                                    entry[internal_key] = value
                            else:
                                entry[internal_key] = value
                        else:
                            internal_key = self.excel_to_internal_headers_map.get(header_from_excel, header_from_excel)
                            entry[internal_key] = None

                    self.history_data.append(entry)
            self.logger_obj.info(f"成功从记录文件加载 {len(self.history_data)} 条数据记录。")
        except FileNotFoundError:
            self.logger_obj.error(f"错误: 记录文件 {normalize_drive_letter(str(self.history_file_path))} 未找到。")
            self.history_data = []
        except PermissionError as e:
            self.logger_obj.error(f"错误: 没有权限读取记录文件 '{normalize_drive_letter(str(self.history_file_path))}'。请检查文件权限。详情: {e}")
            self.history_data = []
        except InvalidFileException as e:
            self.logger_obj.error(f"错误: 记录文件 '{normalize_drive_letter(str(self.history_file_path))}' 格式无效或已损坏。详情: {e}")
            self.history_data = []
        except Exception as e:
            self.logger_obj.error(f"错误: 从记录文件 {normalize_drive_letter(str(self.history_file_path))} 加载数据记录失败: 未知错误: {e}")
            self.history_data = []


    def add_history_entry(self, entry_data: Dict[str, Any]):
        """
        向内存中的数据记录列表添加一条新的操作记录或日志条目。
        Args:
            entry_data (Dict[str, Any]): 包含所有历史记录数据的字典，键名应与 field_definitions 中的 internal_key 对应。
                                          例如：{'timestamp': '...', 'source_path': Path(...), ...}
        """
        if 'timestamp' not in entry_data:
            entry_data['timestamp'] = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        self.history_data.append(entry_data)
        self.logger_obj.info(f"操作记录成功添加至内存: 条目数据: {entry_data.get('timestamp', '未知时间')}")

    def _prepare_excel_for_saving(self) -> Optional[Tuple[Workbook, Any]]:
        """
        辅助方法：删除旧的Excel文件并创建新的Workbook和Sheet，写入表头。
        主要改动：根据 self.field_definitions 动态生成Excel表头。
        Returns:
            Optional[Tuple[Workbook, Worksheet]]: 成功则返回Workbook和Worksheet对象，否则返回None。
        """
        if self.history_file_path.exists():
            try:
                os.remove(str(self.history_file_path))
                self.logger_obj.info(f"已删除旧的记录文件: {normalize_drive_letter(str(self.history_file_path))}")
            except PermissionError as e:
                self.logger_obj.critical(f"致命错误: 无法删除旧的记录文件 '{normalize_drive_letter(str(self.history_file_path))}'，可能文件被占用。请关闭Excel中打开的记录文件。错误详情: {e}")
                return None
            except OSError as e:
                self.logger_obj.critical(f"致命错误: 删除旧的记录文件 '{normalize_drive_letter(str(self.history_file_path))}' 时发生操作系统错误。数据记录将无法保存。错误详情: {e}")
                return None
            except Exception as e:
                self.logger_obj.critical(f"致命错误: 删除旧的记录文件 '{normalize_drive_letter(str(self.history_file_path))}' 时发生未知错误。数据记录将无法保存。错误详情: {e}")
                return None

        wb = Workbook()
        ws = wb.active
        ws.title = self.sheet_name

        excel_headers = []
        for fd in self.field_definitions:
            excel_headers.append(fd["excel_header"])
            if fd.get("is_path", False):
                excel_headers.append(f"{fd['excel_header']}超链接")

        ws.append(excel_headers)
        return wb, ws

    def _write_history_data_to_sheet(self, ws: Any):
        """
        辅助方法：将内存中的数据记录写入到指定的Worksheet，并设置超链接。
        主要改动：根据 self.field_definitions 动态写入数据和设置超链接。
        """
        current_excel_headers = [cell.value for cell in ws[1]]

        for entry in self.history_data:
            row_data = []
            hyperlink_info_to_set = []
            current_col_idx = 1

            for fd in self.field_definitions:
                internal_key = fd["internal_key"]
                excel_header = fd["excel_header"]
                is_path = fd.get("is_path", False)

                value = entry.get(internal_key, "N/A")

                if is_path and isinstance(value, Path):
                    row_data.append(self._get_normalized_path_string(value) if value else "N/A")
                elif isinstance(value, Path):
                    row_data.append(str(value))
                else:
                    row_data.append(value)
                current_col_idx += 1

                if is_path:
                    file_path_obj = entry.get(internal_key)
                    display_text = fd.get("hyperlink_display_text", f"打开{excel_header.replace('绝对路径', '').replace('本地', '')}")
                    not_exist_text = fd.get("hyperlink_not_exist_text", f"{excel_header.replace('绝对路径', '').replace('本地', '')}不存在")

                    link_info = self._create_hyperlink_info(file_path_obj, display_text, not_exist_text)
                    row_data.append(link_info["display_text"])

                    hyperlink_info_to_set.append({
                        "column": current_col_idx,
                        "location": link_info["location"],
                        "display_text": link_info["display_text"],
                        "source_description": f"数据记录文件 ({excel_header} - 行: {ws.max_row + 1})"
                    })
                    current_col_idx += 1

            ws.append(row_data)

            new_row_idx = ws.max_row
            for link_data in hyperlink_info_to_set:
                link_cell = ws.cell(row=new_row_idx, column=link_data["column"])
                set_hyperlink_and_style(
                    link_cell,
                    link_data["location"],
                    link_data["display_text"],
                    self.logger_obj,
                    source_description=link_data["source_description"]
                )

    def _create_cached_snapshot(self, final_history_excel_path: Path) -> None:
        """
        原理：将保存成功的主历史记录Excel文件复制到缓存目录，作为快照。
        实现过程：检查缓存目录是否存在，生成带时间戳的文件名，然后执行复制操作。
        主要改动点：这是原 _handle_history_caching 函数的核心逻辑，现在作为类方法。
        """
        if self.cache_folder_path:
            if create_directory_if_not_exists(self.cache_folder_path, self.logger_obj):
                current_timestamp_for_cache = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
                # 保持文件名一致性，使用scan_history_cached_前缀，因为这是扫描项目的历史记录
                cached_history_file_name = f"scan_history_cached_{current_timestamp_for_cache}.xlsx"
                cached_history_file_path = self.cache_folder_path / cached_history_file_name

                self.logger_obj.info(f"开始复制历史记录到缓存文件夹: {normalize_drive_letter(str(cached_history_file_path))}")
                copy_cache_success = copy_file(final_history_excel_path, cached_history_file_path, self.logger_obj)

                if copy_cache_success:
                    self.logger_obj.info("历史记录已成功复制到缓存文件夹。")
                    if self.files_to_open_at_end is not None:
                        self.files_to_open_at_end.append(cached_history_file_path)
                else:
                    self.logger_obj.error("历史记录复制到缓存文件夹失败。")
            else:
                self.logger_obj.error(f"无法创建缓存文件夹: {normalize_drive_letter(str(self.cache_folder_path))}，将无法复制历史记录。")
        else:
            self.logger_obj.info("未配置缓存文件夹路径，跳过历史记录缓存。")


    def save_history_to_excel(self) -> bool:
        """
        将内存中的所有数据记录保存到Excel文件。
        Returns:
            bool: 如果保存成功返回True，否则返回False。
        """
        self.logger_obj.info(f"开始将内存中的数据记录保存到Excel: {normalize_drive_letter(str(self.history_file_path))}")

        excel_preparation = self._prepare_excel_for_saving()
        if excel_preparation is None:
            return False

        wb, ws = excel_preparation

        try:
            self._write_history_data_to_sheet(ws)

            # 设置所有列宽
            set_fixed_column_widths(ws, FIXED_COLUMN_WIDTH, self.logger_obj)

            wb.save(str(self.history_file_path))
            self.logger_obj.info(f"成功将数据记录保存到Excel: {normalize_drive_letter(str(self.history_file_path))}")

            # 调用内部缓存方法
            self._create_cached_snapshot(self.history_file_path) # 成功保存后，立即生成缓存快照

            return True
        except PermissionError as e:
            self.logger_obj.error(f"错误: 没有权限写入记录文件 '{normalize_drive_letter(str(self.history_file_path))}'，或文件被占用。请关闭文件。详情: {e}")
            return False
        except Exception as e:
            self.logger_obj.error(f"错误: 将数据记录保存到Excel失败 {normalize_drive_letter(str(self.history_file_path))}: 未知错误: {e}")
            return False